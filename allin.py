from __future__ import annotations
def cg_guess_id_from_symbol(symbol: str) -> Optional[str]:
    s = symbol.upper()
    static = {
        "BTC": "bitcoin", "ETH": "ethereum", "ATOM": "cosmos", "DYDX": "dydx-chain",
        "BABY": "babylon", "DYM": "dymension", "AURA": "aura-network", "ZETA": "zetachain",
        "SOL": "solana", "BNB": "binancecoin", "TON": "the-open-network",
        "USDT": "tether", "USDC": "usd-coin"
    }
    return static.get(s)
# --- NHẬP NHANH CÂU TỰ NHIÊN ---
VIET_NUM = {
    "k": 1000, "nghin": 1000, "nghìn": 1000, "ngan": 1000, "ngàn": 1000,
    "tr": 1_000_000, "trieu": 1_000_000, "triệu": 1_000_000,
    "củ": 1_000_000, "cu": 1_000_000
}

def parse_vietnamese_amount(text):
    """
    Chuyển các chuỗi tiền tệ tiếng Việt sang số nguyên (VND).
    Ví dụ:
        '1tr5' -> 1_500_000
        '2 củ' -> 2_000_000
        '2k1' -> 2_100
    """
    text = text.strip().lower().replace(",", "").replace(" ", "")

    import re
    # Dạng '1tr5' hoặc '2củ3' => triệu + trăm nghìn
    m = re.match(r"^(\d+)(tr|trieu|triệu|củ|cu)(\d+)$", text)
    if m:
        tr_part = int(m.group(1)) * 1_000_000
        sub_part = int(m.group(3)) * 100_000
        return tr_part + sub_part

    # Dạng '2k1' => nghìn + trăm
    m = re.match(r"^(\d+)(k|nghin|nghìn|ngan|ngàn)(\d+)$", text)
    if m:
        k_part = int(m.group(1)) * 1_000
        sub_part = int(m.group(3)) * 100
        return k_part + sub_part

    # Dạng '2tr', '3k', '4 củ'
    for key, multiplier in VIET_NUM.items():
        if text.endswith(key):
            num = text[:-len(key)]
            if num.isdigit():
                return int(num) * multiplier

    # Chỉ là số
    if text.isdigit():
        return int(text)

    return None

CATEGORY_KEYWORDS = {
    "Ăn uống": [
        "ăn", "ăn sáng", "bữa sáng", "sáng", "trưa", "chiều", "tối",
        "cafe", "cà phê", "trà sữa", "quán", "nhậu", "uống", "đi ăn", "trà đá", 
        "buffet", "liên hoan", "bánh", "kem", "đồ uống", "thức ăn", "đặt món"
    ],
    "Di chuyển": [
        "taxi", "grab", "xăng", "vé xe", "tàu", "vé tàu", "xe buýt", "bus", "gửi xe", "bến",
        "đi lại", "đi xe", "xe khách", "máy bay", "vé máy bay", "xích lô", "ô tô", "đi grab", "ship"
    ],
    "Mua sắm": [
        "mua", "đồ", "áo", "quần", "giày", "shopee", "lazada", "tiki", "siêu thị",
        "mua hàng", "đặt hàng", "order", "đi chợ", "mỹ phẩm", "trang sức", "phụ kiện"
    ],
    "Hóa đơn": [
        "điện", "nước", "wifi", "internet", "cước", "hóa đơn", "tiền nhà",
        "truyền hình", "gas", "rác thải", "điện thoại", "sim", "4g", "5g"
    ],
    "Gia đình": [
        "gia đình", "bà ngoại", "bố mẹ", "con", "vợ", "chồng",
        "ông bà", "anh chị", "em", "gia tộc", "cháu"
    ],
    "Sức khỏe": [
        "khám", "thuốc", "bệnh viện", "y tế",
        "khám bệnh", "xét nghiệm", "phẫu thuật", "tiêm", "bảo hiểm y tế", "bảo hiểm sức khỏe"
    ],
    "Giải trí": [
        "xem phim", "anime", "game", "giải trí",
        "nhạc", "hát", "karaoke", "du lịch", "chơi", "coi phim", "xem ca nhạc", "thể thao"
    ],
    "Đầu tư": [
        "đầu tư", "crypto", "coin", "chứng khoán",
        "cổ phiếu", "trái phiếu", "mua coin", "mua cổ", "vàng", "bất động sản"
    ],
    "Tiết kiệm": [
        "tiết kiệm", "gửi tiết kiệm",
        "mở sổ", "sổ tiết kiệm", "để dành", "dành dụm"
    ],
    "Khác": [
        "từ thiện", "quà", "ủng hộ", "học phí", "đào tạo", "khóa học", "mua tên miền", "dịch vụ"
    ]
}

INCOME_CATEGORY_KEYWORDS = {
    "Lương": [
        "lương", "tiền công", "tiền lương", "lương tháng", "salary",
        "lương cơ bản", "lương chính", "lương thời vụ", "lương ngày",
        "lương tuần", "phụ cấp", "trợ cấp", "tiền công nhật"
    ],
    "Thưởng": [
        "thưởng", "bonus", "thưởng tết", "thưởng quý", "thưởng năm",
        "thưởng doanh số", "thưởng năng suất", "thưởng thêm"
    ],
    "Kinh doanh": [
        "kinh doanh", "bán hàng", "doanh thu", "hoa hồng", "chốt đơn",
        "bán online", "bán trực tiếp", "lợi nhuận shop", "thu bán hàng",
        "thu từ bán", "thu cửa hàng", "thu quán"
    ],
    "Được cho": [
        "được cho", "được tặng", "cho tiền", "mừng tuổi", "mừng cưới",
        "tiền mừng", "quà", "cho", "biếu", "tặng", "hỗ trợ",
        "cha mẹ cho", "người thân cho", "bạn bè cho"
    ],
    "Thu hồi nợ": [
        "trả nợ", "thu hồi nợ", "được trả", "trả lại", "người nợ trả",
        "hoàn tiền", "hoàn vốn", "nhận lại tiền"
    ],
    "Bán tài sản": [
        "bán tài sản", "bán xe", "bán nhà", "thanh lý", "bán đất",
        "bán đồ", "thanh lý hàng", "bán máy", "bán điện thoại",
        "bán vàng", "bán máy tính", "bán laptop", "bán đồ cũ"
    ],
    "Lãi tiết kiệm": [
        "lãi tiết kiệm", "tiền gửi", "gửi tiết kiệm", "lãi ngân hàng",
        "lãi sổ tiết kiệm", "tiền lãi tiết kiệm"
    ],
    "Lãi đầu tư": [
        "lãi đầu tư", "cổ tức", "trái tức", "lợi nhuận", "lãi cổ phiếu",
        "lãi coin", "lãi crypto", "lãi chứng khoán", "lãi quỹ",
        "lợi nhuận đầu tư", "lãi trái phiếu", "lãi bất động sản", "Bán Atom", "Atom", "P2P", "Crypto"
    ],
    "Khác": [
        "thu nhập khác", "nguồn khác", "không xác định", "thu lặt vặt"
    ]
}

def guess_income_source(text: str) -> str:
    low = text.lower()
    for src, kws in INCOME_CATEGORY_KEYWORDS.items():
        for kw in kws:
            if kw in low:
                return src
    return "Khác"

# Tạo list từ khóa thu nhập từ dict INCOME_CATEGORY_KEYWORDS
INCOME_KEYWORDS = [kw for kws in INCOME_CATEGORY_KEYWORDS.values() for kw in kws]
DATE_KEYWORDS = {
    "hôm nay": 0, "hom nay": 0, "nay": 0,
    "hôm qua": -1, "hom qua": -1, "qua": -1,
    "hôm kia": -2, "hom kia": -2
}

import re as _re2

# -*- coding: utf-8 -*-
"""
Univer Crypto Portfolio Bot
- Telegram bot (python-telegram-bot v21 async)
- Portfolio in USD using CoinGecko
- Donut (allocation by value) + Table image on /cp and 📋 Danh mục
- Import/Export CSV hoặc Excel with clear success/error counts
"""

import os
import io
import csv
import sqlite3
import datetime as _dt
import tempfile
import asyncio
import traceback
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch, Circle
from matplotlib.colors import LinearSegmentedColormap
import numpy as np
from typing import List, Dict, Tuple, Optional
import requests
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None
from urllib.parse import quote

from telegram import (
    Update, InlineKeyboardMarkup, InlineKeyboardButton, InputFile
)
from telegram.ext import (
    Application, ApplicationBuilder, CommandHandler, ContextTypes,
    CallbackQueryHandler, MessageHandler, filters
)

# Charts
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import gridspec

# ===================== CONFIG =====================
BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "8035366092:AAGe1lnDSnSSKueP9fSNuBG7di0n2mOebms")  # NOTE: hardcoded fallback for testing
DB_PATH = os.environ.get("UNIVER_DB_PATH", "univer_all_in_one.db")
CRYPTO_DB_PATH = DB_PATH  # merge: use one SQLite file for all modules
COINGECKO_BASE = "https://api.coingecko.com/api/v3"

# --- lazy migration guard for crypto DB ---
# ===================== DB =====================


def run_migrations():
    conn = sqlite3.connect(CRYPTO_DB_PATH)
    cur = conn.cursor()
    cur.execute("PRAGMA user_version")
    ver_row = cur.fetchone()
    ver = ver_row[0] if ver_row and ver_row[0] is not None else 0

    # Create required tables (idempotent)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS crypto_trades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            symbol TEXT NOT NULL,
            cg_id TEXT,
            side TEXT NOT NULL,
            qty REAL NOT NULL,
            price_usd REAL NOT NULL,
            fee_usd REAL DEFAULT 0,
            note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS crypto_map (
            symbol TEXT PRIMARY KEY,
            cg_id TEXT NOT NULL,
            name TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS crypto_prices (
            cg_id TEXT NOT NULL,
            price_usd REAL NOT NULL,
            fetched_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Bump version at least to 1
    if ver < 1:
        cur.execute("PRAGMA user_version = 1")

    conn.commit()
    conn.close()


_MIGRATIONS_DONE = False
def ensure_migrated():
    global _MIGRATIONS_DONE
    if not _MIGRATIONS_DONE:
        run_migrations()
        _MIGRATIONS_DONE = True


def db_conn():
    ensure_migrated()
    return sqlite3.connect(CRYPTO_DB_PATH)

def db_crypto_upsert_map(symbol: str, cg_id: str, name: Optional[str] = None):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO crypto_map(symbol, cg_id, name)
        VALUES (?, ?, ?)
        ON CONFLICT(symbol) DO UPDATE SET cg_id=excluded.cg_id, name=COALESCE(excluded.name, crypto_map.name)
    """, (symbol.upper(), cg_id, name))
    conn.commit(); conn.close()

def db_crypto_get_map(symbol: str) -> Optional[str]:
    conn = db_conn(); cur = conn.cursor()
    cur.execute("SELECT cg_id FROM crypto_map WHERE symbol = ?", (symbol.upper(),))
    row = cur.fetchone(); conn.close()
    return row[0] if row else None

def db_crypto_add_trade(user_id: int, symbol: str, side: str, qty: float, price_usd: float,
                        note: str = "", created_at: Optional[_dt.datetime] = None,
                        cg_id: Optional[str] = None, fee_usd: float = 0.0):
    conn = db_conn(); cur = conn.cursor()
    if created_at is None:
        created_at = _dt.datetime.now()
    cur.execute("""
        INSERT INTO crypto_trades (user_id, symbol, cg_id, side, qty, price_usd, fee_usd, note, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (user_id, symbol.upper(), cg_id, side.upper(), qty, price_usd, fee_usd, note, created_at.isoformat()))
    conn.commit(); conn.close()

def db_crypto_positions(user_id: int):
    """Aggregate net qty and invested USD per symbol for a user; attach a reasonable cg_id per symbol."""
    ensure_migrated()
    conn = db_conn(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT symbol,
                   SUM(CASE WHEN side='BUY' THEN qty ELSE -qty END) AS qty_net,
                   SUM(CASE WHEN side='BUY' THEN qty*price_usd + fee_usd ELSE -(qty*price_usd) END) AS invested_usd
            FROM crypto_trades
            WHERE user_id = ?
            GROUP BY symbol
            HAVING ABS(qty_net) > 1e-12
            ORDER BY symbol
        """, (user_id,))
        rows = cur.fetchall()
    except sqlite3.OperationalError:
        # If tables are missing for some reason, migrate and retry once
        run_migrations()
        cur.execute("""
            SELECT symbol,
                   SUM(CASE WHEN side='BUY' THEN qty ELSE -qty END) AS qty_net,
                   SUM(CASE WHEN side='BUY' THEN qty*price_usd + fee_usd ELSE -(qty*price_usd) END) AS invested_usd
            FROM crypto_trades
            WHERE user_id = ?
            GROUP BY symbol
            HAVING ABS(qty_net) > 1e-12
            ORDER BY symbol
        """, (user_id,))
        rows = cur.fetchall()

    result = []
    for symbol, qty_net, invested_usd in rows:
        # Prefer the latest non-null cg_id in trades; otherwise fallback to mapping table; otherwise guess
        cur.execute("""
            SELECT cg_id
            FROM crypto_trades
            WHERE user_id = ? AND symbol = ? AND cg_id IS NOT NULL
            ORDER BY created_at DESC
            LIMIT 1
        """, (user_id, symbol))
        row = cur.fetchone()
        cg_id = (row[0] if row and row[0] else None) or db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
        result.append({
            "symbol": symbol,
            "qty": float(qty_net or 0.0),
            "invested_usd": float(invested_usd or 0.0),
            "cg_id": cg_id
        })
    conn.close()
    return result


def db_crypto_all_trades(user_id: int):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT symbol, COALESCE(cg_id,''), side, qty, price_usd, fee_usd, COALESCE(note,''), created_at
        FROM crypto_trades
        WHERE user_id = ?
        ORDER BY created_at ASC
    """, (user_id,))
    rows = cur.fetchall(); conn.close()
    return rows

# ===================== COINGECKO =====================
def cg_guess_id_from_symbol(symbol: str) -> Optional[str]:
    s = symbol.upper()
    static = {
        "BTC": "bitcoin", "ETH": "ethereum", "ATOM": "cosmos", "DYDX": "dydx-chain",
        "BABY": "babylon", "DYM": "dymension", "AURA": "aura-network", "ZETA": "zetachain",
        "SOL": "solana", "BNB": "binancecoin", "TON": "the-open-network",
        "USDT": "tether", "USDC": "usd-coin"
    }
    return static.get(s)

def cg_simple_price_usd(cg_ids: List[str]) -> Dict[str, float]:
    if not cg_ids: return {}
    ids_param = ",".join(sorted(set(cg_ids)))
    url = f"{COINGECKO_BASE}/simple/price?ids={quote(ids_param)}&vs_currencies=usd"
    try:
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            data = r.json()
            return {k: float(v.get("usd", 0.0) or 0.0) for k, v in data.items()}
    except Exception:
        pass
    return {}

# ===================== HISTORY SERIES (for performance chart) =====================
def cg_market_chart_usd(cg_id: str, days: int = 365):
    """Return dict {label: price}. Daily for days>1, hourly for days<=1."""
    interval = "hourly" if days <= 1 else "daily"
    url = f"{COINGECKO_BASE}/coins/{quote(cg_id)}/market_chart?vs_currency=usd&days={days}&interval={interval}"
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200:
            return {}
        data = r.json()
        out = {}
        for ts, price in data.get("prices", []):
            dt = _dt.datetime.utcfromtimestamp(ts/1000.0)
            label = dt.strftime('%H:%M') if days <= 1 else dt.date().isoformat()
            out[label] = float(price)
        return out
    except Exception:
        return {}

def portfolio_history_series(user_id: int, days: int = 365):
    """Compute portfolio value series for last `days` using users' net quantities and CoinGecko prices."""
    rows = db_crypto_all_trades(user_id)
    if not rows:
        return [], []

    # Aggregate trades per symbol and get cg_ids
    trades_by_sym = {}
    cg_ids = {}
    for symbol, cg_id, side, qty, price, fee, note, created_at in rows:
        sym = symbol.upper()
        trades_by_sym.setdefault(sym, []).append((side, float(qty), created_at))
        if cg_id: cg_ids[sym] = cg_id

    # Fill cg_id gaps with mapping/static guesses
    for sym in trades_by_sym:
        if sym not in cg_ids:
            guess = db_crypto_get_map(sym) or cg_guess_id_from_symbol(sym)
            if guess: cg_ids[sym] = guess

    if not cg_ids:
        return [], []

    # Fetch price series for each symbol
    price_hist = {}
    all_labels = set()
    for sym, cg in cg_ids.items():
        ph = cg_market_chart_usd(cg, days=days)
        if not ph: 
            continue
        price_hist[sym] = ph
        all_labels.update(ph.keys())

    if not price_hist:
        return [], []

    labels = sorted(list(all_labels))

    # Normalize trade timestamps to labels (date component)
    norm_trades_by_sym = {}
    for sym, arr in trades_by_sym.items():
        tmp = []
        for side, qty, at in arr:
            try:
                d = _dt.datetime.fromisoformat(at).date().isoformat()
            except Exception:
                try:
                    d = str(at).split(" ")[0]
                except Exception:
                    d = labels[0]
            tmp.append((d, side, qty))
        tmp.sort(key=lambda x: x[0])
        norm_trades_by_sym[sym] = tmp

    # Rolling net qty per label
    qty_by_sym_by_label = {sym: {} for sym in trades_by_sym}
    running_qty = {sym: 0.0 for sym in trades_by_sym}
    for lb in labels:
        for sym in trades_by_sym:
            for td, side, qty in [t for t in norm_trades_by_sym[sym] if t[0] == lb]:
                running_qty[sym] += qty if side == "BUY" else -qty
            qty_by_sym_by_label[sym][lb] = running_qty[sym]

    # Compute portfolio value per label
    values = []
    for lb in labels:
        total = 0.0
        for sym in trades_by_sym:
            price = price_hist.get(sym, {}).get(lb, None)
            if price is None: 
                continue
            total += qty_by_sym_by_label[sym].get(lb, 0.0) * price
        values.append(total)

    return labels, values

# ===================== CHART (Donut + Table) =====================
def gen_portfolio_donut_image(positions, prices) -> str:
    """
    Generate a modern, professional portfolio donut chart with advanced styling
    """
    # Modern color palette
    COLORS = {
        'bg': '#0a0e27',
        'card_bg': '#151932',
        'text_primary': '#ffffff',
        'text_secondary': '#94a3b8',
        'accent': '#00d4ff',
        'chart_colors': [
            '#00d4ff', '#00ff88', '#ffd700', '#ff6b6b', '#c56cf0',
            '#4834d4', '#22a6b3', '#f0932b', '#eb4d4b', '#6ab04c',
            '#30336b', '#95afc0', '#535c68', '#ff9ff3', '#48dbfb'
        ],
        'glow': '#00d4ff'
    }
    
    # Prepare data
    tokens_data = []
    total_value = 0
    
    for p in positions:
        sym = p.get("symbol", "").upper()
        qty = float(p.get("qty", 0) or 0)
        cg = p.get("cg_id") or cg_guess_id_from_symbol(sym) or ""
        price_now = float(prices.get(cg, 0.0) if cg else 0.0)
        value = qty * price_now
        
        if value > 0:  # Only include tokens with positive value
            tokens_data.append({
                'symbol': sym or "N/A",
                'value': value,
                'qty': qty,
                'price': price_now
            })
            total_value += value
    
    # Sort by value descending
    tokens_data.sort(key=lambda x: x['value'], reverse=True)
    
    # If no data, create dummy data
    if not tokens_data:
        tokens_data = [{'symbol': 'N/A', 'value': 1.0, 'qty': 0, 'price': 0}]
        total_value = 1.0
    
    # Limit to top 10 tokens and group others
    if len(tokens_data) > 10:
        top_tokens = tokens_data[:10]
        others_value = sum(t['value'] for t in tokens_data[10:])
        if others_value > 0:
            top_tokens.append({
                'symbol': 'Others',
                'value': others_value,
                'qty': 0,
                'price': 0
            })
        tokens_data = top_tokens
    
    # Calculate percentages
    for token in tokens_data:
        token['percentage'] = (token['value'] / total_value * 100) if total_value > 0 else 0
    
    # Create figure with dark background
    fig = plt.figure(figsize=(16, 10), facecolor=COLORS['bg'])
    
    # Create main plot area
    ax = fig.add_subplot(121, facecolor=COLORS['bg'])
    ax.set_aspect('equal')
    
    # Prepare pie chart data
    sizes = [token['value'] for token in tokens_data]
    labels = [token['symbol'] for token in tokens_data]
    colors = COLORS['chart_colors'][:len(tokens_data)]
    
    # Create donut chart with explosion effect for top 3
    explode = [0.03 if i < 3 else 0 for i in range(len(sizes))]
    
    # Draw outer ring (shadow effect)
    shadow_wedges, _ = ax.pie(sizes, 
                              radius=1.15,
                              colors=['#000000'] * len(sizes),
                              wedgeprops=dict(width=0.45, edgecolor='none', alpha=0.3),
                              startangle=90)
    
    # Draw main donut
    wedges, texts, autotexts = ax.pie(sizes,
                                       labels=None,  # We'll add custom labels
                                       autopct='',   # We'll add custom text
                                       startangle=90,
                                       colors=colors,
                                       explode=explode,
                                       wedgeprops=dict(width=0.4, 
                                                      edgecolor=COLORS['bg'],
                                                      linewidth=2),
                                       pctdistance=0.8)
    
    # Add glow effect to wedges
    for i, wedge in enumerate(wedges):
        wedge.set_alpha(0.9)
        if i < 3:  # Add extra glow to top 3
            wedge.set_linewidth(3)
            wedge.set_edgecolor(colors[i])
            wedge.set_alpha(1.0)
    
    # Add center circle with total value
    centre_circle = Circle((0, 0), 0.55, fc=COLORS['card_bg'], linewidth=3, 
                          edgecolor=COLORS['accent'], alpha=0.95)
    ax.add_artist(centre_circle)
    
    # Add center text
    ax.text(0, 0.1, '💰 TỔNG GIÁ TRỊ', ha='center', va='center',
            fontsize=12, color=COLORS['text_secondary'], weight='bold')
    ax.text(0, -0.05, f'${total_value:,.2f}', ha='center', va='center',
            fontsize=20, color=COLORS['accent'], weight='bold')
    
    # Add percentage labels on the donut
    for i, (wedge, token) in enumerate(zip(wedges, tokens_data)):
        ang = (wedge.theta2 - wedge.theta1) / 2. + wedge.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        
        # Position for percentage text
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        
        if token['percentage'] > 3:  # Only show percentage if > 3%
            ax.annotate(f"{token['percentage']:.1f}%",
                       xy=(x*0.8, y*0.8),
                       xytext=(x*0.8, y*0.8),
                       ha='center', va='center',
                       fontsize=10,
                       color='white',
                       weight='bold')
    
    # Create legend area (right side)
    ax2 = fig.add_subplot(122, facecolor=COLORS['bg'])
    ax2.axis('off')
    
    # Title for legend area
    ax2.text(0.5, 0.95, '📊 CHI TIẾT DANH MỤC', 
            transform=ax2.transAxes,
            ha='center', fontsize=16, weight='bold', 
            color=COLORS['accent'])
    
    # Create detailed legend with cards
    y_position = 0.85
    for i, token in enumerate(tokens_data):
        if i >= 12:  # Limit legend items
            break
            
        # Create card background
        card = FancyBboxPatch((0.05, y_position - 0.06), 0.9, 0.055,
                              boxstyle="round,pad=0.01",
                              facecolor=COLORS['card_bg'],
                              edgecolor=colors[i],
                              linewidth=2,
                              alpha=0.8,
                              transform=ax2.transAxes)
        ax2.add_patch(card)
        
        # Rank badge
        rank_text = f"#{i+1}" if i < 9 else "Other"
        ax2.text(0.1, y_position - 0.03, rank_text,
                transform=ax2.transAxes,
                fontsize=10, weight='bold',
                color=colors[i], va='center')
        
        # Token symbol
        ax2.text(0.2, y_position - 0.02, token['symbol'],
                transform=ax2.transAxes,
                fontsize=12, weight='bold',
                color=COLORS['text_primary'], va='center')
        
        # Value
        ax2.text(0.5, y_position - 0.02, f"${token['value']:,.2f}",
                transform=ax2.transAxes,
                fontsize=11,
                color=COLORS['text_primary'], va='center')
        
        # Percentage with bar
        pct = token['percentage']
        ax2.text(0.75, y_position - 0.02, f"{pct:.1f}%",
                transform=ax2.transAxes,
                fontsize=11, weight='bold',
                color=colors[i], va='center')
        
        # Mini progress bar
        bar_bg = FancyBboxPatch((0.82, y_position - 0.035), 0.12, 0.01,
                                boxstyle="round,pad=0",
                                facecolor=COLORS['card_bg'],
                                edgecolor=COLORS['text_secondary'],
                                linewidth=1,
                                alpha=0.5,
                                transform=ax2.transAxes)
        ax2.add_patch(bar_bg)
        
        bar_fill = FancyBboxPatch((0.82, y_position - 0.035), 
                                  0.12 * (pct/100), 0.01,
                                  boxstyle="round,pad=0",
                                  facecolor=colors[i],
                                  edgecolor='none',
                                  alpha=0.8,
                                  transform=ax2.transAxes)
        ax2.add_patch(bar_fill)
        
        y_position -= 0.07
    
    # Add statistics box at bottom
    stats_y = 0.15
    stats_box = FancyBboxPatch((0.05, stats_y - 0.12), 0.9, 0.11,
                               boxstyle="round,pad=0.02",
                               facecolor=COLORS['card_bg'],
                               edgecolor=COLORS['accent'],
                               linewidth=2,
                               alpha=0.9,
                               transform=ax2.transAxes)
    ax2.add_patch(stats_box)
    
    # Statistics content
    num_tokens = len([t for t in tokens_data if t['symbol'] != 'Others'])
    avg_holding = total_value / max(num_tokens, 1)
    max_holding = max(tokens_data, key=lambda x: x['value'])
    
    ax2.text(0.5, stats_y - 0.02, '📈 THỐNG KÊ',
            transform=ax2.transAxes,
            ha='center', fontsize=11, weight='bold',
            color=COLORS['accent'])
    
    stats_text = f"Số token: {num_tokens} | TB/token: ${avg_holding:,.2f}\n"
    stats_text += f"Lớn nhất: {max_holding['symbol']} (${max_holding['value']:,.2f})"
    
    ax2.text(0.5, stats_y - 0.07, stats_text,
            transform=ax2.transAxes,
            ha='center', fontsize=9,
            color=COLORS['text_secondary'],
            linespacing=1.5)
    
    # Add main title
    fig.suptitle('🎯 PHÂN BỔ DANH MỤC ĐẦU TƯ CRYPTO', 
                fontsize=20, weight='bold', 
                color=COLORS['accent'], y=0.98)
    
    # Add subtitle with timestamp
    timestamp = _dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    fig.text(0.5, 0.94, f'Cập nhật: {timestamp}',
            ha='center', fontsize=10,
            color=COLORS['text_secondary'], style='italic')
    
    # Add decorative elements
    # Top gradient line
    gradient_line = plt.Line2D([0.1, 0.9], [0.92, 0.92],
                              linewidth=2, color=COLORS['accent'],
                              alpha=0.5, transform=fig.transFigure)
    fig.add_artist(gradient_line)
    
    # Bottom gradient line
    gradient_line2 = plt.Line2D([0.1, 0.9], [0.02, 0.02],
                               linewidth=2, color=COLORS['accent'],
                               alpha=0.5, transform=fig.transFigure)
    fig.add_artist(gradient_line2)
    
    # Adjust layout
    plt.subplots_adjust(left=0.05, right=0.95, top=0.90, bottom=0.05)
    
    # Save with high quality
    fd, path = tempfile.mkstemp(prefix="portfolio_donut_modern_", suffix=".png")
    os.close(fd)
    plt.savefig(path, dpi=200, bbox_inches='tight',
                facecolor=COLORS['bg'], edgecolor='none')
    plt.close(fig)
    
    return path

def gen_portfolio_table_image(positions, prices) -> str:
    """
    Generate a modern, professional portfolio table image
    """
    # Prepare data
    rows = []
    total_inv = total_val = 0.0
    
    for p in positions:
        sym = p.get("symbol", "")
        qty = float(p.get("qty", 0) or 0)
        inv = float(max(p.get("invested_usd", 0) or 0, 0))
        cg = p.get("cg_id") or cg_guess_id_from_symbol(sym) or ""
        price_now = float(prices.get(cg, 0.0) if cg else 0.0)
        value = qty * price_now
        pnl = value - inv
        pct = (pnl/inv*100) if inv > 1e-12 else 0.0
        
        rows.append({
            'symbol': sym.upper(),
            'qty': f"{qty:,.4f}".rstrip('0').rstrip('.'),
            'price': f"${price_now:,.2f}",
            'value': f"${value:,.2f}",
            'invested': f"${inv:,.2f}",
            'pnl': f"${pnl:+,.2f}",
            'pct': pct,
            'pct_str': f"{pct:+.2f}%"
        })
        
        total_inv += inv
        total_val += value
    
    total_pnl = total_val - total_inv
    total_pct = (total_pnl/total_inv*100) if total_inv > 1e-12 else 0.0
    
    # Modern color scheme
    COLORS = {
        'bg_dark': '#1a1a2e',
        'bg_header': '#16213e',
        'bg_row1': '#0f3460',
        'bg_row2': '#0e2954',
        'bg_total': '#16213e',
        'text_light': '#e8e8e8',
        'text_dim': '#a8a8a8',
        'accent': '#00d4ff',
        'positive': '#00ff88',
        'negative': '#ff4757',
        'neutral': '#ffd700',
        'border': '#2a2a3e'
    }
    
    # Setup figure with dark background
    fig = plt.figure(figsize=(16, max(6, 2 + 0.5 * len(rows))), facecolor=COLORS['bg_dark'])
    ax = fig.add_subplot(111, facecolor=COLORS['bg_dark'])
    ax.axis('off')
    
    # Headers with icons (using Unicode symbols)
    headers = [
        "🪙 Token",
        "📊 Số lượng",
        "💵 Giá hiện tại",
        "💰 Giá trị",
        "📈 Vốn đầu tư",
        "📊 Lãi/Lỗ",
        "📈 %"
    ]
    
    # Prepare table data
    table_data = [headers]
    
    if rows:
        for row in rows:
            table_data.append([
                row['symbol'],
                row['qty'],
                row['price'],
                row['value'],
                row['invested'],
                row['pnl'],
                row['pct_str']
            ])
    else:
        table_data.append(["N/A", "0", "$0.00", "$0.00", "$0.00", "$0.00", "+0.00%"])
    
    # Total row
    table_data.append([
        "📊 TỔNG CỘNG",
        "",
        "",
        f"${total_val:,.2f}",
        f"${total_inv:,.2f}",
        f"${total_pnl:+,.2f}",
        f"{total_pct:+.2f}%"
    ])
    
    # Create table with custom styling
    table = ax.table(cellText=table_data, loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1, 2.2)
    
    # Style the table
    for i, row in enumerate(table_data):
        for j, cell in enumerate(row):
            cell_obj = table[(i, j)]
            
            # Header row styling
            if i == 0:
                cell_obj.set_facecolor(COLORS['bg_header'])
                cell_obj.set_text_props(weight='bold', color=COLORS['accent'])
                cell_obj.set_height(0.08)
            # Total row styling
            elif i == len(table_data) - 1:
                cell_obj.set_facecolor(COLORS['bg_total'])
                cell_obj.set_text_props(weight='bold', color=COLORS['accent'])
                cell_obj.set_height(0.08)
                # Color code total P/L
                if j == 5 or j == 6:
                    if total_pnl > 0:
                        cell_obj.set_text_props(color=COLORS['positive'], weight='bold')
                    elif total_pnl < 0:
                        cell_obj.set_text_props(color=COLORS['negative'], weight='bold')
                    else:
                        cell_obj.set_text_props(color=COLORS['neutral'], weight='bold')
            # Data rows styling
            else:
                # Alternating row colors
                if i % 2 == 0:
                    cell_obj.set_facecolor(COLORS['bg_row1'])
                else:
                    cell_obj.set_facecolor(COLORS['bg_row2'])
                
                cell_obj.set_text_props(color=COLORS['text_light'])
                
                # Special styling for specific columns
                if j == 0:  # Token symbol
                    cell_obj.set_text_props(weight='bold', color=COLORS['text_light'])
                elif j == 5 or j == 6:  # P/L columns
                    if i - 1 < len(rows):
                        pct_val = rows[i-1]['pct']
                        if pct_val > 0:
                            cell_obj.set_text_props(color=COLORS['positive'], weight='bold')
                        elif pct_val < 0:
                            cell_obj.set_text_props(color=COLORS['negative'], weight='bold')
                        else:
                            cell_obj.set_text_props(color=COLORS['neutral'])
            
            # Set border
            cell_obj.set_edgecolor(COLORS['border'])
            cell_obj.set_linewidth(0.5)
    
    # Add title with custom styling
    title_text = "💼 DANH MỤC ĐẦU TƯ CRYPTO"
    subtitle_text = f"Tổng giá trị: ${total_val:,.2f} | Lãi/Lỗ: ${total_pnl:+,.2f} ({total_pct:+.2f}%)"
    
    # Main title
    ax.text(0.5, 1.08, title_text, 
            transform=ax.transAxes, 
            ha='center', 
            fontsize=18, 
            weight='bold', 
            color=COLORS['accent'])
    
    # Subtitle with color coding
    subtitle_color = COLORS['positive'] if total_pnl > 0 else COLORS['negative'] if total_pnl < 0 else COLORS['neutral']
    ax.text(0.5, 1.03, subtitle_text,
            transform=ax.transAxes,
            ha='center',
            fontsize=12,
            color=subtitle_color,
            style='italic')
    
    # Add performance indicator bar
    if total_inv > 0:
        # Create a small performance bar at the bottom
        bar_ax = fig.add_axes([0.1, 0.02, 0.8, 0.02])
        bar_ax.set_xlim(0, 1)
        bar_ax.set_ylim(0, 1)
        bar_ax.axis('off')
        
        # Background bar
        bg_bar = mpatches.Rectangle((0, 0), 1, 1, 
                                   facecolor=COLORS['bg_header'], 
                                   edgecolor=COLORS['border'],
                                   linewidth=1)
        bar_ax.add_patch(bg_bar)
        
        # Performance bar
        perf_ratio = min(abs(total_pct) / 100, 1.0)  # Cap at 100%
        bar_color = COLORS['positive'] if total_pnl > 0 else COLORS['negative']
        
        if total_pnl >= 0:
            perf_bar = mpatches.Rectangle((0.5, 0), perf_ratio * 0.5, 1, 
                                         facecolor=bar_color, alpha=0.7)
        else:
            perf_bar = mpatches.Rectangle((0.5 - perf_ratio * 0.5, 0), perf_ratio * 0.5, 1,
                                         facecolor=bar_color, alpha=0.7)
        bar_ax.add_patch(perf_bar)
        
        # Center line
        bar_ax.axvline(x=0.5, color=COLORS['accent'], linewidth=2, alpha=0.8)
        
        # Labels
        bar_ax.text(0, -0.5, '-100%', ha='left', va='top', 
                   fontsize=8, color=COLORS['text_dim'])
        bar_ax.text(0.5, -0.5, '0%', ha='center', va='top', 
                   fontsize=8, color=COLORS['text_dim'])
        bar_ax.text(1, -0.5, '+100%', ha='right', va='top', 
                   fontsize=8, color=COLORS['text_dim'])
    
    # Add timestamp
    timestamp = _dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ax.text(0.99, -0.05, f"📅 Cập nhật: {timestamp}",
            transform=ax.transAxes,
            ha='right',
            fontsize=8,
            color=COLORS['text_dim'],
            style='italic')
    
    # Adjust layout
    plt.subplots_adjust(top=0.85, bottom=0.1)
    
    # Save with high quality
    fd, path = tempfile.mkstemp(prefix="portfolio_modern_", suffix=".png")
    os.close(fd)
    plt.savefig(path, dpi=200, bbox_inches='tight', 
                facecolor=COLORS['bg_dark'], edgecolor='none')
    plt.close(fig)
    
    return path

# ===================== COMBINED IMAGE: Performance(1Y) + Donut + Table =====================
def gen_portfolio_overview_with_perf(user_id: int, positions, prices) -> str:
    # Data for donut/table
    labels, values, invested_list, pnl_list, prices_now, qtys = [], [], [], [], [], []
    for p in positions:
        sym = p.get("symbol", "")
        qty = float(p.get("qty", 0.0) or 0.0)
        invested = float(max(p.get("invested_usd", 0.0) or 0.0, 0.0))
        cg_id = p.get("cg_id") or cg_guess_id_from_symbol(sym) or ""
        price_now = float(prices.get(cg_id, 0.0) if cg_id else 0.0)
        value_now = qty * price_now
        pnl = value_now - invested

        labels.append(sym or "N/A")
        qtys.append(qty)
        values.append(value_now)
        invested_list.append(invested)
        pnl_list.append(pnl)
        prices_now.append(price_now)

    total_invested = sum(invested_list)
    total_value = sum(values)
    total_pnl = total_value - total_invested
    total_pct = (total_pnl / total_invested * 100.0) if total_invested > 1e-12 else 0.0

    # History 1Y
    dates, hist_values = portfolio_history_series(user_id, days=365)

    # Layout: 3 rows
    # Row0: Performance line (1Y)
    # Row1: Donut (allocation)
    # Row2: Table
    fig = plt.figure(figsize=(14, 12))
    gs = gridspec.GridSpec(3, 1, height_ratios=[1.6, 1.2, 1.8])

    # Performance
    ax_perf = fig.add_subplot(gs[0, 0])
    if not dates or not hist_values:
        dates = ["N/A"]; hist_values = [0.0]
    ax_perf.plot(dates, hist_values)
    ax_perf.set_title("Hiệu suất danh mục (1Y)")
    ax_perf.set_ylabel("USD")
    ax_perf.tick_params(axis='x', rotation=45)

    # Donut
    ax_donut = fig.add_subplot(gs[1, 0])
    labs = labels[:] if labels else ["N/A"]
    vals = values[:] if labels else [1.0]
    tot = sum(vals)
    fracs = [1.0/len(labs)]*len(labs) if tot<=0 else [v/tot for v in vals]
    ax_donut.pie(fracs, labels=labs, autopct="%1.1f%%", wedgeprops=dict(width=0.4))
    ax_donut.set_title("Tỷ trọng danh mục theo giá trị hiện tại")

    # Table
    ax_table = fig.add_subplot(gs[2, 0])
    ax_table.axis("off")
    header = ["Token", "Qty", "Giá now", "Giá trị", "Vốn", "P/L", "P/L %"]
    table_data = [header]
    if labels:
        for i, sym in enumerate(labels):
            inv = invested_list[i]
            val = values[i]
            pnl = pnl_list[i]
            qty_txt = f"{qtys[i]:g}"
            price_now_txt = f"{prices_now[i]:,.2f}"
            pct = (pnl/inv*100) if inv > 1e-12 else 0.0
            table_data.append([sym, qty_txt, price_now_txt, f"{val:,.2f}", f"{inv:,.2f}", f"{pnl:,.2f}", f"{pct:+.2f}%"])
    else:
        table_data.append(["N/A", "0", "0.00", "0.00", "0.00", "0.00", "+0.00%"])
    table_data.append(["TỔNG", "", "", f"{total_value:,.2f}", f"{total_invested:,.2f}", f"{total_pnl:,.2f}", f"{total_pct:+.2f}%"])
    table = ax_table.table(cellText=table_data, loc="center", cellLoc="center")
    table.scale(1, 1.2)
    ax_table.set_title("Chi tiết danh mục (bảng)")

    fig.tight_layout()
    fd, path = tempfile.mkstemp(prefix="cp_overview_", suffix=".png"); os.close(fd)
    plt.savefig(path, dpi=160); plt.close(fig)
    return path

# ===================== UI =====================
CB_CRYPTO_MAIN = "CRYPTO_MAIN"
CB_CRYPTO_PORTFOLIO = "CRYPTO_PORTFOLIO"
CB_CRYPTO_ADD = "CRYPTO_ADD"
CB_CRYPTO_SELL = "CRYPTO_SELL"
CB_CRYPTO_MAP = "CRYPTO_MAP"
CB_CRYPTO_EXPORT = "CRYPTO_EXPORT"
CB_CRYPTO_IMPORT = "CRYPTO_IMPORT"
CB_BACK_HOME = "BACK_HOME"

def main_menu_keyboard():
    return InlineKeyboardMarkup([[InlineKeyboardButton("💹 Crypto", callback_data=CB_CRYPTO_MAIN)]])

def crypto_menu_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Danh mục", callback_data=CB_CRYPTO_PORTFOLIO)],
        [InlineKeyboardButton("➕ Mua", callback_data=CB_CRYPTO_ADD), InlineKeyboardButton("➖ Bán", callback_data=CB_CRYPTO_SELL)],
        [InlineKeyboardButton("🔗 Map token", callback_data=CB_CRYPTO_MAP)],
        [InlineKeyboardButton("⬇️ Nhập CSV hoặc Excel", callback_data=CB_CRYPTO_IMPORT), InlineKeyboardButton("⬆️ Xuất CSV hoặc Excel", callback_data=CB_CRYPTO_EXPORT)],
        [InlineKeyboardButton("⬅️ Về trang chính", callback_data=CB_BACK_HOME)]
    ])

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Chọn từ menu:", reply_markup=main_menu_keyboard())

async def crypto_import_hint_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    # Build and send an import template CSV
    tmp_fd, tmp = tempfile.mkstemp(prefix="crypto_import_template_", suffix=".csv"); os.close(tmp_fd)
    with open(tmp, "w", newline="", encoding="utf-8") as tf:
        tw = csv.writer(tf)
        tw.writerow(["symbol","cg_id","side","qty","price_usd","fee_usd","note","created_at"])
        # sample rows
        tw.writerow(["ATOM","cosmos","BUY","1","10.0","0","","2025-01-01 12:00:00"])
        tw.writerow(["BTC","bitcoin","BUY","0.01","50000","2.5","spot on binance","2025-02-01 20:30:00"])
    try:
        with open(tmp, "rb") as f:
            await q.message.reply_document(f, filename="crypto_import_template.csv", caption="Mẫu CSV hoặc Excel nhập danh mục")
    finally:
        try:
            os.remove(tmp)
        except Exception:
            pass
    help_text = (
    "Gửi *file CSV hoặc Excel (.xlsx/.xlsm)* rồi **REPLY** vào file đó bằng lệnh `/cp_import`.\n"
    "Lưu ý: *Không dùng caption*, vì nhiều cấu hình bot không nhận lệnh trong caption."
)
    await q.edit_message_text(help_text, parse_mode='Markdown', reply_markup=crypto_menu_keyboard())

async def on_crypto_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    # Hiện danh mục (ảnh) trước
    try:
        await send_portfolio_images(q.message.chat.id, q.from_user.id, context)
    except Exception as e:
        await q.message.reply_text(f"Chưa có dữ liệu crypto. Dùng /cp_add để thêm.\nLỗi: {e}")
    # Sau đó hiện các tùy chọn ở dưới, vẫn có nút 📋 Danh mục
    await q.message.reply_text('Chọn thao tác khác:', reply_markup=crypto_menu_keyboard())

async def show_home_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    await q.edit_message_text("🏠 Về trang chính", reply_markup=main_menu_keyboard())

async def quick_hint(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    q = update.callback_query; await q.answer()
    await q.edit_message_text(f"👉 Gõ lệnh:\n`{text}`\n\nHoặc quay lại menu.", parse_mode="Markdown", reply_markup=crypto_menu_keyboard())

async def crypto_add_hint_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Hướng dẫn mua
    await quick_hint(update, context, "/cp_add BTC 0.1 30000 [ghi_chu]")


async def crypto_sell_hint_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Hướng dẫn bán
    await quick_hint(update, context, "/cp_sell BTC 0.05 35000 [ghi_chu]")


async def crypto_map_hint_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Hướng dẫn map token
    await quick_hint(update, context, "/cp_map BTC bitcoin")

def _build_positions_and_prices(user_id: int):
    positions = db_crypto_positions(user_id)
    ids = [p["cg_id"] or cg_guess_id_from_symbol(p["symbol"]) for p in positions if (p["cg_id"] or cg_guess_id_from_symbol(p["symbol"]))]
    prices = cg_simple_price_usd(ids)
    return positions, prices

# ===================== Portfolio actions =====================

def format_portfolio_text(user_id: int) -> str:
    # Trả về tóm tắt danh mục dạng Markdown
    try:
        positions, prices = _build_positions_and_prices(user_id)
    except Exception:
        positions, prices = [], {}
    if not positions:
        return "Danh mục trống. Dùng /cp_add để thêm giao dịch."
    lines = ["*📋 Danh mục hiện tại*"]
    total_invested = 0.0
    total_value = 0.0
    row_lines = []
    for p in positions:
        sym = p.get("symbol","?")
        qty = float(p.get("qty") or 0.0)
        invested = float(p.get("invested_usd") or 0.0)
        cg_id = p.get("cg_id") or cg_guess_id_from_symbol(sym)
        price = float(prices.get(cg_id, 0.0))
        value = qty * price
        pnl = value - invested
        pnl_pct = (pnl / invested * 100.0) if abs(invested) > 1e-9 else 0.0
        total_invested += invested
        total_value += value
        row_lines.append(f"- {sym}: {qty:g} × ${price:,.2f} = ${value:,.2f}  (vốn: ${invested:,.2f}, PnL: {pnl:+,.2f} | {pnl_pct:+.2f}%)")
    tot_pnl = total_value - total_invested
    tot_pct = (tot_pnl / total_invested * 100.0) if abs(total_invested) > 1e-9 else 0.0
    lines += row_lines[:25]
    if len(row_lines) > 25:
        lines.append(f"... và {len(row_lines)-25} mã nữa")
    lines.append("")
    lines.append(f"*Tổng*: vốn ${total_invested:,.2f} → giá trị ${total_value:,.2f} | *PnL*: {tot_pnl:+,.2f} ({tot_pct:+.2f}%)")
    return "\n".join(lines)


async def send_portfolio_images(chat_id: int, user_id: int, context: ContextTypes.DEFAULT_TYPE):
    positions, prices = _build_positions_and_prices(user_id)
    if not positions:
        await context.bot.send_message(chat_id, "Danh mục trống. Dùng /cp_add để thêm giao dịch.")
        return
    try:
        donut = await asyncio.to_thread(gen_portfolio_donut_image, positions, prices)
        table = await asyncio.to_thread(gen_portfolio_table_image, positions, prices)
        with open(donut, "rb") as f1:
            await context.bot.send_photo(chat_id, f1)
        with open(table, "rb") as f2:
            await context.bot.send_photo(chat_id, f2, caption="📋 Danh mục (donut + bảng)")
    except Exception as e:
        # Report error details to user
        await context.bot.send_message(chat_id, f"⚠️ Lỗi khi tạo/gửi ảnh danh mục: {e!r}")
    finally:
        for p in ("donut", "table"):
            try:
                fp = locals().get(p)
                if fp and isinstance(fp, str) and os.path.exists(fp):
                    os.remove(fp)
            except Exception:
                pass

async def show_portfolio_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    await send_portfolio_images(q.message.chat.id, q.from_user.id, context)
    await q.message.reply_text("Chọn thao tác khác:", reply_markup=crypto_menu_keyboard())

async def cp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await send_portfolio_images(update.effective_chat.id, update.effective_user.id, context)

# ===================== Import/Export =====================
async def crypto_export_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    user_id = q.from_user.id
    rows = db_crypto_all_trades(user_id)
    if not rows:
        await q.edit_message_text("Chưa có giao dịch để xuất.", reply_markup=crypto_menu_keyboard())
        return
    headers = ["symbol","cg_id","side","qty","price_usd","fee_usd","note","created_at"]
    fd, path = tempfile.mkstemp(prefix="portfolio_", suffix=".csv"); os.close(fd)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(headers); [w.writerow(r) for r in rows]
    with open(path, "rb") as f:
        await q.message.reply_document(f, filename="crypto_portfolio.csv", caption="⬆️ Danh mục crypto (CSV hoặc Excel)")
    try:
        os.remove(path)
    except Exception:
        pass
    await q.edit_message_text(
        "Đã gửi *file CSV hoặc Excel (.xlsx/.xlsm)* rồi **REPLY** vào file đó bằng lệnh `/cp_import`.\n"
        "Lưu ý: *Không dùng caption*, vì nhiều cấu hình bot không nhận lệnh trong caption.",
        parse_mode="Markdown", reply_markup=crypto_menu_keyboard()
    )

def _parse_trades_from_excel(bio) -> list[dict]:
    """
    Read Excel (xlsx/xlsm) from a BytesIO and return list of row dicts.
    Accepts BOTH:
      A) Import template headers: Date, Exchange, Symbol, Pair, Side, Quantity, Price, Fee, FeeCurrency, Note, CreatedAt
      B) Export headers: symbol, cg_id, side, qty, price_usd, fee_usd, note, created_at
    Sheet: prefers 'Trades', otherwise first sheet.
    """
    if load_workbook is None:
        raise RuntimeError("Excel import requires openpyxl. Please install: pip install openpyxl")
    bio.seek(0)
    wb = load_workbook(filename=bio, read_only=True, data_only=True)
    ws = wb["Trades"] if "Trades" in wb.sheetnames else wb.active

    # Read header row
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    headers = []
    for c in header_cells:
        val = c.value
        if isinstance(val, str):
            headers.append(val.strip())
        else:
            headers.append(val if val is not None else "")
    # Build a lowercase index map
    norm = {str(h).strip().lower(): i for i, h in enumerate(headers) if str(h).strip()}

    def get_val(row_vals, *names):
        for name in names:
            k = name.lower()
            if k in norm and norm[k] < len(row_vals):
                return row_vals[norm[k]]
        return None

    rows: list[dict] = []

    # Iterate rows
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all((v is None or (isinstance(v, str) and v.strip() == "")) for v in r):
            continue

        # Map flexible headers to canonical keys expected by cp_import_cmd
        date_val = get_val(r, "date", "created_at")  # use created_at if no date
        exchange = get_val(r, "exchange")
        symbol = get_val(r, "symbol", "asset", "token")
        pair = get_val(r, "pair")
        side = get_val(r, "side", "action")
        quantity = get_val(r, "quantity", "qty", "amount")
        price = get_val(r, "price", "price_usd")
        fee = get_val(r, "fee", "fee_usd")
        fee_currency = get_val(r, "feecurrency", "fee_currency", "fee cur", "fee currency")
        note = get_val(r, "note", "memo", "comment")
        created_at = get_val(r, "created_at")  # may duplicate date_val

        # If the export format is used, set sensible defaults
        if fee_currency in (None, "") and get_val(r, "fee_usd") is not None:
            fee_currency = "USD"

        # Build canonical dict
        row = {
            "date": date_val,
            "exchange": exchange,
            "symbol": symbol,
            "pair": pair,
            "side": side,
            "quantity": quantity,
            "price": price,
            "fee": fee,
            "fee_currency": fee_currency,
            "note": note,
            "created_at": created_at,
        }
        rows.append(row)

    # If we ended up with zero rows, provide a helpful error
    if not rows:
        have = list(norm.keys())
        raise ValueError(f"No data rows found. Detected headers: {have}. "
                         f"Expected either import template or export columns.")
    return rows

def _normalize_datetime_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Try pandas-like parser via datetime if it's a float/excel serial date we can't handle directly
    # Handle Excel datetime objects (openpyxl may give datetime)
    try:
        # If already datetime
        if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
            return val.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    # Try common formats
    fmts = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"]
    for f in fmts:
        try:
            return _dt.datetime.strptime(s, f).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
    # Fallback: try fromisoformat
    try:
        return _dt.datetime.fromisoformat(s).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    raise ValueError(f"Invalid datetime: {s}")

async def cp_import_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):

    user = update.effective_user
    doc = None
    if update.message and update.message.document:
        doc = update.message.document
    if not doc and update.message and update.message.reply_to_message and update.message.reply_to_message.document:
        doc = update.message.reply_to_message.document
    if not doc:
        await update.message.reply_text(
            "Hãy gửi *file CSV hoặc Excel (xlsx/xlsm)* rồi reply `/cp_import` vào file đã gửi.",
            parse_mode="Markdown"
        )
        return

    await update.message.reply_text("📥 Đang xử lý file, vui lòng chờ...")

    fobj = await context.bot.get_file(doc.file_id)
    bio = io.BytesIO(); await fobj.download_to_memory(out=bio); bio.seek(0)

    filename = (doc.file_name or "").lower()
    mime = (doc.mime_type or "").lower()

    ok = fail = 0

    try:
        if filename.endswith((".xlsx",".xlsm")) or "spreadsheetml" in mime:
            # Excel path
            rows = _parse_trades_from_excel(bio)
            for row in rows:
                try:
                    symbol = (row.get("symbol") or "").strip().upper()
                    side = (row.get("side") or "").strip().upper()
                    qty = float(row.get("quantity") or 0)
                    price = float(row.get("price") or 0)
                    fee = float(row.get("fee") or 0) if str(row.get("fee") or "").strip() != "" else 0.0
                    note = (row.get("note") or "").strip()
                    at_raw = row.get("created_at")
                    at_norm = None
                    if at_raw not in (None, ""):
                        try:
                            at_norm = _normalize_datetime_str(at_raw)
                            created_at = _dt.datetime.fromisoformat(at_norm)
                        except Exception:
                            created_at = None
                    else:
                        created_at = None
                    cg_id = db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
                    if cg_id:
                        db_crypto_upsert_map(symbol, cg_id)
                    if not symbol or side not in ("BUY","SELL") or qty<=0 or price<=0:
                        fail += 1; continue
                    db_crypto_add_trade(user.id, symbol, side, qty, price, note=note, cg_id=cg_id, fee_usd=fee, created_at=created_at)
                    ok += 1
                except Exception:
                    fail += 1
        else:
            # CSV hoặc Excel path (UTF-8)
            reader = csv.DictReader(io.TextIOWrapper(bio, encoding="utf-8"))
            for row in reader:
                try:
                    symbol = (row.get("symbol") or row.get("Symbol") or "").strip().upper()
                    side = (row.get("side") or row.get("Side") or "").strip().upper()
                    qty = float(row.get("qty") or row.get("quantity") or row.get("Quantity") or 0)
                    price = float(row.get("price") or row.get("price_usd") or row.get("Price") or 0)
                    fee = float(row.get("fee") or row.get("Fee") or 0) if str(row.get("fee") or row.get("Fee") or "").strip() != "" else 0.0
                    note = (row.get("note") or row.get("Note") or "").strip()
                    cg_id = (row.get("cg_id") or row.get("CG_ID") or row.get("cgid") or "").strip() or None
                    at_raw = (row.get("created_at") or row.get("CreatedAt") or "").strip()
                    if not symbol or side not in ("BUY","SELL") or qty<=0 or price<=0:
                        fail += 1; continue
                    created_at = None
                    if at_raw:
                        try:
                            created_at = _dt.datetime.fromisoformat(at_raw)
                        except Exception:
                            created_at = None
                    if not cg_id:
                        cg_id = db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
                    if cg_id:
                        db_crypto_upsert_map(symbol, cg_id)
                    db_crypto_add_trade(user.id, symbol, side, qty, price, note=note, cg_id=cg_id, fee_usd=fee, created_at=created_at)
                    ok += 1
                except Exception:
                    fail += 1
    except RuntimeError as e:
        await update.message.reply_text(f"❌ {e}")
        return

    await update.message.reply_text(f"✅ Đã nhập {ok} giao dịch, lỗi {fail}. Dùng /cp để xem danh mục.")

    
    # Safety net: if any unhandled error occurs, inform user
    try:
        pass
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi không xác định khi nhập: {type(e).__name__}: {e}")

# ===================== Commands =====================

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = """*📖 HƯỚNG DẪN NHANH*

*1) Mở menu chính*
• Gõ /start để hiện menu nút.

*2) Crypto*
• Bấm nút *Crypto*: bot sẽ hiện *danh mục đầu tư* (ảnh donut + bảng) và các nút thao tác.
• *Mua*: dùng lệnh `/cp_add SYMBOL QTY PRICE_USD [ghi_chu]`
  └ Ví dụ: `/cp_add BTC 0.02 60000 mua thử`
• *Bán*: dùng lệnh `/cp_sell SYMBOL QTY PRICE_USD [ghi_chu]`
  └ Ví dụ: `/cp_sell BTC 0.01 62000 chốt lời`
• *Map token* (gắn mã với CoinGecko ID) dùng lệnh `/cp_map SYMBOL CG_ID`
  └ Ví dụ: `/cp_map TON the-open-network`, `/cp_map ATOM cosmos`
• *Nhập danh mục*: bấm "⬇️ Nhập CSV hoặc Excel" để xem *mẫu file*. Sau đó *gửi file* và *REPLY* vào file đó bằng lệnh `/cp_import`.
• *Xuất danh mục*: bấm "⬆️ Xuất CSV hoặc Excel" để tải file.
• *Xem lại danh mục*: bấm "📋 Danh mục" hoặc gõ `/cp`.

*3) Gợi ý cú pháp nhanh*
• Mua: `/cp_add BTC 0.1 30000 [ghi_chu]`
• Bán: `/cp_sell BTC 0.05 35000 [ghi_chu]`
• Map: `/cp_map BTC bitcoin`
• Xem: `/cp`
• Nhập: gửi file rồi reply: `/cp_import`

_Mẹo: Khi bấm các nút *Mua / Bán / Map token* trong menu Crypto, bot sẽ hiển thị lại hướng dẫn cú pháp ngay bên dưới._
"""
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=main_menu_keyboard())

def _parse_trade_args(args: List[str]):
    if len(args) < 3: return None
    symbol = args[0].upper()
    try: qty = float(args[1]); price = float(args[2])
    except Exception: return None
    note = " ".join(args[3:]).strip() if len(args) > 3 else ""
    return symbol, qty, price, note

async def cp_map(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_migrated()
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Cú pháp: /cp_map SYMBOL CG_ID")
        return
    symbol, cg_id = args[0].upper(), args[1]
    db_crypto_upsert_map(symbol, cg_id)
    await update.message.reply_text(f"✅ Đã map {symbol} -> {cg_id}")

async def cp_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_migrated()
    parsed = _parse_trade_args(context.args)
    if not parsed:
        await update.message.reply_text("Cú pháp: /cp_add SYMBOL QTY PRICE_USD [note]")
        return
    symbol, qty, price, note = parsed
    if qty <= 0 or price <= 0:
        await update.message.reply_text("Số lượng & giá phải > 0"); return
    cg_id = db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
    if cg_id and not db_crypto_get_map(symbol):
        db_crypto_upsert_map(symbol, cg_id)
    if cg_id and not db_crypto_get_map(symbol):
        db_crypto_upsert_map(symbol, cg_id)
    if cg_id: db_crypto_upsert_map(symbol, cg_id)
    db_crypto_add_trade(update.effective_user.id, symbol, "BUY", qty, price, note=note, cg_id=cg_id)
    await update.message.reply_text(f"✅ BUY {symbol}: {qty:g} @ {price:,.2f} USD")

async def cp_sell(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parsed = _parse_trade_args(context.args)
    if not parsed:
        await update.message.reply_text("Cú pháp: /cp_sell SYMBOL QTY PRICE_USD [note]")
        return
    symbol, qty, price, note = parsed
    if qty <= 0 or price <= 0:
        await update.message.reply_text("Số lượng & giá phải > 0"); return
    cg_id = db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
    if cg_id: db_crypto_upsert_map(symbol, cg_id)
    db_crypto_add_trade(update.effective_user.id, symbol, "SELL", qty, price, note=note, cg_id=cg_id)
    await update.message.reply_text(f"✅ SELL {symbol}: {qty:g} @ {price:,.2f} USD")

# ===================== Error Handler =====================
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        err = "".join(traceback.format_exception(None, context.error, context.error.__traceback__))
    except Exception:
        err = str(context.error)
    print("===== ERROR ====="); print("Update:", update); print(err)
    try:
        chat_id = update.effective_chat.id if update and hasattr(update, "effective_chat") else None
        if chat_id:
            await context.bot.send_message(chat_id, "⚠️ Có lỗi xảy ra. Hãy thử lại.")
    except Exception:
        pass

# ===================== App =====================

# ===== Crypto UI Handlers (moved earlier) =====
async def on_crypto_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    try:
        await send_portfolio_images(chat_id, user_id, context)
    except Exception as e:
        if update.message:
            await update.message.reply_text(f"Chưa có dữ liệu crypto. Dùng /cp_add để thêm.\nLỗi: {e}")
    if update.message:
        await update.message.reply_text('Chọn thao tác khác:', reply_markup=crypto_menu_keyboard())

async def on_crypto_menu_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    text = format_portfolio_text(q.from_user.id)
    await q.message.reply_text(text, parse_mode='Markdown', reply_markup=crypto_menu_keyboard())
async def on_crypto_portfolio_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    if 'send_portfolio_images' in globals():
        await send_portfolio_images(q.message.chat.id, q.from_user.id, context)
async def on_crypto_import_hint_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    help_text = (
    "Gửi *file CSV hoặc Excel (.xlsx/.xlsm)* rồi **REPLY** vào file đó bằng lệnh `/cp_import`.\n"
    "Lưu ý: *Không dùng caption*, vì nhiều cấu hình bot không nhận lệnh trong caption."
)
    await q.edit_message_text(help_text, parse_mode='Markdown', reply_markup=crypto_menu_keyboard())
async def on_crypto_export_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    if 'cp_export_cmd' in globals():
        await cp_export_cmd(update, context)
    else:
        await q.message.reply_text("Nhập: /cp_export để xuất CSV/XLSX")



def build_app() -> Application:
    run_migrations()
            
            

    app = Application.builder().token(BOT_TOKEN).build()

    # Commands
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("cp", cp))
    app.add_handler(CommandHandler("cp_add", cp_add))
    app.add_handler(CommandHandler("cp_sell", cp_sell))
    app.add_handler(CommandHandler("cp_map", cp_map))
    app.add_handler(CommandHandler("cp_import", cp_import_cmd))
    # Text buttons / entries
    app.add_handler(MessageHandler(filters.Regex(r"(?i)^(?:🪙|💎)?\s*crypto$"), on_crypto_entry))

    # Callbacks
    app.add_handler(CallbackQueryHandler(on_crypto_menu, pattern=f"^{CB_CRYPTO_MAIN}$"))
    app.add_handler(CallbackQueryHandler(show_portfolio_cb, pattern=f"^{CB_CRYPTO_PORTFOLIO}$"))
    app.add_handler(CallbackQueryHandler(crypto_import_hint_cb, pattern=f"^{CB_CRYPTO_IMPORT}$"))
    app.add_handler(CallbackQueryHandler(crypto_export_cb, pattern=f"^{CB_CRYPTO_EXPORT}$"))
    app.add_handler(CallbackQueryHandler(crypto_add_hint_cb, pattern=f"^{CB_CRYPTO_ADD}$"))
    app.add_handler(CallbackQueryHandler(crypto_sell_hint_cb, pattern=f"^{CB_CRYPTO_SELL}$"))
    app.add_handler(CallbackQueryHandler(crypto_map_hint_cb, pattern=f"^{CB_CRYPTO_MAP}$"))

    app.add_handler(CallbackQueryHandler(show_home_cb, pattern=f"^{CB_BACK_HOME}$"))

    app.add_error_handler(error_handler)
    return app

# ==== END: Inlined from EM ĐÃ OK.py ====
from datetime import datetime as _dt2, timedelta as _td2

def parse_amount_loose(text: str) -> float | None:
    t = text.lower().replace(" ", "")
    # Chuẩn hóa đơn vị
    t = (t.replace("triệu","tr").replace("trieu","tr")
           .replace("nghìn","k").replace("nghin","k")
           .replace("ngàn","k").replace("ngan","k"))
    import re as _re2local
    # Dạng '1tr5' => 1_500_000 ; '2k5' => 2_500
    m = _re2local.match(r'^(\d+)(tr)(\d+)$', t)
    if m:
        return int(m.group(1)) * 1_000_000 + int(m.group(3)) * 100_000
    m = _re2local.match(r'^(\d+)(k)(\d+)$', t)
    if m:
        return int(m.group(1)) * 1_000 + int(m.group(3)) * 100
    # Dạng số kèm đơn vị đơn lẻ: 200k, 3tr, 1.5tr, 2.75k ...
    m = _re2local.search(r'(\d{1,3}(?:[.,]\d{3})+|\d+(?:[.,]\d+)?)(k|tr)?', t)
    if not m: 
        return None
    num, unit = m.group(1), m.group(2) or ""
    num = num.replace(".", "").replace(",", ".")
    try:
        val = float(num)
    except ValueError:
        return None
    if unit == "k": 
        val *= 1000
    elif unit == "tr": 
        val *= 1_000_000
    return round(val)

def guess_type(text: str) -> str:
    low = text.lower()
    for kw in INCOME_KEYWORDS:
        if kw in low:
            return "income"
    return "expense"

def guess_category(text: str) -> str:
    low = text.lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        for kw in kws:
            if kw in low:
                return cat
    return "Khác"

def guess_date(text: str):
    low = text.lower()
    for k, delta in DATE_KEYWORDS.items():
        if k in low:
            return _dt2.now() + _td2(days=delta)
    m = _re2.search(r'(\d{1,2})[\/\-](\d{1,2})(?:[\/\-](\d{2,4}))?', low)
    if m:
        d, mth, yr = int(m.group(1)), int(m.group(2)), m.group(3)
        year = int(yr) if yr else _dt2.now().year
        try:
            return _dt2(year, mth, d)
        except ValueError:
            pass
    return _dt2.now()

def clean_note(text: str) -> str:
    note = _re2.sub(r'\d{1,3}(?:[.,]\d{3})+|\d+(?:[.,]\d+)?\s*(k|tr|triệu|nghìn|ngàn)?', '', text, flags=_re2.IGNORECASE)
    return note.strip(" -–—,.;:") or "Không ghi chú"

def nlp_parse_transaction(raw: str):
    amount = parse_amount_loose(raw)
    if amount is None or amount <= 0: return None
    tx_type = guess_type(raw)
    created_at = guess_date(raw)
    note = clean_note(raw)
    if tx_type == "income":
        category = guess_income_source(raw)  # dùng danh mục như 'Nguồn' cho income
    else:
        category = guess_category(raw)
    return {"type": tx_type, "amount": amount, "category": category, "note": note, "created_at": created_at}
import logging
import sqlite3
import datetime
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import pandas as pd
import io

# helper: normalize a date/datetime to 'YYYY-MM-DD' string
def _fmt_day_label(d):
    try:
        # datetime -> date
        d = d.date()
    except AttributeError:
        pass
    return f"{d:%Y-%m-%d}" if hasattr(d, 'strftime') else str(d)
 # Để xử lý ảnh và file trong bộ nhớ
import os # Để xóa file tạm nếu cần
import numpy as np
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler

# --- HÀM MÃ HÓA TÊN NGƯỜI DÙNG ---
def mask_user_name(full_name: str) -> str:
    """Ẩn tên người dùng, chỉ hiển thị 3 ký tự cuối."""
    full_name = full_name.strip()
    if len(full_name) <= 3:
        return '****' + full_name
    return '****' + full_name[-3:]

# --- CẤU HÌNH ---
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

# Kênh Telegram để đăng thông báo giao dịch mới.
CHANNEL_ID = "@congthuchi" 

# Trạng thái cho ConversationHandler
CHOOSING_INCOME_AMOUNT, CHOOSING_INCOME_NOTE, CHOOSING_INCOME_DATE, CHOOSING_INCOME_SOURCE, CONFIRM_INCOME_AMOUNT = range(5)
CHOOSING_EXPENSE_AMOUNT, CHOOSING_EXPENSE_NOTE, CHOOSING_EXPENSE_DATE, CHOOSING_EXPENSE_CATEGORY, CONFIRM_EXPENSE_AMOUNT = range(5, 10)
CHOOSING_ADD_TYPE = 10
CHOOSING_REPORT_PERIOD = 11
CHOOSING_SUPPLEMENT_TYPE = 12
DELETE_CHOOSING_ACTION = 13
DELETE_INDIVIDUAL_LISTING = 14
DELETE_INDIVIDUAL_CONFIRM = 15
RESET_DATA_CONFIRM_STEP1 = 16
RESET_DATA_CONFIRM_STEP2 = 17
CHOOSING_CHART_PERIOD = 18
EXPORT_CHOOSING_FORMAT = 19
# Thêm trạng thái mới cho việc nhập file
CHOOSING_IMPORT_ACTION = 20
IMPORT_FILE_UPLOADED = 21

# DB_PATH unified above
LARGE_AMOUNT_THRESHOLD = 100_000_000

# --- CÁC HÀM XỬ LÝ LOGIC (Không thay đổi) ---
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            amount REAL,
            note TEXT,
            category TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS incomes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            amount REAL,
            source TEXT,
            note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS budgets (
            user_id INTEGER,
            category TEXT,
            amount REAL,
            PRIMARY KEY (user_id, category)
        )
    """)
    conn.commit()
    conn.close()

def run_migrations():
    """Simple schema migrations using PRAGMA user_version."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("PRAGMA user_version")
    ver = cur.fetchone()[0] or 0
    # v1 -> v2 : add budgets table if not exists
    if ver < 2:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS budgets (
                user_id INTEGER,
                category TEXT,
                amount REAL,
                PRIMARY KEY (user_id, category)
            )
        """
        )
        cur.execute("PRAGMA user_version = 2")
        conn.commit()
    conn.close()
def parse_amount(text: str) -> float:
    # Ưu tiên dùng bộ phân tích rút gọn tiếng Việt (hỗ trợ 2k5, 1tr5, 'củ', ...)
    v = parse_vietnamese_amount(text)
    if v is not None:
        return float(v)
    # Thử bộ lỏng
    v2 = parse_amount_loose(text)
    if v2 is not None:
        return float(v2)
    # Cuối cùng: loại bỏ ký tự không số rồi parse
    import re as _re_fallback
    t = str(text).lower().strip()
    t = t.replace(",", "").replace(" ", "")
    if t.endswith("tr"):
        try:
            return float(t[:-2]) * 1_000_000
        except Exception:
            pass
    if t.endswith("k"):
        try:
            return float(t[:-1]) * 1_000
        except Exception:
            pass
    return float(_re_fallback.sub(r'[^0-9.]', '', t) or 0)

def parse_date(date_str: str) -> datetime.datetime:
    today = datetime.date.today()
    try:
        return datetime.datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        try:
            return datetime.datetime.strptime(f"{date_str}/{today.year}", "%d/%m/%Y")
        except ValueError:
            raise ValueError("Định dạng ngày không hợp lệ. Vui lòng nhập DD/MM/YYYY hoặc DD/MM.")

def db_add_income(user_id: int, amount: float, source: str, note: str, created_at: datetime.datetime = None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    if created_at is None:
        created_at = datetime.datetime.now()
    cursor.execute("INSERT INTO incomes (user_id, amount, source, note, created_at) VALUES (?, ?, ?, ?, ?)", 
                   (user_id, amount, source, note, created_at.isoformat()))
    conn.commit()
    conn.close()

def db_add_expense(user_id: int, amount: float, note: str, category: str, created_at: datetime.datetime = None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    if created_at is None:
        created_at = datetime.datetime.now()
    cursor.execute("INSERT INTO expenses (user_id, amount, note, category, created_at) VALUES (?, ?, ?, ?, ?)",
                   (user_id, amount, note, category, created_at.isoformat()))
    conn.commit()
    conn.close()

def db_list_expenses_grouped(user_id: int, start_date=None, end_date=None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    query = "SELECT category, SUM(amount) FROM expenses WHERE user_id = ?"
    params = [user_id]
    if start_date and end_date:
        end_date_inclusive = end_date + datetime.timedelta(days=1)
        query += " AND created_at >= ? AND created_at < ?"
        params.extend([start_date.isoformat(), end_date_inclusive.isoformat()])
    query += " GROUP BY category ORDER BY SUM(amount) DESC"
    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    conn.close()
    return rows

def db_list_incomes_grouped(user_id: int, start_date=None, end_date=None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    query = "SELECT source, SUM(amount) FROM incomes WHERE user_id = ?"
    params = [user_id]
    if start_date and end_date:
        end_date_inclusive = end_date + datetime.timedelta(days=1)
        query += " AND created_at >= ? AND created_at < ?"
        params.extend([start_date.isoformat(), end_date_inclusive.isoformat()])
    query += " GROUP BY source ORDER BY SUM(amount) DESC"
    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    conn.close()
    return rows

def db_get_monthly_report(user_id, start_date=None, end_date=None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    expense_query = "SELECT strftime('%Y-%m', created_at) as month, SUM(amount) FROM expenses WHERE user_id = ?"
    income_query = "SELECT strftime('%Y-%m', created_at) as month, SUM(amount) FROM incomes WHERE user_id = ?"
    params = [user_id]
    if start_date and end_date:
        end_date_inclusive = end_date + datetime.timedelta(days=1)
        expense_query += " AND created_at >= ? AND created_at < ?"
        income_query += " AND created_at >= ? AND created_at < ?"
        params.extend([start_date.isoformat(), end_date_inclusive.isoformat()])
    expense_query += " GROUP BY month ORDER BY month"
    income_query += " GROUP BY month ORDER BY month"
    cursor.execute(expense_query, tuple(params))
    expenses = cursor.fetchall()
    cursor.execute(income_query, tuple(params))
    incomes = cursor.fetchall()
    conn.close()
    return expenses, incomes

def db_get_combined_summary(user_id: int, start_date=None, end_date=None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    income_query = "SELECT SUM(amount) FROM incomes WHERE user_id = ?"
    expense_query = "SELECT SUM(amount) FROM expenses WHERE user_id = ?"
    params = [user_id]
    if start_date and end_date:
        end_date_inclusive = end_date + datetime.timedelta(days=1)
        income_query += " AND created_at >= ? AND created_at < ?"
        expense_query += " AND created_at >= ? AND created_at < ?"
        params.extend([start_date.isoformat(), end_date_inclusive.isoformat()])
    cursor.execute(income_query, tuple(params))
    income = cursor.fetchone()[0] or 0
    cursor.execute(expense_query, tuple(params))
    expense = cursor.fetchone()[0] or 0
    conn.close()
    return income, expense, income - expense

def db_export_data(user_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 'Chi tiêu' as type, amount, note, category as label, created_at, id
        FROM expenses WHERE user_id = ?
        UNION ALL
        SELECT 'Thu nhập' as type, amount, note, source as label, created_at, id
        FROM incomes WHERE user_id = ?
        ORDER BY created_at DESC
    """, (user_id, user_id))
    rows = cursor.fetchall()
    headers = ['Loại', 'Số tiền', 'Ghi chú', 'Danh mục/Nguồn', 'Ngày', 'ID_Giao_dich']
    conn.close()
    return pd.DataFrame(rows, columns=headers)

def db_get_last_n_transactions(user_id: int, limit: int = 5):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, 'expense' as type, amount, note, category, created_at FROM expenses WHERE user_id = ?
        UNION ALL
        SELECT id, 'income' as type, amount, note, source, created_at FROM incomes WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT ?
    """, (user_id, user_id, limit))
    transactions = cursor.fetchall()
    conn.close()
    return transactions

def db_delete_transaction(transaction_id: int, transaction_type: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    table_name = "expenses" if transaction_type == 'expense' else "incomes"
    cursor.execute(f"DELETE FROM {table_name} WHERE id = ?", (transaction_id,))
    conn.commit()
    conn.close()

def db_reset_all_data(user_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM expenses WHERE user_id = ?", (user_id,))
    cursor.execute("DELETE FROM incomes WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

# Thêm hàm mới để xử lý file nhập
def db_add_transactions_from_df(user_id: int, df: pd.DataFrame):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Chuyển đổi tên cột cho đồng nhất
    df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]
    
    # Kiểm tra các cột cần thiết
    required_cols = ['loại', 'số_tiền', 'ghi_chú', 'danh_mục/nguồn', 'ngày']
    if not all(col in df.columns for col in required_cols):
        raise ValueError("File thiếu các cột bắt buộc: Loại, Số tiền, Ghi chú, Danh mục/Nguồn, Ngày")

    # Xử lý các giao dịch
    success_count = {'income': 0, 'expense': 0}
    for index, row in df.iterrows():
        try:
            trans_type = row['loại'].strip().lower()
            amount = float(row['số_tiền'])
            note = str(row['ghi_chú']).strip()
            cat_or_src = str(row['danh_mục/nguồn']).strip()
            created_at = pd.to_datetime(row['ngày'])

            if trans_type == 'thu nhập':
                cursor.execute("INSERT INTO incomes (user_id, amount, source, note, created_at) VALUES (?, ?, ?, ?, ?)",
                               (user_id, amount, cat_or_src, note, created_at.isoformat()))
                success_count['income'] += 1
            elif trans_type == 'chi tiêu':
                cursor.execute("INSERT INTO expenses (user_id, amount, note, category, created_at) VALUES (?, ?, ?, ?, ?)",
                               (user_id, amount, note, cat_or_src, created_at.isoformat()))
                success_count['expense'] += 1
        except Exception as e:
            logger.error(f"Lỗi khi xử lý hàng {index+1}: {e}")
            continue

    conn.commit()
    conn.close()
    return success_count

def db_set_budget(user_id: int, category: str, amount: float):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO budgets (user_id, category, amount)
        VALUES (?, ?, ?)
        ON CONFLICT(user_id, category) DO UPDATE SET amount = excluded.amount
        """,
        (user_id, category, amount)
    )
    conn.commit()
    conn.close()

def db_get_budgets(user_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT category, amount FROM budgets WHERE user_id = ?", (user_id,))
    rows = cursor.fetchall()
    conn.close()
    return {r[0]: r[1] for r in rows}

def db_delete_budget(user_id: int, category: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM budgets WHERE user_id = ? AND category = ?", (user_id, category))
    conn.commit()
    conn.close()

def get_month_range(dt: datetime.datetime):
    start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month = (start.replace(day=28) + datetime.timedelta(days=4)).replace(hour=0, minute=0, second=0, microsecond=0)
    end = (next_month - datetime.timedelta(days=next_month.day)).replace(hour=23, minute=59, second=59, microsecond=999999)
    return start, end

def db_sum_expense_by_category_in_period(user_id: int, category: str, start_date: datetime.datetime, end_date: datetime.datetime) -> float:
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT COALESCE(SUM(amount), 0) FROM expenses
        WHERE user_id = ? AND category = ? AND created_at >= ? AND created_at <= ?
        """,
        (user_id, category, start_date.isoformat(), end_date.isoformat())
    )
    total = cursor.fetchone()[0] or 0.0
    conn.close()
    return float(total)

def get_main_menu_keyboard():
    keyboard = [
        [KeyboardButton("➕ Thêm"), KeyboardButton("🗑️ Xóa"), KeyboardButton("📈 Báo Cáo")],
        [KeyboardButton("📊 Biểu Đồ"), KeyboardButton("📤 Xuất File"), KeyboardButton("📥 Nhập File")],
        [KeyboardButton("🪙 Crypto"), KeyboardButton("❓ Hướng Dẫn")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False, is_persistent=True)

# --- CÁC HÀM XỬ LÝ BOT ---

def anonymize_name(name: str) -> str:
    if not name:
        return "****"
    if len(name) < 2:
        return "*"
    masked_part = '*' * (len(name) - 1)
    last_char = name[-1]
    return f"{masked_part}{last_char}"

async def post_to_channel(context: ContextTypes.DEFAULT_TYPE, message: str):
    try:
        await context.bot.send_message(chat_id=CHANNEL_ID, text=message, parse_mode='Markdown')
        logger.info(f"Đã đăng tin thành công lên kênh {CHANNEL_ID}")
    except Exception as e:
        logger.error(f"Lỗi không thể đăng tin lên kênh {CHANNEL_ID}: {e}. "
                     "Hãy chắc chắn rằng bot là Quản trị viên của kênh và có quyền đăng tin nhắn.")

async def send_or_update_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, message_text: str = "Đây là menu chính của bạn:"):
    # This function is meant to ensure the main menu is always visible and in a consistent state.
    # It tries to edit if possible, otherwise sends a new message.
    if update.callback_query:
        try:
            # Try to edit the message the callback came from
            await update.callback_query.edit_message_text(message_text, reply_markup=get_main_menu_keyboard())
        except Exception:
            # If editing fails (e.g., message was deleted or too old), send a new message
            await update.callback_query.message.reply_text(message_text, reply_markup=get_main_menu_keyboard())
    elif update.message:
        await update.message.reply_text(message_text, reply_markup=get_main_menu_keyboard())

# Hàm chung để kết thúc hội thoại và dọn dẹp
async def end_current_conversation_and_reset(update: Update, context: ContextTypes.DEFAULT_TYPE, message_text: str) -> int:
    # Xóa dữ liệu người dùng liên quan đến hội thoại hiện tại
    context.user_data.clear()
    
    # Gửi tin nhắn hủy bỏ/hoàn tất và hiển thị lại menu chính
    await send_or_update_main_menu(update, context, message_text)
    
    return ConversationHandler.END

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    init_db()
    
    run_migrations()
            
            
    return await end_current_conversation_and_reset(update, context, 
        f"Chào mừng bạn đến với Bot Quản lý Tài chính!\n\n"
        f"Sử dụng các nút bên dưới hoặc gõ các lệnh để tương tác.\n"
    )

async def show_main_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await end_current_conversation_and_reset(update, context, "Đây là menu chính của bạn:")

async def cancel_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Hàm hủy bỏ chung cho CommandHandler /cancel"""
    return await end_current_conversation_and_reset(update, context, 'Đã hủy bỏ thao tác.')

async def cancel_conversation_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Hàm hủy bỏ chung cho CallbackQueryHandler (nút ❌ Hủy bỏ)"""
    return await end_current_conversation_and_reset(update, context, 'Đã hủy bỏ thao tác.')

async def handle_add_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Ensure current conversation is ended cleanly before starting new one
    context.user_data.clear() # Clear any remnants from previous conversations
    keyboard = [
        [InlineKeyboardButton("Thu nhập (Hôm nay)", callback_data="add_type_income_today")],
        [InlineKeyboardButton("Chi tiêu (Hôm nay)", callback_data="add_type_expense_today")],
        [InlineKeyboardButton("Bổ sung giao dịch", callback_data="add_type_supplement")],
        [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text("Bạn muốn thêm loại giao dịch nào?", reply_markup=reply_markup)
    else:
        await update.message.reply_text("Bạn muốn thêm loại giao dịch nào?", reply_markup=reply_markup)
    return CHOOSING_ADD_TYPE

async def choose_add_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data
    if choice == "add_type_income_today":
        context.user_data['transaction_type'] = "income"
        context.user_data['created_at'] = datetime.datetime.now()
        await query.edit_message_text("Nhập số tiền thu nhập (VD: 5tr, 200k):")
        return CHOOSING_INCOME_AMOUNT
    elif choice == "add_type_expense_today":
        context.user_data['transaction_type'] = "expense"
        context.user_data['created_at'] = datetime.datetime.now()
        await query.edit_message_text("Nhập số tiền chi tiêu (VD: 5tr2, 100k):")
        return CHOOSING_EXPENSE_AMOUNT
    elif choice == "add_type_supplement":
        keyboard = [
            [InlineKeyboardButton("Thu nhập", callback_data="supplement_type_income")],
            [InlineKeyboardButton("Chi tiêu", callback_data="supplement_type_expense")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Bạn muốn bổ sung Thu nhập hay Chi tiêu?", reply_markup=reply_markup)
        return CHOOSING_SUPPLEMENT_TYPE
    return await cancel_conversation_callback(update, context)

async def choose_supplement_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data
    context.user_data.pop('created_at', None) # Clear any pre-existing date for supplement
    if choice == "supplement_type_income":
        context.user_data['transaction_type'] = "income"
        await query.edit_message_text("Nhập số tiền thu nhập (VD: 5tr, 200k):")
        return CHOOSING_INCOME_AMOUNT
    elif choice == "supplement_type_expense":
        context.user_data['transaction_type'] = "expense"
        await query.edit_message_text("Nhập số tiền chi tiêu (VD: 5tr2, 100k):")
        return CHOOSING_EXPENSE_AMOUNT
    return await cancel_conversation_callback(update, context)

async def get_income_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        amount = parse_amount(update.message.text)
        if amount <= 0:
            await update.message.reply_text("Số tiền phải lớn hơn 0. Vui lòng nhập lại:")
            return CHOOSING_INCOME_AMOUNT
        if amount > LARGE_AMOUNT_THRESHOLD:
            context.user_data['amount_to_confirm'] = amount
            keyboard = [
                [InlineKeyboardButton("✅ Đúng vậy", callback_data="confirm_amount_yes")],
                [InlineKeyboardButton("❌ Nhập lại", callback_data="confirm_amount_no")],
                [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text(
                f"Bạn chắc chắn muốn ghi nhận khoản thu nhập {amount:,.0f} đ chứ? Số tiền này khá lớn.",
                reply_markup=reply_markup
            )
            return CONFIRM_INCOME_AMOUNT
        else:
            context.user_data['amount'] = amount
            await update.message.reply_text("Tuyệt vời! Bây giờ, nhập ghi chú cho khoản thu này:")
            return CHOOSING_INCOME_NOTE
    except (ValueError, TypeError):
        await update.message.reply_text("Số tiền không hợp lệ. Vui lòng nhập lại (VD: 5tr, 200k):")
        return CHOOSING_INCOME_AMOUNT

async def confirm_income_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_amount_yes":
        context.user_data['amount'] = context.user_data.pop('amount_to_confirm')
        await query.edit_message_text("Đã xác nhận. Bây giờ, nhập ghi chú cho khoản thu này:")
        return CHOOSING_INCOME_NOTE
    elif query.data == "confirm_amount_no":
        context.user_data.pop('amount_to_confirm', None)
        await query.edit_message_text("Vui lòng nhập lại số tiền thu nhập:")
        return CHOOSING_INCOME_AMOUNT
    return await cancel_conversation_callback(update, context)

async def get_income_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['note'] = update.message.text
    if 'created_at' in context.user_data: # If date was set (today)
        sources = ["Lương","Thưởng","Kinh doanh","Được cho","Thu hồi nợ","Bán tài sản","Lãi tiết kiệm","Lãi đầu tư","Khác"]
        keyboard = [
            [InlineKeyboardButton("Lương", callback_data="income_source_Lương"), InlineKeyboardButton("Thưởng", callback_data="income_source_Thưởng")],
            [InlineKeyboardButton("Kinh doanh", callback_data="income_source_Kinh doanh"), InlineKeyboardButton("Được cho", callback_data="income_source_Được cho")],
            [InlineKeyboardButton("Thu hồi nợ", callback_data="income_source_Thu hồi nợ"), InlineKeyboardButton("Bán tài sản", callback_data="income_source_Bán tài sản")],
            [InlineKeyboardButton("Lãi tiết kiệm", callback_data="income_source_Lãi tiết kiệm"), InlineKeyboardButton("Lãi đầu tư", callback_data="income_source_Lãi đầu tư")],
            [InlineKeyboardButton("Khác", callback_data="income_source_Khác")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Chọn nguồn thu nhập:", reply_markup=reply_markup)
        return CHOOSING_INCOME_SOURCE
    else: # Need to ask for date
        keyboard = [
            [InlineKeyboardButton("Hôm nay", callback_data="date_today_income")],
            [InlineKeyboardButton("Ngày khác", callback_data="date_other_income")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Giao dịch này xảy ra vào khi nào?", reply_markup=reply_markup)
        return CHOOSING_INCOME_DATE

async def choose_income_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data
    if choice == "date_today_income":
        context.user_data['created_at'] = datetime.datetime.now()
        sources = ["Lương","Thưởng","Kinh doanh","Được cho","Thu hồi nợ","Bán tài sản","Lãi tiết kiệm","Lãi đầu tư","Khác"]
        keyboard = [
            [InlineKeyboardButton("Lương", callback_data="income_source_Lương"), InlineKeyboardButton("Thưởng", callback_data="income_source_Thưởng")],
            [InlineKeyboardButton("Kinh doanh", callback_data="income_source_Kinh doanh"), InlineKeyboardButton("Được cho", callback_data="income_source_Được cho")],
            [InlineKeyboardButton("Thu hồi nợ", callback_data="income_source_Thu hồi nợ"), InlineKeyboardButton("Bán tài sản", callback_data="income_source_Bán tài sản")],
            [InlineKeyboardButton("Lãi tiết kiệm", callback_data="income_source_Lãi tiết kiệm"), InlineKeyboardButton("Lãi đầu tư", callback_data="income_source_Lãi đầu tư")],
            [InlineKeyboardButton("Khác", callback_data="income_source_Khác")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Chọn nguồn thu nhập:", reply_markup=reply_markup)
        return CHOOSING_INCOME_SOURCE
    elif choice == "date_other_income":
        await query.edit_message_text("Vui lòng nhập ngày của khoản thu nhập (VD: 25/06/2024 hoặc 15/1):")
        return CHOOSING_INCOME_DATE
    return await cancel_conversation_callback(update, context)

async def get_income_date_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['created_at'] = parse_date(update.message.text)
        sources = ["Lương","Thưởng","Kinh doanh","Được cho","Thu hồi nợ","Bán tài sản","Lãi tiết kiệm","Lãi đầu tư","Khác"]
        keyboard = [
            [InlineKeyboardButton("Lương", callback_data="income_source_Lương"), InlineKeyboardButton("Thưởng", callback_data="income_source_Thưởng")],
            [InlineKeyboardButton("Kinh doanh", callback_data="income_source_Kinh doanh"), InlineKeyboardButton("Được cho", callback_data="income_source_Được cho")],
            [InlineKeyboardButton("Thu hồi nợ", callback_data="income_source_Thu hồi nợ"), InlineKeyboardButton("Bán tài sản", callback_data="income_source_Bán tài sản")],
            [InlineKeyboardButton("Lãi tiết kiệm", callback_data="income_source_Lãi tiết kiệm"), InlineKeyboardButton("Lãi đầu tư", callback_data="income_source_Lãi đầu tư")],
            [InlineKeyboardButton("Khác", callback_data="income_source_Khác")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Chọn nguồn thu nhập:", reply_markup=reply_markup)
        return CHOOSING_INCOME_SOURCE
    except ValueError as e:
        await update.message.reply_text(f"{e} Vui lòng nhập lại ngày:")
        return CHOOSING_INCOME_DATE

async def get_income_source(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    source = query.data.replace("income_source_", "")
    user_id = query.from_user.id
    user_name = query.from_user.full_name
    amount = context.user_data.get('amount')
    note = context.user_data.get('note')
    created_at = context.user_data.get('created_at')

    db_add_income(user_id, amount, source, note, created_at)
    await query.edit_message_text(f"✅ Đã lưu thu nhập: {amount:,.0f} đ - {source} ({note}) vào ngày {created_at.strftime('%d/%m/%Y')}")
    
    anonymized_user_name = anonymize_name(user_name)
    post_message = (
        f"🔔 *Giao dịch mới được ghi nhận*\n\n"
        f"**Loại:** Thu nhập\n"
        f"**Số tiền:** `{amount:,.0f} đ`\n"
        f"**Nguồn:** {source}\n"
        f"**Ghi chú:** {note}\n"
        f"**Ngày:** {created_at.strftime('%d/%m/%Y')}\n\n"
        f"👤 *Người dùng:* `{anonymized_user_name}`"
    )
    await post_to_channel(context, post_message)
    return await end_current_conversation_and_reset(update, context, "Thu nhập đã được ghi nhận thành công.")

async def get_expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        amount = parse_amount(update.message.text)
        if amount <= 0:
            await update.message.reply_text("Số tiền phải lớn hơn 0. Vui lòng nhập lại:")
            return CHOOSING_EXPENSE_AMOUNT
        if amount > LARGE_AMOUNT_THRESHOLD:
            context.user_data['amount_to_confirm'] = amount
            keyboard = [
                [InlineKeyboardButton("✅ Đúng vậy", callback_data="confirm_amount_yes")],
                [InlineKeyboardButton("❌ Nhập lại", callback_data="confirm_amount_no")],
                [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text(
                f"Bạn chắc chắn muốn ghi nhận khoản chi tiêu {amount:,.0f} đ chứ? Số tiền này khá lớn.",
                reply_markup=reply_markup
            )
            return CONFIRM_EXPENSE_AMOUNT
        else:
            context.user_data['amount'] = amount
            await update.message.reply_text("Tuyệt vời! Bây giờ, nhập ghi chú cho khoản chi này:")
            return CHOOSING_EXPENSE_NOTE
    except (ValueError, TypeError):
        await update.message.reply_text("Số tiền không hợp lệ. Vui lòng nhập lại (VD: 5tr2, 100k):")
        return CHOOSING_EXPENSE_AMOUNT

async def confirm_expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_amount_yes":
        context.user_data['amount'] = context.user_data.pop('amount_to_confirm')
        await query.edit_message_text("Đã xác nhận. Bây giờ, nhập ghi chú cho khoản chi này:")
        return CHOOSING_EXPENSE_NOTE
    elif query.data == "confirm_amount_no":
        context.user_data.pop('amount_to_confirm', None)
        await query.edit_message_text("Vui lòng nhập lại số tiền chi tiêu:")
        return CHOOSING_EXPENSE_AMOUNT
    return await cancel_conversation_callback(update, context)

async def get_expense_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['note'] = update.message.text
    if 'created_at' in context.user_data: # If date was set (today)
        categories = ['Ăn uống', 'Di chuyển', 'Mua sắm', 'Hóa đơn', 'Gia đình', 'Sức khỏe', 'Giải trí', 'Đầu tư', 'Tiết kiệm', 'Khác']
        keyboard = [
            [InlineKeyboardButton("Ăn uống", callback_data="expense_category_Ăn uống"), InlineKeyboardButton("Di chuyển", callback_data="expense_category_Di chuyển")],
            [InlineKeyboardButton("Mua sắm", callback_data="expense_category_Mua sắm"), InlineKeyboardButton("Hóa đơn", callback_data="expense_category_Hóa đơn")],
            [InlineKeyboardButton("Gia đình", callback_data="expense_category_Gia đình"), InlineKeyboardButton("Sức khỏe", callback_data="expense_category_Sức khỏe")],
            [InlineKeyboardButton("Giải trí", callback_data="expense_category_Giải trí"), InlineKeyboardButton("Đầu tư", callback_data="expense_category_Đầu tư")],
            [InlineKeyboardButton("Tiết kiệm", callback_data="expense_category_Tiết kiệm"), InlineKeyboardButton("Khác", callback_data="expense_category_Khác")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Chọn danh mục chi tiêu:", reply_markup=reply_markup)
        return CHOOSING_EXPENSE_CATEGORY
    else: # Need to ask for date
        keyboard = [
            [InlineKeyboardButton("Hôm nay", callback_data="date_today_expense")],
            [InlineKeyboardButton("Ngày khác", callback_data="date_other_expense")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Giao dịch này xảy ra vào khi nào?", reply_markup=reply_markup)
        return CHOOSING_EXPENSE_DATE

async def choose_expense_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data
    if choice == "date_today_expense":
        context.user_data['created_at'] = datetime.datetime.now()
        categories = ['Ăn uống', 'Di chuyển', 'Mua sắm', 'Hóa đơn', 'Gia đình', 'Sức khỏe', 'Giải trí', 'Đầu tư', 'Tiết kiệm', 'Khác']
        keyboard = [
            [InlineKeyboardButton("Ăn uống", callback_data="expense_category_Ăn uống"), InlineKeyboardButton("Di chuyển", callback_data="expense_category_Di chuyển")],
            [InlineKeyboardButton("Mua sắm", callback_data="expense_category_Mua sắm"), InlineKeyboardButton("Hóa đơn", callback_data="expense_category_Hóa đơn")],
            [InlineKeyboardButton("Gia đình", callback_data="expense_category_Gia đình"), InlineKeyboardButton("Sức khỏe", callback_data="expense_category_Sức khỏe")],
            [InlineKeyboardButton("Giải trí", callback_data="expense_category_Giải trí"), InlineKeyboardButton("Đầu tư", callback_data="expense_category_Đầu tư")],
            [InlineKeyboardButton("Tiết kiệm", callback_data="expense_category_Tiết kiệm"), InlineKeyboardButton("Khác", callback_data="expense_category_Khác")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Chọn danh mục chi tiêu:", reply_markup=reply_markup)
        return CHOOSING_EXPENSE_CATEGORY
    elif choice == "date_other_expense":
        await query.edit_message_text("Vui lòng nhập ngày của khoản chi tiêu (VD: 25/06/2024 hoặc 15/1):")
        return CHOOSING_EXPENSE_DATE
    return await cancel_conversation_callback(update, context)

async def get_expense_date_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['created_at'] = parse_date(update.message.text)
        categories = ['Ăn uống', 'Di chuyển', 'Mua sắm', 'Hóa đơn', 'Gia đình', 'Sức khỏe', 'Giải trí', 'Đầu tư', 'Tiết kiệm', 'Khác']
        keyboard = [
            [InlineKeyboardButton("Ăn uống", callback_data="expense_category_Ăn uống"), InlineKeyboardButton("Di chuyển", callback_data="expense_category_Di chuyển")],
            [InlineKeyboardButton("Mua sắm", callback_data="expense_category_Mua sắm"), InlineKeyboardButton("Hóa đơn", callback_data="expense_category_Hóa đơn")],
            [InlineKeyboardButton("Gia đình", callback_data="expense_category_Gia đình"), InlineKeyboardButton("Sức khỏe", callback_data="expense_category_Sức khỏe")],
            [InlineKeyboardButton("Giải trí", callback_data="expense_category_Giải trí"), InlineKeyboardButton("Đầu tư", callback_data="expense_category_Đầu tư")],
            [InlineKeyboardButton("Tiết kiệm", callback_data="expense_category_Tiết kiệm"), InlineKeyboardButton("Khác", callback_data="expense_category_Khác")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Chọn danh mục chi tiêu:", reply_markup=reply_markup)
        return CHOOSING_EXPENSE_CATEGORY
    except ValueError as e:
        await update.message.reply_text(f"{e} Vui lòng nhập lại ngày:")
        return CHOOSING_EXPENSE_DATE

async def get_expense_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    category = query.data.replace("expense_category_", "")
    user_id = query.from_user.id
    user_name = query.from_user.full_name
    amount = context.user_data.get('amount')
    note = context.user_data.get('note')
    created_at = context.user_data.get('created_at')

    db_add_expense(user_id, amount, note, category, created_at)
    await query.edit_message_text(f"✅ Đã lưu chi tiêu: {amount:,.0f} đ - {note} ({category}) vào ngày {created_at.strftime('%d/%m/%Y')}")
    
    anonymized_user_name = anonymize_name(user_name)
    post_message = (
        f"🔔 *Giao dịch mới được ghi nhận*\n\n"
        f"**Loại:** Chi tiêu\n"
        f"**Số tiền:** `{amount:,.0f} đ`\n"
        f"**Danh mục:** {category}\n"
        f"**Ghi chú:** {note}\n"
        f"**Ngày:** {created_at.strftime('%d/%m/%Y')}\n\n"
        f"👤 *Người dùng:* `{anonymized_user_name}`"
    )
    await post_to_channel(context, post_message)
    return await end_current_conversation_and_reset(update, context, "Chi tiêu đã được ghi nhận thành công.")

async def handle_delete_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Ensure current conversation is ended cleanly before starting new one
    context.user_data.clear()
    keyboard = [
        [InlineKeyboardButton("Xóa giao dịch", callback_data="delete_action_individual")],
        [InlineKeyboardButton("Đặt lại dữ liệu", callback_data="delete_action_reset_all")],
        [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text("Bạn muốn thực hiện thao tác nào?", reply_markup=reply_markup)
    else:
        await update.message.reply_text("Bạn muốn thực hiện thao tác nào?", reply_markup=reply_markup)
    return DELETE_CHOOSING_ACTION

async def choose_delete_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data
    if choice == "delete_action_individual":
        return await delete_start_individual(update, context)
    elif choice == "delete_action_reset_all":
        keyboard = [
            [InlineKeyboardButton("✅ Có, tôi chắc chắn", callback_data="reset_confirm_1_yes")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            "Bạn có chắc chắn muốn xóa TẤT CẢ dữ liệu tài chính của mình không? "
            "Hành động này không thể hoàn tác.",
            reply_markup=reply_markup
        )
        return RESET_DATA_CONFIRM_STEP1
    return await cancel_conversation_callback(update, context)

async def delete_start_individual(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    transactions = db_get_last_n_transactions(user_id, limit=5)
    if not transactions:
        message_text = "Bạn chưa có giao dịch nào để xóa."
        return await end_current_conversation_and_reset(update, context, message_text)
        
    message_text = "Dưới đây là 5 giao dịch gần nhất của bạn:\n\n"
    keyboard = []
    for i, t in enumerate(transactions):
        trans_id, trans_type_str, amount, note, cat_or_src, created_at_str = t
        trans_type = "Chi tiêu" if trans_type_str == "expense" else "Thu nhập"
        created_at = datetime.datetime.fromisoformat(created_at_str)
        message_text += f"*{i+1}.* [{trans_type}] {amount:,.0f} đ - `{note}` ({cat_or_src}) - {created_at.strftime('%d/%m/%Y')}\n"
        keyboard.append([InlineKeyboardButton(f"Xóa mục {i+1}", callback_data=f"delete_id_{trans_id}_{trans_type_str}")])
    keyboard.append([InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(message_text, reply_markup=reply_markup, parse_mode='Markdown')
    else:
        await update.message.reply_text(message_text, reply_markup=reply_markup, parse_mode='Markdown')
    return DELETE_INDIVIDUAL_LISTING

async def delete_choose_item(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    parts = query.data.split('_')
    trans_id = int(parts[2])
    trans_type = parts[3]
    context.user_data['id_to_delete'] = trans_id
    context.user_data['type_to_delete'] = trans_type
    keyboard = [[InlineKeyboardButton("✅ Xác nhận XÓA", callback_data="confirm_delete_yes")], [InlineKeyboardButton("❌ Hủy", callback_data="cancel_action")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(f"Bạn có chắc chắn muốn xóa giao dịch ID {trans_id} ({'Thu nhập' if trans_type == 'income' else 'Chi tiêu'}) này không?", reply_markup=reply_markup)
    return DELETE_INDIVIDUAL_CONFIRM

async def delete_confirm_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_delete_yes":
        trans_id = context.user_data.get('id_to_delete')
        trans_type = context.user_data.get('type_to_delete')
        db_delete_transaction(trans_id, trans_type)
        return await end_current_conversation_and_reset(update, context, f"✅ Đã xóa thành công giao dịch ID {trans_id}.")
    else:
        return await end_current_conversation_and_reset(update, context, "Đã hủy bỏ thao tác xóa.")

async def reset_data_confirmation_1_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "reset_confirm_1_yes":
        keyboard = [
            [InlineKeyboardButton("🔥 Xóa TẤT CẢ dữ liệu 🔥", callback_data="reset_confirm_2_yes")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            "Đây là bước xác nhận cuối cùng. Bạn có THỰC SỰ muốn xóa tất cả dữ liệu? "
            "Sau khi xóa, dữ liệu sẽ không thể khôi phục.",
            reply_markup=reply_markup
        )
        return RESET_DATA_CONFIRM_STEP2
    return await cancel_conversation_callback(update, context)

async def reset_data_confirmation_2_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "reset_confirm_2_yes":
        user_id = query.from_user.id
        db_reset_all_data(user_id)
        return await end_current_conversation_and_reset(update, context, "✅ Đã xóa toàn bộ dữ liệu tài chính của bạn thành công.")
    else:
        return await end_current_conversation_and_reset(update, context, "Đã hủy bỏ thao tác đặt lại dữ liệu.")

def get_period_dates(period_choice: str):
    today = datetime.datetime.now()
    start_date, end_date, title_suffix = None, None, ""
    if period_choice == "week":
        start_of_week = today - datetime.timedelta(days=today.weekday())
        start_date = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_of_week + datetime.timedelta(days=6, hours=23, minutes=59, seconds=59)
        title_suffix = f"tuần này ({start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m')})"
    elif period_choice == "month":
        start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = start_date.replace(day=28) + datetime.timedelta(days=4)  # Đảm bảo chuyển sang tháng sau
        end_date = (next_month - datetime.timedelta(days=next_month.day)).replace(hour=23, minute=59, second=59)
        title_suffix = f"tháng này ({today.strftime('%m/%Y')})"
    elif period_choice == "year":
        start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(month=12, day=31, hour=23, minute=59, second=59)
        title_suffix = f"năm nay ({today.year})"
    elif period_choice == "all":
        title_suffix = "tất cả thời gian"
    else:
        return None, None, None
    return start_date, end_date, title_suffix

async def report_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Khởi tạo quy trình tạo báo cáo - cho phép người dùng chọn kỳ báo cáo"""
    # Đảm bảo cuộc hội thoại hiện tại được kết thúc sạch sẽ trước khi bắt đầu cuộc hội thoại mới
    context.user_data.clear()
    
    keyboard = [
        [InlineKeyboardButton("📅 Theo tuần", callback_data="report_period_week")],
        [InlineKeyboardButton("📆 Theo tháng", callback_data="report_period_month")],
        [InlineKeyboardButton("🗓️ Theo năm", callback_data="report_period_year")],
        [InlineKeyboardButton("⏰ Toàn bộ thời gian", callback_data="report_period_all")],
        [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    welcome_message = (
        "📊 *CHỌN KỲ BÁO CÁO*\n\n"
        "Vui lòng chọn khoảng thời gian bạn muốn xem báo cáo tài chính:"
    )
    
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(
            welcome_message, 
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
    else:
        await update.message.reply_text(
            welcome_message, 
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
    
    return CHOOSING_REPORT_PERIOD

async def generate_report_for_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Tạo và hiển thị báo cáo tài chính theo kỳ đã chọn"""
    query = update.callback_query
    await query.answer()
    
    # Xử lý hủy bỏ
    if query.data == "cancel_action":
        return await cancel_conversation_callback(update, context)
    
    # Lấy thông tin kỳ báo cáo
    period_choice = query.data.replace("report_period_", "")
    user_id = query.from_user.id
    
    # Hiển thị thông báo đang xử lý
    loading_message = "⏳ *Đang tạo báo cáo...*\n\nVui lòng chờ trong giây lát..."
    message_to_delete = await query.edit_message_text(
        loading_message, 
        parse_mode='Markdown'
    )
    
    # Lấy dữ liệu từ database
    start_date, end_date, title_suffix = get_period_dates(period_choice)
    income, expense, balance = db_get_combined_summary(user_id, start_date, end_date)
    incomes_grouped = db_list_incomes_grouped(user_id, start_date, end_date)
    expenses_grouped = db_list_expenses_grouped(user_id, start_date, end_date)
    
    # Tạo header báo cáo
    report_header = (
        f"📋 *BÁO CÁO TÀI CHÍNH {title_suffix.upper()}*\n"
        f"{'=' * 35}\n\n"
    )
    
    # Tạo phần tổng quan
    summary_section = (
        f"💼 *TỔNG QUAN TÀI CHÍNH*\n"
        f"┏━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"┃ 📈 Tổng thu nhập: `{income:,.0f} VND`\n"
        f"┃ 📉 Tổng chi tiêu: `{expense:,.0f} VND`\n"
        f"┃ 💰 Số dư còn lại: `{balance:,.0f} VND`\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    )
    
    # Tạo phần chi tiết thu nhập
    income_section = ""
    if incomes_grouped:
        income_section += "💵 *CHI TIẾT THU NHẬP*\n"
        for i, (source, total) in enumerate(incomes_grouped, 1):
            percentage = (total / income * 100) if income > 0 else 0
            income_section += f"{i}. **{source}**: `{total:,.0f} VND` ({percentage:.1f}%)\n"
        income_section += "\n"
    else:
        income_section += "⚠️ *Chưa có thu nhập nào trong kỳ này*\n\n"
    
    # Tạo phần chi tiết chi tiêu
    expense_section = ""
    if expenses_grouped:
        expense_section += "💸 *CHI TIẾT CHI TIÊU*\n"
        for i, (category, total) in enumerate(expenses_grouped, 1):
            percentage = (total / expense * 100) if expense > 0 else 0
            expense_section += f"{i}. **{category}**: `{total:,.0f} VND` ({percentage:.1f}%)\n"
        expense_section += "\n"
    else:
        expense_section += "✅ *Không có chi tiêu nào trong kỳ này*\n\n"
    

    # === CẢNH BÁO NGÂN SÁCH (chỉ hiển thị trong báo cáo THÁNG) ===
    budget_alerts = ""
    budgets = db_get_budgets(user_id)
    if period_choice == "month" and budgets and expenses_grouped:
        budget_alerts += "⚠️ *CẢNH BÁO NGÂN SÁCH*\n"
        for i, (category, total) in enumerate(expenses_grouped, 1):
            limit_amt = budgets.get(category)
            if not limit_amt:
                continue
            ratio = total / limit_amt if limit_amt > 0 else 0
            if ratio >= 1.0:
                budget_alerts += f"- {category}: `{total:,.0f}/{limit_amt:,.0f}` đ  🚨 *VƯỢT*\n"
            elif ratio >= 0.9:
                budget_alerts += f"- {category}: `{total:,.0f}/{limit_amt:,.0f}` đ  ⚠️ *GẦN*\n"
        if budget_alerts.strip() == "⚠️ *CẢNH BÁO NGÂN SÁCH*":
            budget_alerts += "✔️ Tất cả danh mục trong hạn mức.\n"
        budget_alerts += "\n"
    else:
        budget_alerts = ""
    # Tạo phần đánh giá tình hình tài chính
    financial_status = ""
    if balance > 0:
        financial_status = "📊 *ĐÁNH GIÁ*: Tình hình tài chính tích cực! 🎉"
    elif balance == 0:
        financial_status = "📊 *ĐÁNH GIÁ*: Thu chi cân bằng ⚖️"
    else:
        financial_status = "📊 *ĐÁNH GIÁ*: Cần cân nhắc chi tiêu ⚠️"
    
    # Ghép tất cả các phần lại
    report_message = (
        report_header + summary_section + income_section + expense_section + budget_alerts + financial_status
    )
    
    # Xóa thông báo loading và gửi báo cáo
    await message_to_delete.delete()
    await query.message.reply_text(report_message, parse_mode='Markdown')
    
    return await end_current_conversation_and_reset(
        update, 
        context, 
        "✅ Báo cáo đã được tạo thành công! Cảm ơn bạn đã sử dụng dịch vụ."
    )

async def chart_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Ensure current conversation is ended cleanly before starting new one
    context.user_data.clear()
    keyboard = [
        [InlineKeyboardButton("Theo tuần", callback_data="chart_period_week")],
        [InlineKeyboardButton("Theo tháng", callback_data="chart_period_month")],
        [InlineKeyboardButton("Theo năm", callback_data="chart_period_year")],
        [InlineKeyboardButton("Toàn bộ thời gian", callback_data="chart_period_all")],
        [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text("Bạn muốn xem biểu đồ theo kỳ nào?", reply_markup=reply_markup)
    else:
        await update.message.reply_text("Bạn muốn xem biểu đồ theo kỳ nào?", reply_markup=reply_markup)
    return CHOOSING_CHART_PERIOD
async def generate_charts_for_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel_action":
        return await cancel_conversation_callback(update, context)

    period_choice = query.data.replace("chart_period_", "")
    user_id = query.from_user.id
    start_date, end_date, title_suffix = get_period_dates(period_choice)
    
    message_to_delete = await query.edit_message_text(f"Đang tạo biểu đồ cho {title_suffix}, vui lòng chờ...")
    
    has_sent_chart = False

    # === PHẦN TẠO BIỂU ĐỒ DONUT MỚI ===
    expenses_data = db_list_expenses_grouped(user_id, start_date, end_date)
    if expenses_data:
        has_sent_chart = True
        labels, sizes = [r[0] for r in expenses_data], [r[1] for r in expenses_data]
        
        # Nếu tổng chi tiêu bằng 0, không vẽ biểu đồ
        if sum(sizes) == 0:
            await query.message.reply_text(f"Không có chi tiêu nào trong {title_suffix} để tạo biểu đồ.")
            plt.close()
            # Xóa tin nhắn "Đang tạo biểu đồ..."
            await message_to_delete.delete()
            return await end_current_conversation_and_reset(update, context, "Hoàn tất.")
            
        fig, ax = plt.subplots(figsize=(8, 8))
        
        # Cấu hình biểu đồ Donut
        wedgeprops = {'width': 0.4, 'edgecolor': 'white', 'linewidth': 2}
        
        # Bảng màu trẻ trung, hiện đại (bạn có thể tùy chỉnh)
        colors = ['#FF7A7A', '#3498DB', '#F1C40F', '#27AE60', '#E74C3C', '#689F38', '#F39C12', '#2980B9']
        
        # Vẽ biểu đồ donut
        patches, texts, autotexts = ax.pie(sizes,
                                           colors=colors,
                                           startangle=90,
                                           counterclock=False,
                                           wedgeprops=wedgeprops,
                                           autopct="%1.1f%%",
                                           pctdistance=0.85)

        # Chỉnh sửa autopct để làm nổi bật phần trăm
        for autotext in autotexts:
            autotext.set_color('black')
            autotext.set_fontsize(20)
            autotext.set_fontweight('bold')
        
        # Cấu hình phần trung tâm của biểu đồ
        center_circle = plt.Circle((0, 0), 0.6, fc='white')
        fig.gca().add_artist(center_circle)
        # Thêm chữ vào giữa biểu đồ (tọa độ (0,0) là tâm)
        plt.text(0, 0, '@quanlythuchixx_bot', horizontalalignment='center', verticalalignment='center', fontsize=12, color='black', alpha=0.3)

        # Tạo legend tùy chỉnh
        # Tạo label string bao gồm tên và phần trăm
        legend_labels = [f'{label} ({size/sum(sizes)*100:.1f}%)' for label, size in zip(labels, sizes)]
        ax.legend(patches, legend_labels, loc="center left", bbox_to_anchor=(1, 0.5), fontsize=12)

        ax.axis("equal")
        plt.title(f"Phân bổ Chi tiêu ({title_suffix})", fontsize=16, fontweight='bold', pad=20)
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight', dpi=150, 
                   facecolor='white', edgecolor='none')
        buf.seek(0)
        
        # Caption chi tiết cho biểu đồ donut
        total_expense = sum(sizes)
        caption = f"🍩 Phân bổ chi tiêu {title_suffix}\n"
        caption += f"💸 Tổng chi tiêu: {total_expense:,} VND\n\n"
        caption += "📊 Chi tiết từng danh mục:\n"
        
        # Sắp xếp theo số tiền giảm dần
        sorted_data = sorted(zip(labels, sizes), key=lambda x: x[1], reverse=True)
        
        for i, (label, amount) in enumerate(sorted_data[:5], 1):  # Chỉ hiển thị top 5
            percentage = (amount / total_expense) * 100
            caption += f"{i}. {label}: {amount:,} VND ({percentage:.1f}%)\n"
        
        if len(sorted_data) > 5:
            remaining_amount = sum(amount for _, amount in sorted_data[5:])
            remaining_percentage = (remaining_amount / total_expense) * 100
            caption += f"... và {len(sorted_data)-5} danh mục khác: {remaining_amount:,} VND ({remaining_percentage:.1f}%)\n"
        
        # Tìm danh mục chi tiêu nhiều nhất
        top_category = sorted_data[0][0]
        top_amount = sorted_data[0][1]
        top_percentage = (top_amount / total_expense) * 100
        
        caption += f"\n🔥 Chi tiêu nhiều nhất: {top_category} ({top_percentage:.1f}%)"
        
        await query.message.reply_photo(photo=buf, caption=caption)
        plt.close(fig)
    
    # === PHẦN BIỂU ĐỒ CỘT CHUYÊN NGHIỆP ===
    expenses, incomes = db_get_monthly_report(user_id, start_date, end_date)
    if expenses or incomes:
        has_sent_chart = True
        months_expenses = [e[0] for e in expenses]
        months_incomes = [i[0] for i in incomes]
        all_months = sorted(list(set(months_expenses + months_incomes)))
        
        exp_dict = {month: amount for month, amount in expenses}
        inc_dict = {month: amount for month, amount in incomes}
        
        exp_values = [exp_dict.get(m, 0) for m in all_months]
        inc_values = [inc_dict.get(m, 0) for m in all_months]
        x = np.arange(len(all_months))
        
        # Thiết lập figure chuyên nghiệp hơn
        fig, ax = plt.subplots(figsize=(12, 7))
        fig.patch.set_facecolor('white')
        
        bar_width = 0.4
        
        # Màu sắc chuyên nghiệp
        income_color = '#2E8B57'  # Xanh lá đậm
        expense_color = '#DC143C'  # Đỏ đậm
        
        # Tạo bars với edge và alpha
        bars_inc = ax.bar(x - bar_width/2, inc_values, width=bar_width, 
                         label='Thu nhập', color=income_color, alpha=0.8,
                         edgecolor='white', linewidth=1)
        bars_exp = ax.bar(x + bar_width/2, exp_values, width=bar_width, 
                         label='Chi tiêu', color=expense_color, alpha=0.8,
                         edgecolor='white', linewidth=1)
        
        # Cải thiện axes
        ax.set_xticks(x)
        ax.set_xticklabels(all_months, rotation=45, ha='right', fontsize=11, fontweight='medium')
        ax.set_ylabel("Số tiền (VND)", fontsize=12, fontweight='bold')
        ax.set_title(f"Báo Cáo Thu Chi Tài Chính ({title_suffix})", 
                    fontsize=14, fontweight='bold', pad=20)
        
        # Grid tinh tế
        ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
        ax.set_axisbelow(True)
        
        # Legend đẹp hơn
        legend = ax.legend(loc='upper left', frameon=True, shadow=True, 
                          fancybox=True, fontsize=11)
        legend.get_frame().set_facecolor('white')
        legend.get_frame().set_alpha(0.9)
        
        # Thêm values trên bars
        for bar in bars_inc:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width()/2, height + max(max(inc_values), max(exp_values)) * 0.01,
                       f'{int(height):,}', ha='center', va='bottom', 
                       fontsize=9, fontweight='bold', color=income_color)
        
        for bar in bars_exp:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width()/2, height + max(max(inc_values), max(exp_values)) * 0.01,
                       f'{int(height):,}', ha='center', va='bottom', 
                       fontsize=9, fontweight='bold', color=expense_color)
        
        # Thêm thông tin tổng kết trong box
        total_income = sum(inc_values)
        total_expense = sum(exp_values)
        net_balance = total_income - total_expense
        
        info_text = f"Tổng thu: {total_income:,} VND\nTổng chi: {total_expense:,} VND\nCân đối: {net_balance:,} VND"
        bbox_color = '#E8F5E8' if net_balance >= 0 else '#FFE8E8'
        text_color = '#2E8B57' if net_balance >= 0 else '#DC143C'
        
        ax.text(0.98, 0.98, info_text, transform=ax.transAxes, fontsize=10,
        verticalalignment='top', horizontalalignment='right', # Đã sửa tại đây
        bbox=dict(boxstyle='round,pad=0.5', facecolor=bbox_color, alpha=0.8),
        color=text_color, fontweight='bold')
        
        # Watermark tinh tế
        ax.text(0.98, 0.02, '@quanlythuchixx_bot', 
               transform=ax.transAxes, ha='right', va='bottom',
               fontsize=9, color='gray', alpha=0.5, style='italic')
        
        plt.tight_layout()
        plt.subplots_adjust(bottom=0.15)
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight', dpi=150, 
                   facecolor='white', edgecolor='none')
        buf.seek(0)
        
        # Caption chi tiết hơn
        caption = f"📊 Báo cáo tài chính {title_suffix}\n"
        caption += f"💰 Tổng thu nhập: {total_income:,} VND\n"
        caption += f"🏷️ Tổng chi tiêu: {total_expense:,} VND\n"
        if net_balance >= 0:
            caption += f"✅ Tiết kiệm: {net_balance:,} VND"
        else:
            caption += f"⚠️ Thâm hụt: {abs(net_balance):,} VND"
        
        await query.message.reply_photo(photo=buf, caption=caption)
        plt.close(fig)

    # === 3) HEATMAP Chi tiêu theo Danh mục theo kỳ (TUẦN/THÁNG/NĂM) ===
    if period_choice in ("week", "month", "year"):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Xác định phạm vi thời gian nếu chưa truyền vào
        today = datetime.date.today()

        if period_choice == "week":
            # mặc định: tuần hiện tại (Thứ 2..CN theo ISO)
            if not (start_date and end_date):
                start_date = today - datetime.timedelta(days=today.weekday())
                end_date = start_date + datetime.timedelta(days=6)
            # nhóm theo ngày
            group_expr = "strftime('%Y-%m-%d', created_at)"
            x_label_name = "day"
            title = "Heatmap Chi Tiêu Theo Danh Mục & Ngày (Tuần này)"
            caption = "🔥 Heatmap chi tiêu theo danh mục & ngày (tuần)"

        elif period_choice == "month":
            # mặc định: tháng hiện tại
            if not (start_date and end_date):
                start_date = today.replace(day=1)
                # ngày cuối tháng
                next_month = (start_date.replace(day=28) + datetime.timedelta(days=4)).replace(day=1)
                end_date = next_month - datetime.timedelta(days=1)
            # nhóm theo ngày
            group_expr = "strftime('%Y-%m-%d', created_at)"
            x_label_name = "day"
            title = "Heatmap Chi Tiêu Theo Danh Mục & Ngày (Tháng này)"
            caption = "🔥 Heatmap chi tiêu theo danh mục & ngày (tháng)"

        else:  # year
            # mặc định: năm hiện tại
            if not (start_date and end_date):
                start_date = today.replace(month=1, day=1)
                end_date = today.replace(month=12, day=31)
            # nhóm theo tháng
            group_expr = "strftime('%Y-%m', created_at)"
            x_label_name = "month"
            title = "Heatmap Chi Tiêu Theo Danh Mục & Tháng (Năm nay)"
            caption = "🔥 Heatmap chi tiêu theo danh mục & tháng (năm)"

        params = [user_id]
        sql = f"""
            SELECT {group_expr} AS x_val,
                   category,
                   SUM(amount) AS total
            FROM expenses
            WHERE user_id = ?
        """

        if start_date and end_date:
            # end_date < next day (bao gồm cả end_date)
            end_date_inclusive = end_date + datetime.timedelta(days=1)
            sql += " AND created_at >= ? AND created_at < ?"
            params.extend([start_date.isoformat(), end_date_inclusive.isoformat()])

        sql += " GROUP BY x_val, category ORDER BY x_val;"
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()
        conn.close()

        if rows:
            has_sent_chart = True

            # Tạo dải nhãn X (liên tục theo ngày hoặc theo tháng)
            if x_label_name == "day":
                # danh sách ngày liên tục từ start_date..end_date
                days = []
                cur = start_date
                while cur <= end_date:
                    days.append(_fmt_day_label(cur))
                    cur += datetime.timedelta(days=1)
                x_vals = days
            else:
                # danh sách tháng liên tục start_date..end_date (YYYY-MM)
                x_vals = []
                y, m = start_date.year, start_date.month
                y_end, m_end = end_date.year, end_date.month
                while (y < y_end) or (y == y_end and m <= m_end):
                    x_vals.append(f"{y:04d}-{m:02d}")
                    # tăng tháng
                    if m == 12:
                        m, y = 1, y + 1
                    else:
                        m += 1

            categories = sorted(list({r[1] for r in rows}))

            # Map dữ liệu vào ma trận
            data_map = {(x, c): 0 for x in x_vals for c in categories}
            for x, c, v in rows:
                if x in x_vals:  # chỉ nhận các điểm nằm trong range đã chuẩn hóa
                    data_map[(x, c)] = v
            heat = np.array([[data_map[(x, c)] for c in categories] for x in x_vals], dtype=float)

            # Vẽ heatmap
            fig, ax = plt.subplots(figsize=(12, 6))
            # Colormap sáng: 0 -> xám nhạt -> cam -> đỏ
            cmap = mcolors.LinearSegmentedColormap.from_list(
                'custom_heat',
                [
                    (0.0, '#f0f0f0'),  # xám rất nhạt cho 0
                    (0.5, '#ffcc66'),  # cam nhạt
                    (1.0, '#ff3300')   # đỏ đậm
                ]
            )
            norm = mcolors.Normalize(vmin=0, vmax=heat.max() if heat.size else 1)
            im = ax.imshow(heat, aspect='auto', cmap=cmap, norm=norm)

            # Thanh màu chú giải
            cbar = plt.colorbar(im, ax=ax)
            cbar.set_label('Số tiền (VND)')

            # Nhãn trục
            ax.set_xticks(np.arange(len(categories)))
            ax.set_xticklabels(categories, rotation=45, ha='right', fontsize=9)
            ax.set_yticks(np.arange(len(x_vals)))
            ax.set_yticklabels(x_vals, fontsize=9)

            # Tiêu đề động
            ax.set_title(title, fontsize=14, fontweight='bold', pad=12)

            # Ghi giá trị lên ô (nếu > 0)
            maxv = heat.max() if heat.size else 0
            for i in range(len(x_vals)):
                for j in range(len(categories)):
                    val = heat[i, j]
                    if val > 0:
                        txt_color = 'white' if maxv > 0 and val > 0.6 * maxv else 'black'
                        ax.text(j, i, f"{int(val):,}", ha='center', va='center',
                                fontsize=8, color=txt_color)

            # Watermark
            ax.text(len(categories) - 0.5, len(x_vals) - 0.5, '@quanlythuchixx_bot',
                    ha='right', va='bottom', fontsize=9, color='gray', alpha=0.5, style='italic')

            # Xuất ảnh
            plt.tight_layout()
            buf = io.BytesIO()
            plt.savefig(buf, format='png', bbox_inches='tight', dpi=150,
                        facecolor='white', edgecolor='none')
            buf.seek(0)

            await query.message.reply_photo(
                photo=buf,
                caption=caption
            )
            plt.close(fig)

async def export_file_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Ensure current conversation is ended cleanly before starting new one
    context.user_data.clear()
    df = db_export_data(update.effective_user.id)
    if not df.empty:
        keyboard = [
            [InlineKeyboardButton("Xuất ra CSV", callback_data="export_csv")],
            [InlineKeyboardButton("Xuất ra Excel", callback_data="export_excel")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        if update.callback_query:
            await update.callback_query.answer()
            await update.callback_query.edit_message_text("Bạn muốn xuất dữ liệu ra định dạng nào?", reply_markup=reply_markup)
        else:
            await update.message.reply_text("Bạn muốn xuất dữ liệu ra định dạng nào?", reply_markup=reply_markup)
        return EXPORT_CHOOSING_FORMAT
    else:
        return await end_current_conversation_and_reset(update, context, "⚠️ Bạn chưa có dữ liệu nào để xuất.")

async def handle_export_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel_action":
        return await cancel_conversation_callback(update, context)
    
    df = db_export_data(query.from_user.id)
    if df.empty: # Should not happen if export_file_start worked, but good for safety
        return await end_current_conversation_and_reset(update, context, "⚠️ Không có dữ liệu để xuất file.")
    
    message_to_edit = await query.edit_message_text("Đang chuẩn bị file, vui lòng chờ...")
    file_name_prefix = f'finance_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}'
    
    if query.data == "export_csv":
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8-sig') 
        await query.message.reply_document(document=csv_buffer.getvalue().encode('utf-8-sig'), filename=f'{file_name_prefix}.csv', caption="📁 Dữ liệu của bạn.")
    elif query.data == "export_excel":
        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False, engine='xlsxwriter')
        await query.message.reply_document(document=excel_buffer.getvalue(), filename=f'{file_name_prefix}.xlsx', caption="📁 Dữ liệu của bạn.")
    
    await message_to_edit.delete()
    return await end_current_conversation_and_reset(update, context, "Dữ liệu đã được xuất thành công.")

async def import_file_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    
    instruction_message = (
        "Bạn có thể nhập dữ liệu tài chính của mình bằng cách tải lên file Excel (.xlsx) hoặc CSV (.csv).\n\n"
        "**Định dạng file yêu cầu:**\n"
        "File của bạn phải có 5 cột với tên chính xác như sau (không phân biệt chữ hoa/thường):\n"
        "`Loại`, `Số tiền`, `Ghi chú`, `Danh mục/Nguồn`, `Ngày`\n\n"
        "**Ví dụ:**\n"
        "| Loại      | Số tiền  | Ghi chú              | Danh mục/Nguồn | Ngày         |\n"
        "|-----------|----------|----------------------|----------------|--------------|\n"
        "| Chi tiêu  | 50000    | Ăn trưa              | Ăn uống        | 2024-07-28   |\n"
        "| Thu nhập  | 10000000 | Lương tháng 7        | Lương          | 2024-07-31   |\n\n"
        "Vui lòng tải file của bạn lên ngay bây giờ."
    )
    
    keyboard = [[InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(instruction_message, reply_markup=reply_markup, parse_mode='Markdown')
    else:
        await update.message.reply_text(instruction_message, reply_markup=reply_markup, parse_mode='Markdown')

    return IMPORT_FILE_UPLOADED

async def handle_import_file_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    document = update.message.document
    
    if not (document.file_name.endswith('.csv') or document.file_name.endswith('.xlsx')):
        await update.message.reply_text("⚠️ Bot chỉ hỗ trợ file .csv hoặc .xlsx. Vui lòng thử lại với định dạng đúng.")
        return IMPORT_FILE_UPLOADED

    # Gửi tin nhắn đang xử lý
    await update.message.reply_text("Đang xử lý file của bạn, vui lòng chờ...")

    try:
        # Tải file về bộ nhớ
        file_bytes = await document.get_file()
        file_stream = io.BytesIO()
        await file_bytes.download_to_memory(file_stream)
        file_stream.seek(0)
        
        # Đọc file dựa trên định dạng
        if document.file_name.endswith('.csv'):
            df = pd.read_csv(file_stream)
        elif document.file_name.endswith('.xlsx'):
            df = pd.read_excel(file_stream)
        
        # Thêm dữ liệu vào DB
        success_count = db_add_transactions_from_df(user_id, df)
        
        income_count = success_count['income']
        expense_count = success_count['expense']
        total_count = income_count + expense_count
        
        message_text = (
            f"✅ **Nhập dữ liệu thành công!**\n"
            f"Đã thêm `{total_count}` giao dịch mới vào hệ thống:\n"
            f"- `{income_count}` khoản thu nhập\n"
            f"- `{expense_count}` khoản chi tiêu"
        )
        await update.message.reply_text(message_text, parse_mode='Markdown')
        
    except ValueError as e:
        await update.message.reply_text(f"❌ Lỗi: {e}. Vui lòng kiểm tra lại định dạng file.")
    except Exception as e:
        logger.error(f"Lỗi khi xử lý file từ user {user_id}: {e}")
        await update.message.reply_text(f"❌ Đã xảy ra lỗi không mong muốn khi xử lý file của bạn. Vui lòng thử lại sau.")

    return await end_current_conversation_and_reset(update, context, "Nhập dữ liệu đã hoàn tất.")

async def show_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await end_current_conversation_and_reset(
        update, context,
        '**HƯỚNG DẪN SỬ DỤNG BOT QUẢN LÝ THU – CHI**\n\n🎯 *Mục tiêu*\n• Ghi nhanh thu/chi hằng ngày\n• Theo dõi báo cáo theo tuần/tháng/năm\n• Cảnh báo ngân sách từng danh mục\n• Biểu đồ trực quan (có heatmap năm)\n\n⚡ *Nhập nhanh (khuyến nghị)*\nGõ tự nhiên, bot tự hiểu số tiền → loại → danh mục → ghi chú → ngày:\n- `20k ăn sáng` → Chi tiêu / **Ăn uống** / 20.000đ / hôm nay\n- `2.5tr tiền nhà hôm qua` → Chi tiêu / **Hóa đơn** / 2.500.000đ / hôm qua\n- `+500k lương` hoặc `thu 500k lương` → **Thu nhập** / 500.000đ / hôm nay\n- `900k grab 05/08` → Chi tiêu / **Di chuyển** / 900.000đ / 05/08/(năm nay)\n\n👉 Đơn vị: `k` = nghìn, `tr` = triệu (vd: `150k`, `3tr`, `1.2tr`…)\n\n➕ *Thêm giao dịch (qua menu)*\n- Thu nhập: số tiền → ghi chú → chọn **Nguồn** (Lương/Thưởng/…)\n- Chi tiêu: số tiền → ghi chú → chọn **Danh mục** (Ăn uống/Di chuyển/…)\n- Có tùy chọn nhập **ngày khác**.\n\n💰 *Ngân sách theo danh mục/tháng*\nĐặt hạn mức chi và nhận **cảnh báo** khi **≥90%** hoặc **vượt**.\n- Đặt: `/ngansach <Danh mục> <Số tiền>`  (vd: `/ngansach Ăn uống 3tr`)\n- Xem: `/xem_ngansach`\n- Xóa: `/xoa_ngansach <Danh mục>` (vd: `/xoa_ngansach Ăn uống`)\nKhi ghi chi tiêu, bot sẽ tự kiểm tra và báo **GẦN**/**VƯỢT** ngay.\n\n📈 *Báo cáo*\n`📈 Báo Cáo` → chọn **Tuần / Tháng / Năm / Toàn thời gian**\n- Tổng thu / Tổng chi / Cân đối\n- Chi tiết theo nguồn thu / danh mục chi\n- **⚠️ CẢNH BÁO NGÂN SÁCH** (nếu gần/vượt hạn mức)\n\n📊 *Biểu đồ*\n`📊 Biểu Đồ` → chọn **Tuần / Tháng / Năm / Toàn thời gian**\n- Donut phân bổ chi tiêu theo danh mục\n- Cột so sánh **Thu nhập vs Chi tiêu** theo tháng\n- **NĂM**: thêm **Heatmap** chi tiêu **(Danh mục × Tháng)** (màu sáng) + watermark `@quanlythuchixx_bot`\n\n📤📥 *Xuất / Nhập dữ liệu*\n- Xuất: CSV/Excel qua `📤 Xuất File`\n- Nhập: CSV/Excel qua `📥 Nhập File`\n  File cần 5 cột: `Loại`, `Số tiền`, `Ghi chú`, `Danh mục/Nguồn`, `Ngày`\n\n🗑️ *Quản lý dữ liệu*\n- Xóa giao dịch gần đây hoặc **đặt lại toàn bộ** (qua `🗑️ Xóa`).\n- **Cẩn thận** khi đặt lại: không thể hoàn tác.\n\n📌 *Mẹo & Quy ước*\n- Số tiền: chấp nhận `k`/`tr`, dấu `,`/`.` (vd: `1,200,000`, `1.2tr`)\n- Ngày: `hôm nay`, `hôm qua`, `hôm kia`, hoặc `dd/mm(/yyyy)`\n- Danh mục/nguồn chọn nhanh bằng **bàn phím 2 cột**.\n\n❓ *Lệnh nhanh*\n- `/ngansach Ăn uống 3tr` – đặt ngân sách\n- `/xem_ngansach` – xem ngân sách & % dùng\n- `/xoa_ngansach Ăn uống` – xóa ngân sách\n- `/baocao`, `/bieudo`, `/xuatfile`, `/nhapfile`, `/menu`, `/huongdan`, `/cancel`\n'
    )

async def quick_entry_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (update.message.text or "").strip()
    parsed = nlp_parse_transaction(text)
    if not parsed:
        return await end_current_conversation_and_reset(update, context, "Không nhận diện được giao dịch. Bạn dùng các nút bên dưới nhé:")
    user_id = update.effective_user.id
    if parsed["type"] == "income":
        source = parsed["category"] if parsed["category"] != "Khác" else "Khác"
        db_add_income(user_id, parsed["amount"], source, parsed["note"], parsed["created_at"])
        msg = (f"✅ Đã lưu *Thu nhập*: `{parsed['amount']:,.0f} đ`\n"
               f"• Nguồn: *{source}*\n• Ghi chú: _{parsed['note']}_\n• Ngày: {parsed['created_at'].strftime('%d/%m/%Y')}")
        await update.message.reply_text(msg, parse_mode='Markdown')
        return await end_current_conversation_and_reset(update, context, "Giao dịch đã ghi nhận. Đây là menu:")
    else:
        db_add_expense(user_id, parsed["amount"], parsed["note"], parsed["category"], parsed["created_at"])
        msg = (f"✅ Đã lưu *Chi tiêu*: `{parsed['amount']:,.0f} đ`\n"
               f"• Danh mục: *{parsed['category']}*\n• Ghi chú: _{parsed['note']}_\n• Ngày: {parsed['created_at'].strftime('%d/%m/%Y')}")
        budgets = db_get_budgets(user_id)
        budget_amt = budgets.get(parsed["category"])
        if budget_amt:
            m_start, m_end = get_month_range(parsed["created_at"])
            spent_now = db_sum_expense_by_category_in_period(user_id, parsed["category"], m_start, m_end)
            ratio = spent_now / budget_amt if budget_amt > 0 else 0
            remain = budget_amt - spent_now
            if ratio >= 1.0:
                msg += (f"\n\n⚠️ *VƯỢT NGÂN SÁCH THÁNG* {parsed['category']}\n"
                        f"• Đã chi: `{spent_now:,.0f} đ` / Hạn mức: `{budget_amt:,.0f} đ`\n"
                        f"• Vượt: `{abs(remain):,.0f} đ`")
            elif ratio >= 0.9:
                msg += (f"\n\n⚠️ *GẦN CHẠM NGÂN SÁCH* {parsed['category']}\n"
                        f"• Đã chi: `{spent_now:,.0f} đ` / Hạn mức: `{budget_amt:,.0f} đ` (≥90%)\n"
                        f"• Còn lại: `{remain:,.0f} đ`")
        await update.message.reply_text(msg, parse_mode='Markdown')
        return await end_current_conversation_and_reset(update, context, "Giao dịch đã ghi nhận. Đây là menu:")

async def budget_set_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /ngansach <Danh mục có thể có khoảng trắng> <Số tiền>
    Ví dụ:
      /ngansach Ăn uống 3tr
      /ngansach Mua sắm 1500000
    """
    try:
        raw = (update.message.text or "").strip()
        tokens = raw.split()
        if len(tokens) < 3:
            return await end_current_conversation_and_reset(update, context,
                "Cú pháp: /ngansach <Danh mục> <Số tiền>\nVD: /ngansach Ăn uống 3tr")

        amount_text = tokens[-1]
        category = " ".join(tokens[1:-1]).strip()

        amount = parse_amount(amount_text)
        if amount <= 0:
            raise ValueError

        db_set_budget(update.effective_user.id, category, amount)
        return await end_current_conversation_and_reset(update, context,
            f"✅ Đã đặt ngân sách *{category}*: `{amount:,.0f} đ`/tháng")
    except Exception:
        return await end_current_conversation_and_reset(update, context,
            "Không hiểu số tiền. Hãy thử: /ngansach Ăn uống 3tr")

async def budget_list_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    budgets = db_get_budgets(user_id)
    if not budgets:
        return await end_current_conversation_and_reset(update, context, "Bạn chưa đặt ngân sách nào. Dùng: /ngansach <Danh mục> <Số tiền>")
    now = datetime.datetime.now()
    m_start, m_end = get_month_range(now)
    lines = ["📋 *NGÂN SÁCH THÁNG NÀY*"]
    for cat, limit_amt in budgets.items():
        spent = db_sum_expense_by_category_in_period(user_id, cat, m_start, m_end)
        ratio = spent / limit_amt if limit_amt > 0 else 0
        status = "✅ OK"
        if ratio >= 1.0: status = "🚨 VƯỢT"
        elif ratio >= 0.9: status = "⚠️ GẦN"
        lines.append(f"- {cat}: {spent:,.0f}/{limit_amt:,.0f} đ  ({ratio*100:.0f}%)  {status}")
    msg = "\n".join(lines)
    return await end_current_conversation_and_reset(update, context, msg)

async def budget_delete_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = update.message.text.split(maxsplit=1)
    if len(args) < 2:
        return await end_current_conversation_and_reset(update, context, "Cú pháp: /xoa_ngansach <Danh mục>")
    category = args[1].strip()
    db_delete_budget(update.effective_user.id, category)
    return await end_current_conversation_and_reset(update, context, f"🗑️ Đã xóa ngân sách cho *{category}*")

def main() -> None:
    application = build_app()
    add_finance_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("them", handle_add_button),
            MessageHandler(filters.Regex("^➕ Thêm$"), handle_add_button)
        ],
        states={
            CHOOSING_ADD_TYPE: [CallbackQueryHandler(choose_add_type_callback, pattern="^add_type_")],
            CHOOSING_SUPPLEMENT_TYPE: [CallbackQueryHandler(choose_supplement_type_callback, pattern="^supplement_type_")],
            CHOOSING_INCOME_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_income_amount)],
            CONFIRM_INCOME_AMOUNT: [CallbackQueryHandler(confirm_income_amount, pattern="^confirm_amount_")],
            CHOOSING_INCOME_NOTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_income_note)],
            CHOOSING_INCOME_DATE: [CallbackQueryHandler(choose_income_date, pattern="^date_"), MessageHandler(filters.TEXT & ~filters.COMMAND, get_income_date_input)],
            CHOOSING_INCOME_SOURCE: [CallbackQueryHandler(get_income_source, pattern="^income_source_")],
            CHOOSING_EXPENSE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_expense_amount)],
            CONFIRM_EXPENSE_AMOUNT: [CallbackQueryHandler(confirm_expense_amount, pattern="^confirm_amount_")],
            CHOOSING_EXPENSE_NOTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_expense_note)],
            CHOOSING_EXPENSE_DATE: [CallbackQueryHandler(choose_expense_date, pattern="^date_"), MessageHandler(filters.TEXT & ~filters.COMMAND, get_expense_date_input)],
            CHOOSING_EXPENSE_CATEGORY: [CallbackQueryHandler(get_expense_category, pattern="^expense_category_")],
        },
        fallbacks=[CommandHandler("cancel", cancel_conversation), CallbackQueryHandler(cancel_conversation_callback, pattern="^cancel_action$")],
        # per_message=True is removed as it causes new convs for each message, not desired here.
        # per_user=True is default, ensuring one active conversation per user.
    )

    delete_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("xoa", handle_delete_button),
            MessageHandler(filters.Regex("^🗑️ Xóa$"), handle_delete_button)
        ],
        states={
            DELETE_CHOOSING_ACTION: [CallbackQueryHandler(choose_delete_action, pattern="^delete_action_")],
            DELETE_INDIVIDUAL_LISTING: [CallbackQueryHandler(delete_choose_item, pattern="^delete_id_")],
            DELETE_INDIVIDUAL_CONFIRM: [CallbackQueryHandler(delete_confirm_action, pattern="^confirm_delete_")],
            RESET_DATA_CONFIRM_STEP1: [CallbackQueryHandler(reset_data_confirmation_1_handler, pattern="^reset_confirm_1_")],
            RESET_DATA_CONFIRM_STEP2: [CallbackQueryHandler(reset_data_confirmation_2_handler, pattern="^reset_confirm_2_")],
        },
        fallbacks=[CommandHandler("cancel", cancel_conversation), CallbackQueryHandler(cancel_conversation_callback, pattern="^cancel_action$")],
    )

    report_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("baocao", report_start),
            MessageHandler(filters.Regex("^📈 Báo Cáo$"), report_start)
        ],
        states={CHOOSING_REPORT_PERIOD: [CallbackQueryHandler(generate_report_for_period, pattern="^report_period_|^cancel_action$")]},
        fallbacks=[CommandHandler("cancel", cancel_conversation), CallbackQueryHandler(cancel_conversation_callback, pattern="^cancel_action$")],
    )

    chart_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("bieudo", chart_start),
            MessageHandler(filters.Regex("^📊 Biểu Đồ$"), chart_start)
        ],
        states={CHOOSING_CHART_PERIOD: [CallbackQueryHandler(generate_charts_for_period, pattern="^chart_period_|^cancel_action$")]},
        fallbacks=[CommandHandler("cancel", cancel_conversation), CallbackQueryHandler(cancel_conversation_callback, pattern="^cancel_action$")],
    )

    export_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("xuatfile", export_file_start),
            MessageHandler(filters.Regex("^📤 Xuất File$"), export_file_start)
        ],
        states={
            EXPORT_CHOOSING_FORMAT: [CallbackQueryHandler(handle_export_choice, pattern="^export_|^cancel_action$")]
        },
        fallbacks=[CommandHandler("cancel", cancel_conversation), CallbackQueryHandler(cancel_conversation_callback, pattern="^cancel_action$")],
    )

    import_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("nhapfile", import_file_start),
            MessageHandler(filters.Regex("^📥 Nhập File$"), import_file_start)
        ],
        states={
            IMPORT_FILE_UPLOADED: [MessageHandler(filters.Document.ALL, handle_import_file_upload)],
        },
        fallbacks=[CommandHandler("cancel", cancel_conversation), CallbackQueryHandler(cancel_conversation_callback, pattern="^cancel_action$")],
    )

    application.add_handler(add_finance_conv_handler)
    application.add_handler(delete_conv_handler)
    application.add_handler(report_conv_handler)
    application.add_handler(chart_conv_handler)
    application.add_handler(export_conv_handler)
    application.add_handler(import_conv_handler)

    # Global CommandHandlers and MessageHandlers should be added AFTER ConversationHandlers
    application.add_handler(CommandHandler("ngansach", budget_set_cmd))
    application.add_handler(CommandHandler("xem_ngansach", budget_list_cmd))
    application.add_handler(CommandHandler("xoa_ngansach", budget_delete_cmd))
    application.add_handler(CommandHandler("menu", show_main_menu_handler))
    application.add_handler(CommandHandler("huongdan", show_help))
    application.add_handler(MessageHandler(filters.Regex("^❓ Hướng Dẫn$"), show_help))
    
    # Global cancel handler: important to break out of any ConversationHandler
    application.add_handler(CommandHandler("cancel", cancel_conversation))
    
    # This catch-all MessageHandler should be the very last handler to ensure ConversationHandlers get priority.
    # It also ensures that if a user types random text not part of any conversation, they get the main menu.
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, quick_entry_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, show_main_menu_handler))

    print("Bot đang chạy...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    try:
        init_db()
    except Exception:
        pass
    main()
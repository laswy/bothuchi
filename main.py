# -*- coding: utf-8 -*-
"""
Univer All-in-One Bot
- Telegram bot (python-telegram-bot v22 async, Python 3.9+)
- Crypto portfolio (CoinGecko, themes, charts)
- Personal finance (income/expense, budgets, reports, charts)
- HTML dashboard server
"""
from __future__ import annotations

import os, io, csv, sqlite3, datetime, tempfile, asyncio, traceback, re, threading, logging, json
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, quote

try:
    from dotenv import load_dotenv
    load_dotenv(dotenv_path=".env")
except ImportError:
    pass

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.font_manager as _fm
import matplotlib.colors as mcolors
from matplotlib.patches import FancyBboxPatch, Circle
from matplotlib import gridspec
import numpy as np
import requests

try:
    import openpyxl
    from openpyxl import load_workbook
except Exception:
    openpyxl = None
    load_workbook = None

from telegram import (
    Update, InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, ContextTypes,
    CallbackQueryHandler, MessageHandler, filters, ConversationHandler
)

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# ===================== CONFIG =====================
BOT_TOKEN        = os.environ["TELEGRAM_BOT_TOKEN"]
DB_PATH          = os.environ.get("UNIVER_DB_PATH", "univer_all_in_one.db")
DASHBOARD_SECRET = os.environ.get("DASHBOARD_SECRET", "")

# Webhook config (để trống = dùng polling)
WEBHOOK_URL    = os.environ.get("WEBHOOK_URL", "").rstrip("/")   # vd: https://mybot.up.railway.app
WEBHOOK_PORT   = int(os.environ.get("WEBHOOK_PORT") or os.environ.get("PORT") or "8443")
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET", "")

# Dashboard port — tránh trùng với webhook port
_default_html = 8080 if WEBHOOK_PORT != 8080 else 8081
HTML_PORT = int(os.environ.get("HTML_PORT", str(_default_html)))

COINGECKO_BASE = "https://api.coingecko.com/api/v3"

OWNER_USER_ID   = 679130099
CHANNEL_CHAT_ID    = int(os.environ.get("CHANNEL_CHAT_ID",    "-1001974996093"))
FINANCE_CHANNEL_ID = os.environ.get("FINANCE_CHANNEL_ID") or CHANNEL_CHAT_ID

LARGE_AMOUNT_THRESHOLD = 100_000_000

# ===================== THEMES =====================
THEMES = {
    "dark": {
        "name": "🌑 Dark",
        "bg": "#0a0e27", "card_bg": "#151932",
        "text_primary": "#ffffff", "text_secondary": "#94a3b8",
        "accent": "#00d4ff",
        "chart_colors": ["#00d4ff","#00ff88","#ffd700","#ff6b6b","#c56cf0",
                         "#4834d4","#22a6b3","#f0932b","#eb4d4b","#6ab04c",
                         "#30336b","#95afc0","#535c68","#ff9ff3","#48dbfb"],
        "bg_dark": "#1a1a2e", "bg_header": "#16213e",
        "bg_row1": "#0f3460", "bg_row2": "#0e2954", "bg_total": "#16213e",
        "text_light": "#e8e8e8", "text_dim": "#a8a8a8",
        "positive": "#00ff88", "negative": "#ff4757", "neutral": "#ffd700",
        "border": "#2a2a3e",
    },
    "light": {
        "name": "☀️ Light",
        "bg": "#f0f4f8", "card_bg": "#ffffff",
        "text_primary": "#1a202c", "text_secondary": "#4a5568",
        "accent": "#3182ce",
        "chart_colors": ["#3182ce","#38a169","#d69e2e","#e53e3e","#805ad5",
                         "#2b6cb0","#2c7a7b","#c05621","#c53030","#276749",
                         "#553c9a","#2d3748","#718096","#d53f8c","#0987a0"],
        "bg_dark": "#f7fafc", "bg_header": "#e2e8f0",
        "bg_row1": "#edf2f7", "bg_row2": "#f7fafc", "bg_total": "#e2e8f0",
        "text_light": "#1a202c", "text_dim": "#718096",
        "positive": "#38a169", "negative": "#e53e3e", "neutral": "#d69e2e",
        "border": "#cbd5e0",
    },
    "neon": {
        "name": "💚 Neon",
        "bg": "#0d0d0d", "card_bg": "#1a1a1a",
        "text_primary": "#00ff41", "text_secondary": "#00cc33",
        "accent": "#00ff41",
        "chart_colors": ["#00ff41","#ff00ff","#00ffff","#ffff00","#ff6600",
                         "#ff0066","#66ff00","#0066ff","#ff3300","#00ff99",
                         "#ff00cc","#ccff00","#00ccff","#ff9900","#9900ff"],
        "bg_dark": "#0d0d0d", "bg_header": "#1a1a1a",
        "bg_row1": "#0a1a0a", "bg_row2": "#0d0d0d", "bg_total": "#1a1a1a",
        "text_light": "#00ff41", "text_dim": "#009922",
        "positive": "#00ff41", "negative": "#ff0066", "neutral": "#ffff00",
        "border": "#003300",
    },
    "bpink": {
        "name": "🩷 Black Pink",
        "bg": "#0d0010", "card_bg": "#1a0020",
        "text_primary": "#ffffff", "text_secondary": "#d4a0d4",
        "accent": "#ff69b4",
        "chart_colors": ["#ff69b4","#ff1493","#da70d6","#ba55d3","#9400d3",
                         "#ff85c8","#ff4da6","#e040fb","#ce93d8","#f48fb1",
                         "#ff80ab","#ea80fc","#b39ddb","#f8bbd0","#e1bee7"],
        "bg_dark": "#0d0010", "bg_header": "#1a0020",
        "bg_row1": "#250030", "bg_row2": "#1a0020", "bg_total": "#2d0040",
        "text_light": "#ffffff", "text_dim": "#d4a0d4",
        "positive": "#ff69b4", "negative": "#ff1744", "neutral": "#e040fb",
        "border": "#4a0060",
    },
}
DEFAULT_THEME = "dark"
_user_themes: dict[int, str] = {}

# ===================== RATE LIMITER =====================
RATE_WINDOW = 60
RATE_MAX_CALLS = 10
_rate_store: dict[int, list] = {}

def is_rate_limited(user_id: int) -> bool:
    now = datetime.datetime.now().timestamp()
    calls = [t for t in _rate_store.get(user_id, []) if now - t < RATE_WINDOW]
    if len(calls) >= RATE_MAX_CALLS:
        _rate_store[user_id] = calls
        return True
    calls.append(now)
    _rate_store[user_id] = calls
    return False

async def check_rate(update: Update) -> bool:
    if is_rate_limited(update.effective_user.id):
        await update.message.reply_text(f"⏳ Gửi lệnh quá nhanh. Chờ {RATE_WINDOW}s rồi thử lại.")
        return True
    return False

# ===================== DB MIGRATIONS =====================
_MIGRATIONS_DONE = False

def run_migrations():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # Crypto tables
    cur.execute("""CREATE TABLE IF NOT EXISTS crypto_trades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL, symbol TEXT NOT NULL, cg_id TEXT,
        side TEXT NOT NULL, qty REAL NOT NULL, price_usd REAL NOT NULL,
        fee_usd REAL DEFAULT 0, note TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS crypto_map (
        symbol TEXT PRIMARY KEY, cg_id TEXT NOT NULL, name TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS crypto_prices (
        cg_id TEXT NOT NULL, price_usd REAL NOT NULL,
        fetched_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS user_settings (
        user_id INTEGER PRIMARY KEY, theme TEXT NOT NULL DEFAULT 'dark')""")
    # Finance tables
    cur.execute("""CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, amount REAL, note TEXT, category TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS incomes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, amount REAL, source TEXT, note TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS budgets (
        user_id INTEGER, category TEXT, amount REAL,
        PRIMARY KEY (user_id, category))""")
    cur.execute("""CREATE TABLE IF NOT EXISTS recurring_transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        ttype TEXT NOT NULL,
        amount REAL NOT NULL,
        category TEXT NOT NULL,
        note TEXT DEFAULT '',
        day_of_month INTEGER NOT NULL,
        month INTEGER,
        last_triggered TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit()
    conn.close()

def ensure_migrated():
    global _MIGRATIONS_DONE
    if not _MIGRATIONS_DONE:
        run_migrations()
        _MIGRATIONS_DONE = True

def db_conn():
    ensure_migrated()
    return sqlite3.connect(DB_PATH)

# ===================== CRYPTO DB =====================
def db_get_theme(user_id: int) -> str:
    if user_id in _user_themes:
        return _user_themes[user_id]
    conn = db_conn(); cur = conn.cursor()
    cur.execute("SELECT theme FROM user_settings WHERE user_id=?", (user_id,))
    row = cur.fetchone(); conn.close()
    theme = row[0] if row and row[0] in THEMES else DEFAULT_THEME
    _user_themes[user_id] = theme
    return theme

def db_set_theme(user_id: int, theme: str):
    if theme not in THEMES: theme = DEFAULT_THEME
    _user_themes[user_id] = theme
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""INSERT INTO user_settings(user_id,theme) VALUES(?,?)
        ON CONFLICT(user_id) DO UPDATE SET theme=excluded.theme""", (user_id, theme))
    conn.commit(); conn.close()

def db_crypto_upsert_map(symbol: str, cg_id: str, name: Optional[str] = None):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""INSERT INTO crypto_map(symbol,cg_id,name) VALUES(?,?,?)
        ON CONFLICT(symbol) DO UPDATE SET cg_id=excluded.cg_id,
        name=COALESCE(excluded.name,crypto_map.name)""",
        (symbol.upper(), cg_id, name))
    conn.commit(); conn.close()

def db_crypto_get_map(symbol: str) -> Optional[str]:
    conn = db_conn(); cur = conn.cursor()
    cur.execute("SELECT cg_id FROM crypto_map WHERE symbol=?", (symbol.upper(),))
    row = cur.fetchone(); conn.close()
    return row[0] if row else None

def db_crypto_add_trade(user_id: int, symbol: str, side: str, qty: float,
                        price_usd: float, note: str = "",
                        created_at: Optional[datetime.datetime] = None,
                        cg_id: Optional[str] = None, fee_usd: float = 0.0):
    conn = db_conn(); cur = conn.cursor()
    if created_at is None: created_at = datetime.datetime.now()
    cur.execute("""INSERT INTO crypto_trades
        (user_id,symbol,cg_id,side,qty,price_usd,fee_usd,note,created_at)
        VALUES(?,?,?,?,?,?,?,?,?)""",
        (user_id, symbol.upper(), cg_id, side.upper(), qty, price_usd,
         fee_usd, note, created_at.isoformat()))
    conn.commit(); conn.close()

def db_crypto_positions(user_id: int):
    ensure_migrated()
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT symbol,
        SUM(CASE WHEN side='BUY' THEN qty ELSE -qty END) AS qty_net,
        SUM(CASE WHEN side='BUY' THEN qty*price_usd+fee_usd ELSE -(qty*price_usd) END) AS invested_usd
        FROM crypto_trades WHERE user_id=?
        GROUP BY symbol HAVING ABS(qty_net)>1e-12 ORDER BY symbol""", (user_id,))
    rows = cur.fetchall()
    result = []
    for symbol, qty_net, invested_usd in rows:
        cur.execute("""SELECT cg_id FROM crypto_trades
            WHERE user_id=? AND symbol=? AND cg_id IS NOT NULL
            ORDER BY created_at DESC LIMIT 1""", (user_id, symbol))
        row = cur.fetchone()
        cg_id = (row[0] if row and row[0] else None) or db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
        result.append({"symbol": symbol, "qty": float(qty_net or 0),
                       "invested_usd": float(invested_usd or 0), "cg_id": cg_id})
    conn.close()
    return result

def db_crypto_all_trades(user_id: int):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT symbol,COALESCE(cg_id,''),side,qty,price_usd,
        fee_usd,COALESCE(note,''),created_at
        FROM crypto_trades WHERE user_id=? ORDER BY created_at ASC""", (user_id,))
    rows = cur.fetchall(); conn.close()
    return rows

def _get_held_qty(user_id: int, symbol: str) -> float:
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT SUM(CASE WHEN side='BUY' THEN qty ELSE -qty END)
        FROM crypto_trades WHERE user_id=? AND symbol=?""",
        (user_id, symbol.upper()))
    row = cur.fetchone(); conn.close()
    return float(row[0] or 0) if row and row[0] is not None else 0.0

# ===================== FINANCE DB =====================
def db_add_income(user_id: int, amount: float, source: str, note: str,
                  created_at: datetime.datetime = None):
    conn = db_conn(); cur = conn.cursor()
    if created_at is None: created_at = datetime.datetime.now()
    cur.execute("INSERT INTO incomes(user_id,amount,source,note,created_at) VALUES(?,?,?,?,?)",
                (user_id, amount, source, note, created_at.isoformat()))
    conn.commit(); conn.close()

def db_add_expense(user_id: int, amount: float, note: str, category: str,
                   created_at: datetime.datetime = None):
    conn = db_conn(); cur = conn.cursor()
    if created_at is None: created_at = datetime.datetime.now()
    cur.execute("INSERT INTO expenses(user_id,amount,note,category,created_at) VALUES(?,?,?,?,?)",
                (user_id, amount, note, category, created_at.isoformat()))
    conn.commit(); conn.close()

def db_list_expenses_grouped(user_id: int, start_date=None, end_date=None):
    conn = db_conn(); cur = conn.cursor()
    q = "SELECT category,SUM(amount) FROM expenses WHERE user_id=?"
    params = [user_id]
    if start_date and end_date:
        q += " AND created_at>=? AND created_at<?"
        params += [start_date.isoformat(), (end_date+datetime.timedelta(days=1)).isoformat()]
    q += " GROUP BY category ORDER BY SUM(amount) DESC"
    cur.execute(q, tuple(params)); rows = cur.fetchall(); conn.close()
    return rows

def db_list_incomes_grouped(user_id: int, start_date=None, end_date=None):
    conn = db_conn(); cur = conn.cursor()
    q = "SELECT source,SUM(amount) FROM incomes WHERE user_id=?"
    params = [user_id]
    if start_date and end_date:
        q += " AND created_at>=? AND created_at<?"
        params += [start_date.isoformat(), (end_date+datetime.timedelta(days=1)).isoformat()]
    q += " GROUP BY source ORDER BY SUM(amount) DESC"
    cur.execute(q, tuple(params)); rows = cur.fetchall(); conn.close()
    return rows

def db_get_combined_summary(user_id: int, start_date=None, end_date=None):
    conn = db_conn(); cur = conn.cursor()
    params = [user_id]
    inc_q = "SELECT SUM(amount) FROM incomes WHERE user_id=?"
    exp_q = "SELECT SUM(amount) FROM expenses WHERE user_id=?"
    if start_date and end_date:
        extra = " AND created_at>=? AND created_at<?"
        inc_q += extra; exp_q += extra
        params += [start_date.isoformat(), (end_date+datetime.timedelta(days=1)).isoformat()]
    cur.execute(inc_q, tuple(params)); income = cur.fetchone()[0] or 0
    cur.execute(exp_q, tuple(params)); expense = cur.fetchone()[0] or 0
    conn.close()
    return income, expense, income - expense

def db_get_monthly_report(user_id, start_date=None, end_date=None):
    conn = db_conn(); cur = conn.cursor()
    params = [user_id]
    exp_q = "SELECT strftime('%Y-%m',created_at) as m,SUM(amount) FROM expenses WHERE user_id=?"
    inc_q = "SELECT strftime('%Y-%m',created_at) as m,SUM(amount) FROM incomes WHERE user_id=?"
    if start_date and end_date:
        extra = " AND created_at>=? AND created_at<?"
        exp_q += extra; inc_q += extra
        params += [start_date.isoformat(), (end_date+datetime.timedelta(days=1)).isoformat()]
    exp_q += " GROUP BY m ORDER BY m"; inc_q += " GROUP BY m ORDER BY m"
    cur.execute(exp_q, tuple(params)); expenses = cur.fetchall()
    cur.execute(inc_q, tuple(params)); incomes = cur.fetchall()
    conn.close()
    return expenses, incomes

def db_export_data(user_id: int):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT 'Chi tiêu',amount,note,category,created_at,id
        FROM expenses WHERE user_id=?
        UNION ALL
        SELECT 'Thu nhập',amount,note,source,created_at,id
        FROM incomes WHERE user_id=?
        ORDER BY created_at DESC""", (user_id, user_id))
    rows = cur.fetchall(); conn.close()
    return rows

def db_get_last_n_transactions(user_id: int, limit: int = 5):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT id,'expense',amount,note,category,created_at FROM expenses WHERE user_id=?
        UNION ALL
        SELECT id,'income',amount,note,source,created_at FROM incomes WHERE user_id=?
        ORDER BY created_at DESC LIMIT ?""", (user_id, user_id, limit))
    rows = cur.fetchall(); conn.close()
    return rows

def db_delete_transaction(transaction_id: int, transaction_type: str):
    conn = db_conn(); cur = conn.cursor()
    table = "expenses" if transaction_type == 'expense' else "incomes"
    cur.execute(f"DELETE FROM {table} WHERE id=?", (transaction_id,))
    conn.commit(); conn.close()

def db_reset_all_data(user_id: int):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM expenses WHERE user_id=?", (user_id,))
    cur.execute("DELETE FROM incomes WHERE user_id=?", (user_id,))
    conn.commit(); conn.close()

def db_set_budget(user_id: int, category: str, amount: float):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""INSERT INTO budgets(user_id,category,amount) VALUES(?,?,?)
        ON CONFLICT(user_id,category) DO UPDATE SET amount=excluded.amount""",
        (user_id, category, amount))
    conn.commit(); conn.close()

def db_get_budgets(user_id: int):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("SELECT category,amount FROM budgets WHERE user_id=?", (user_id,))
    rows = cur.fetchall(); conn.close()
    return {r[0]: r[1] for r in rows}

def db_delete_budget(user_id: int, category: str):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM budgets WHERE user_id=? AND category=?", (user_id, category))
    conn.commit(); conn.close()

def get_month_range(dt: datetime.datetime):
    start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_m = (start.replace(day=28)+datetime.timedelta(days=4)).replace(day=1)
    end = (next_m-datetime.timedelta(days=1)).replace(hour=23, minute=59, second=59)
    return start, end

def db_sum_expense_by_category_in_period(user_id: int, category: str,
                                          start_date: datetime.datetime,
                                          end_date: datetime.datetime) -> float:
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT COALESCE(SUM(amount),0) FROM expenses
        WHERE user_id=? AND category=? AND created_at>=? AND created_at<=?""",
        (user_id, category, start_date.isoformat(), end_date.isoformat()))
    val = cur.fetchone()[0] or 0.0; conn.close()
    return float(val)
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import urlparse, parse_qs

# ── SECTION 1: DB HELPERS ──────────────────────────────

def db_crypto_get_trades(user_id: int):
    """Return list of tuples (id,symbol,cg_id,side,qty,price_usd,fee_usd,note,created_at) DESC."""
    try:
        conn = db_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT id,symbol,COALESCE(cg_id,''),side,qty,price_usd,
                   COALESCE(fee_usd,0),COALESCE(note,''),created_at
            FROM crypto_trades WHERE user_id=?
            ORDER BY created_at DESC LIMIT 50
        """, (user_id,))
        rows = cur.fetchall(); conn.close(); return rows
    except Exception:
        return []


def db_crypto_delete_trade(trade_id: int):
    """DELETE FROM crypto_trades WHERE id=?"""
    conn = db_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM crypto_trades WHERE id=?", (trade_id,))
    conn.commit(); conn.close()


def db_crypto_reset_all(user_id: int):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM crypto_trades WHERE user_id=?", (user_id,))
    conn.commit(); conn.close()


def db_crypto_update_trade(trade_id: int, qty: float, price: float, fee: float, note: str):
    """UPDATE crypto_trades SET qty,price_usd,fee_usd,note WHERE id=?"""
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE crypto_trades SET qty=?,price_usd=?,fee_usd=?,note=? WHERE id=?
    """, (qty, price, fee, note, trade_id))
    conn.commit(); conn.close()


def db_update_income(income_id: int, amount: float, source: str, note: str, date_str: str):
    """UPDATE incomes SET amount,source,note,created_at WHERE id=?"""
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE incomes SET amount=?,source=?,note=?,created_at=? WHERE id=?
    """, (amount, source, note, date_str, income_id))
    conn.commit(); conn.close()


def db_update_expense(expense_id: int, amount: float, category: str, note: str, date_str: str):
    """UPDATE expenses SET amount,category,note,created_at WHERE id=?"""
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE expenses SET amount=?,category=?,note=?,created_at=? WHERE id=?
    """, (amount, category, note, date_str, expense_id))
    conn.commit(); conn.close()


def db_get_incomes_list(user_id: int, limit: int = 50):
    """Return list of (id,amount,source,note,created_at) ORDER BY created_at DESC."""
    try:
        conn = db_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT id,amount,COALESCE(source,''),COALESCE(note,''),created_at
            FROM incomes WHERE user_id=?
            ORDER BY created_at DESC LIMIT ?
        """, (user_id, limit))
        rows = cur.fetchall(); conn.close(); return rows
    except Exception:
        return []


def db_get_expenses_list(user_id: int, limit: int = 50):
    """Return list of (id,amount,category,note,created_at) ORDER BY created_at DESC."""
    try:
        conn = db_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT id,amount,COALESCE(category,''),COALESCE(note,''),created_at
            FROM expenses WHERE user_id=?
            ORDER BY created_at DESC LIMIT ?
        """, (user_id, limit))
        rows = cur.fetchall(); conn.close(); return rows
    except Exception:
        return []


def db_get_incomes_filtered(user_id: int, date_from=None, date_to=None,
                             category=None, search=None, limit: int = 500):
    try:
        conn = db_conn(); cur = conn.cursor()
        sql = "SELECT id,amount,COALESCE(source,''),COALESCE(note,''),created_at FROM incomes WHERE user_id=?"
        params: list = [user_id]
        if date_from:
            sql += " AND created_at >= ?"; params.append(date_from)
        if date_to:
            sql += " AND created_at <= ?"; params.append(date_to + " 23:59:59")
        if category:
            sql += " AND LOWER(source) LIKE ?"; params.append(f"%{category.lower()}%")
        if search:
            sql += " AND (LOWER(note) LIKE ? OR LOWER(source) LIKE ?)"; params += [f"%{search.lower()}%", f"%{search.lower()}%"]
        sql += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        cur.execute(sql, params)
        rows = cur.fetchall(); conn.close(); return rows
    except Exception:
        return []


def db_get_expenses_filtered(user_id: int, date_from=None, date_to=None,
                              category=None, search=None, limit: int = 500):
    try:
        conn = db_conn(); cur = conn.cursor()
        sql = "SELECT id,amount,COALESCE(category,''),COALESCE(note,''),created_at FROM expenses WHERE user_id=?"
        params: list = [user_id]
        if date_from:
            sql += " AND created_at >= ?"; params.append(date_from)
        if date_to:
            sql += " AND created_at <= ?"; params.append(date_to + " 23:59:59")
        if category:
            sql += " AND LOWER(category) LIKE ?"; params.append(f"%{category.lower()}%")
        if search:
            sql += " AND (LOWER(note) LIKE ? OR LOWER(category) LIKE ?)"; params += [f"%{search.lower()}%", f"%{search.lower()}%"]
        sql += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        cur.execute(sql, params)
        rows = cur.fetchall(); conn.close(); return rows
    except Exception:
        return []


def db_get_finance_categories(user_id: int):
    try:
        conn = db_conn(); cur = conn.cursor()
        cur.execute("SELECT DISTINCT source FROM incomes WHERE user_id=? AND source IS NOT NULL AND source!='' ORDER BY source", (user_id,))
        inc_cats = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT category FROM expenses WHERE user_id=? AND category IS NOT NULL AND category!='' ORDER BY category", (user_id,))
        exp_cats = [r[0] for r in cur.fetchall()]
        conn.close()
        return {"income": inc_cats, "expense": exp_cats}
    except Exception:
        return {"income": [], "expense": []}


# ── Recurring transactions DB ──────────────────────────

def db_recurring_add(user_id, ttype, amount, category, note, day, month=None):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""INSERT INTO recurring_transactions
        (user_id,ttype,amount,category,note,day_of_month,month)
        VALUES (?,?,?,?,?,?,?)""", (user_id, ttype, amount, category, note or '', day, month))
    conn.commit(); conn.close()

def db_recurring_list(user_id):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT id,ttype,amount,category,note,day_of_month,month,last_triggered
        FROM recurring_transactions WHERE user_id=? ORDER BY month,day_of_month""", (user_id,))
    rows = cur.fetchall(); conn.close(); return rows

def db_recurring_get(rec_id):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT id,user_id,ttype,amount,category,note,day_of_month,month,last_triggered
        FROM recurring_transactions WHERE id=?""", (rec_id,))
    row = cur.fetchone(); conn.close(); return row

def db_recurring_update(rec_id, amount, category, note, day, month):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""UPDATE recurring_transactions
        SET amount=?,category=?,note=?,day_of_month=?,month=? WHERE id=?""",
        (amount, category, note, day, month, rec_id))
    conn.commit(); conn.close()

def db_recurring_delete(rec_id):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM recurring_transactions WHERE id=?", (rec_id,))
    conn.commit(); conn.close()

def db_recurring_update_triggered(rec_id, period_key):
    conn = db_conn(); cur = conn.cursor()
    cur.execute("UPDATE recurring_transactions SET last_triggered=? WHERE id=?", (period_key, rec_id))
    conn.commit(); conn.close()

def db_recurring_all():
    conn = db_conn(); cur = conn.cursor()
    cur.execute("""SELECT id,user_id,ttype,amount,category,note,day_of_month,month,last_triggered
        FROM recurring_transactions""")
    rows = cur.fetchall(); conn.close(); return rows


# ── SECTION 2: _DashboardHandler ───────────────────────

# ===================== NLP PARSERS (Vietnamese) =====================
VIET_NUM = {
    "k": 1000, "nghin": 1000, "nghìn": 1000, "ngan": 1000, "ngàn": 1000,
    "tr": 1_000_000, "trieu": 1_000_000, "triệu": 1_000_000,
    "củ": 1_000_000, "cu": 1_000_000
}

CATEGORY_KEYWORDS = {
    "Ăn uống": ["ăn","ăn sáng","bữa sáng","sáng","trưa","chiều","tối","cafe","cà phê","trà sữa","quán","nhậu","uống","đi ăn","trà đá","buffet","liên hoan","bánh","kem","đồ uống","thức ăn","đặt món"],
    "Di chuyển": ["taxi","grab","xăng","vé xe","tàu","vé tàu","xe buýt","bus","gửi xe","bến","đi lại","đi xe","xe khách","máy bay","vé máy bay","xích lô","ô tô","đi grab","ship"],
    "Mua sắm": ["mua","đồ","áo","quần","giày","shopee","lazada","tiki","siêu thị","mua hàng","đặt hàng","order","đi chợ","mỹ phẩm","trang sức","phụ kiện"],
    "Hóa đơn": ["điện","nước","wifi","internet","cước","hóa đơn","tiền nhà","truyền hình","gas","rác thải","điện thoại","sim","4g","5g"],
    "Gia đình": ["gia đình","bà ngoại","bố mẹ","con","vợ","chồng","ông bà","anh chị","em","gia tộc","cháu"],
    "Sức khỏe": ["khám","thuốc","bệnh viện","y tế","khám bệnh","xét nghiệm","phẫu thuật","tiêm","bảo hiểm y tế","bảo hiểm sức khỏe"],
    "Giải trí": ["xem phim","anime","game","giải trí","nhạc","hát","karaoke","du lịch","chơi","coi phim","xem ca nhạc","thể thao"],
    "Đầu tư": ["đầu tư","crypto","coin","chứng khoán","cổ phiếu","trái phiếu","mua coin","mua cổ","vàng","bất động sản"],
    "Tiết kiệm": ["tiết kiệm","gửi tiết kiệm","mở sổ","sổ tiết kiệm","để dành","dành dụm"],
    "Khác": ["từ thiện","quà","ủng hộ","học phí","đào tạo","khóa học","mua tên miền","dịch vụ"]
}

INCOME_CATEGORY_KEYWORDS = {
    "Lương": ["lương","tiền công","tiền lương","lương tháng","salary","lương cơ bản","lương chính","lương thời vụ","lương ngày","lương tuần","phụ cấp","trợ cấp","tiền công nhật"],
    "Thưởng": ["thưởng","bonus","thưởng tết","thưởng quý","thưởng năm","thưởng doanh số","thưởng năng suất","thưởng thêm"],
    "Kinh doanh": ["kinh doanh","bán hàng","doanh thu","hoa hồng","chốt đơn","bán online","bán trực tiếp","lợi nhuận shop","thu bán hàng","thu từ bán","thu cửa hàng","thu quán"],
    "Được cho": ["được cho","được tặng","cho tiền","mừng tuổi","mừng cưới","tiền mừng","quà","cho","biếu","tặng","hỗ trợ","cha mẹ cho","người thân cho","bạn bè cho"],
    "Thu hồi nợ": ["trả nợ","thu hồi nợ","được trả","trả lại","người nợ trả","hoàn tiền","hoàn vốn","nhận lại tiền"],
    "Bán tài sản": ["bán tài sản","bán xe","bán nhà","thanh lý","bán đất","bán đồ","thanh lý hàng","bán máy","bán điện thoại","bán vàng","bán máy tính","bán laptop","bán đồ cũ"],
    "Lãi tiết kiệm": ["lãi tiết kiệm","tiền gửi","gửi tiết kiệm","lãi ngân hàng","lãi sổ tiết kiệm","tiền lãi tiết kiệm"],
    "Lãi đầu tư": ["lãi đầu tư","cổ tức","trái tức","lợi nhuận","lãi cổ phiếu","lãi coin","lãi crypto","lãi chứng khoán","lãi quỹ","lợi nhuận đầu tư","lãi trái phiếu","lãi bất động sản","P2P","Crypto"],
    "Khác": ["thu nhập khác","nguồn khác","không xác định","thu lặt vặt"]
}

INCOME_KEYWORDS = [kw for kws in INCOME_CATEGORY_KEYWORDS.values() for kw in kws]
DATE_KEYWORDS = {
    "hôm nay": 0, "hom nay": 0, "nay": 0,
    "hôm qua": -1, "hom qua": -1, "qua": -1,
    "hôm kia": -2, "hom kia": -2
}

def parse_vietnamese_amount(text: str):
    text = text.strip().lower().replace(",","").replace(" ","")
    m = re.match(r"^(\d+)(tr|trieu|triệu|củ|cu)(\d+)$", text)
    if m:
        return int(m.group(1))*1_000_000 + int(m.group(3))*100_000
    m = re.match(r"^(\d+)(k|nghin|nghìn|ngan|ngàn)(\d+)$", text)
    if m:
        return int(m.group(1))*1_000 + int(m.group(3))*100
    for key, mult in VIET_NUM.items():
        if text.endswith(key):
            num = text[:-len(key)]
            if num.isdigit():
                return int(num)*mult
    if text.isdigit():
        return int(text)
    return None

def parse_amount_loose(text: str):
    t = text.lower().replace(" ","")
    t = (t.replace("triệu","tr").replace("trieu","tr")
          .replace("nghìn","k").replace("nghin","k")
          .replace("ngàn","k").replace("ngan","k"))
    m = re.match(r'^(\d+)(tr)(\d+)$', t)
    if m: return int(m.group(1))*1_000_000 + int(m.group(3))*100_000
    m = re.match(r'^(\d+)(k)(\d+)$', t)
    if m: return int(m.group(1))*1_000 + int(m.group(3))*100
    m = re.search(r'(\d{1,3}(?:[.,]\d{3})+|\d+(?:[.,]\d+)?)(k|tr)?', t)
    if not m: return None
    num, unit = m.group(1), m.group(2) or ""
    num = num.replace(".","").replace(",",".")
    try: val = float(num)
    except ValueError: return None
    if unit == "k": val *= 1000
    elif unit == "tr": val *= 1_000_000
    return round(val)

def parse_amount(text: str) -> float:
    v = parse_vietnamese_amount(text)
    if v is not None: return float(v)
    v2 = parse_amount_loose(text)
    if v2 is not None: return float(v2)
    t = str(text).lower().strip().replace(",","").replace(" ","")
    if t.endswith("tr"):
        try: return float(t[:-2])*1_000_000
        except Exception: pass
    if t.endswith("k"):
        try: return float(t[:-1])*1_000
        except Exception: pass
    return float(re.sub(r'[^0-9.]','',t) or 0)

def parse_date(date_str: str) -> datetime.datetime:
    today = datetime.date.today()
    try: return datetime.datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError: pass
    try: return datetime.datetime.strptime(f"{date_str}/{today.year}", "%d/%m/%Y")
    except ValueError: raise ValueError("Định dạng ngày không hợp lệ. Dùng DD/MM/YYYY hoặc DD/MM.")

def guess_type(text: str) -> str:
    low = text.lower()
    for kw in INCOME_KEYWORDS:
        if kw in low: return "income"
    return "expense"

def guess_category(text: str) -> str:
    low = text.lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        for kw in kws:
            if kw in low: return cat
    return "Khác"

def guess_income_source(text: str) -> str:
    low = text.lower()
    for src, kws in INCOME_CATEGORY_KEYWORDS.items():
        for kw in kws:
            if kw in low: return src
    return "Khác"

def guess_date(text: str):
    low = text.lower()
    for k, delta in DATE_KEYWORDS.items():
        if k in low:
            return datetime.datetime.now() + datetime.timedelta(days=delta)
    m = re.search(r'(\d{1,2})[\/\-](\d{1,2})(?:[\/\-](\d{2,4}))?', low)
    if m:
        d, mth, yr = int(m.group(1)), int(m.group(2)), m.group(3)
        year = int(yr) if yr else datetime.datetime.now().year
        try: return datetime.datetime(year, mth, d)
        except ValueError: pass
    return datetime.datetime.now()

def clean_note(text: str) -> str:
    # Remove "ngày DD/MM/YYYY" then bare date patterns before stripping numbers
    # (otherwise the numbers are removed first and bare slashes like "//" remain)
    note = re.sub(r'ngày\s+\d{1,2}[\/\-]\d{1,2}(?:[\/\-]\d{2,4})?', '', text, flags=re.IGNORECASE)
    note = re.sub(r'\d{1,2}[\/\-]\d{1,2}(?:[\/\-]\d{2,4})?', '', note)
    note = re.sub(r'\d{1,3}(?:[.,]\d{3})+\s*đ?|\d+(?:[.,]\d+)?\s*(k|tr|triệu|nghìn|ngàn|đ)?', '', note, flags=re.IGNORECASE)
    note = re.sub(r'\s+', ' ', note)
    return note.strip(" -–—,.;:/") or "Không ghi chú"

def nlp_parse_transaction(raw: str):
    amount = parse_amount_loose(raw)
    if amount is None or amount <= 0: return None
    tx_type = guess_type(raw)
    created_at = guess_date(raw)
    note = clean_note(raw)
    category = guess_income_source(raw) if tx_type == "income" else guess_category(raw)
    return {"type": tx_type, "amount": amount, "category": category, "note": note, "created_at": created_at}

# ===================== COINGECKO =====================
def cg_guess_id_from_symbol(symbol: str) -> Optional[str]:
    static = {
        "BTC":"bitcoin","ETH":"ethereum","ATOM":"cosmos","DYDX":"dydx-chain",
        "BABY":"babylon","DYM":"dymension","AURA":"aura-network","ZETA":"zetachain",
        "SOL":"solana","BNB":"binancecoin","TON":"the-open-network",
        "USDT":"tether","USDC":"usd-coin"
    }
    return static.get(symbol.upper())

def cg_simple_price_usd(cg_ids: list[str]) -> dict[str, float]:
    if not cg_ids: return {}
    ids_param = ",".join(sorted(set(cg_ids)))
    url = f"{COINGECKO_BASE}/simple/price?ids={quote(ids_param)}&vs_currencies=usd"
    try:
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            return {k: float(v.get("usd",0) or 0) for k,v in r.json().items()}
    except Exception: pass
    return {}

def cg_market_chart_usd(cg_id: str, days: int = 365):
    interval = "hourly" if days <= 1 else "daily"
    url = f"{COINGECKO_BASE}/coins/{quote(cg_id)}/market_chart?vs_currency=usd&days={days}&interval={interval}"
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200: return {}
        out = {}
        for ts, price in r.json().get("prices", []):
            dt = datetime.datetime.fromtimestamp(ts/1000.0, tz=datetime.timezone.utc)
            label = dt.strftime('%H:%M') if days<=1 else dt.date().isoformat()
            out[label] = float(price)
        return out
    except Exception: return {}

def portfolio_history_series(user_id: int, days: int = 365):
    rows = db_crypto_all_trades(user_id)
    if not rows: return [], []
    trades_by_sym = {}; cg_ids = {}
    for symbol, cg_id, side, qty, price, fee, note, created_at in rows:
        sym = symbol.upper()
        trades_by_sym.setdefault(sym,[]).append((side, float(qty), created_at))
        if cg_id: cg_ids[sym] = cg_id
    for sym in trades_by_sym:
        if sym not in cg_ids:
            guess = db_crypto_get_map(sym) or cg_guess_id_from_symbol(sym)
            if guess: cg_ids[sym] = guess
    if not cg_ids: return [], []
    price_hist = {}; all_labels = set()
    for sym, cg in cg_ids.items():
        ph = cg_market_chart_usd(cg, days=days)
        if ph: price_hist[sym] = ph; all_labels.update(ph.keys())
    if not price_hist: return [], []
    labels = sorted(list(all_labels))
    norm_trades = {}
    for sym, arr in trades_by_sym.items():
        tmp = []
        for side, qty, at in arr:
            try: d = datetime.datetime.fromisoformat(at).date().isoformat()
            except Exception: d = str(at).split(" ")[0] if " " in str(at) else labels[0]
            tmp.append((d, side, qty))
        norm_trades[sym] = sorted(tmp, key=lambda x: x[0])
    qty_map = {sym: {} for sym in trades_by_sym}
    running = {sym: 0.0 for sym in trades_by_sym}
    for lb in labels:
        for sym in trades_by_sym:
            for td, side, qty in [t for t in norm_trades[sym] if t[0]==lb]:
                running[sym] += qty if side=="BUY" else -qty
            qty_map[sym][lb] = running[sym]
    values = []
    for lb in labels:
        total = sum(qty_map[sym].get(lb,0)*price_hist.get(sym,{}).get(lb,0) for sym in trades_by_sym)
        values.append(total)
    return labels, values

# ===================== CRYPTO CHARTS =====================
def _chart_font(size: int, bold: bool = False):
    weight = 'bold' if bold else 'normal'
    for name in ['Noto Sans','DejaVu Sans']:
        try: return _fm.FontProperties(family=name, size=size, weight=weight)
        except Exception: pass
    return _fm.FontProperties(size=size, weight=weight)

def gen_portfolio_donut_image(positions, prices, theme: str = DEFAULT_THEME) -> str:
    C = THEMES.get(theme, THEMES[DEFAULT_THEME])
    tokens_data = []
    total_value = 0.0
    for p in positions:
        sym = p.get("symbol","").upper(); qty = float(p.get("qty",0) or 0)
        cg = p.get("cg_id") or cg_guess_id_from_symbol(sym) or ""
        price_now = float(prices.get(cg,0) if cg else 0)
        value = qty * price_now
        if value > 0:
            tokens_data.append({'symbol':sym,'value':value,'qty':qty,'price':price_now})
            total_value += value
    tokens_data.sort(key=lambda x: x['value'], reverse=True)
    if not tokens_data:
        tokens_data = [{'symbol':'N/A','value':1.0,'qty':0,'price':0}]; total_value = 1.0
    if len(tokens_data) > 10:
        others = sum(t['value'] for t in tokens_data[10:])
        tokens_data = tokens_data[:10]
        if others > 0: tokens_data.append({'symbol':'Others','value':others,'qty':0,'price':0})
    for t in tokens_data: t['pct'] = t['value']/total_value*100 if total_value>0 else 0
    n = len(tokens_data); chart_colors = C['chart_colors'][:n]
    fig = plt.figure(figsize=(14,8), facecolor=C['bg'])
    ax_donut = fig.add_axes([0.02,0.08,0.44,0.80])
    ax_legend = fig.add_axes([0.48,0.05,0.50,0.85])
    ax_donut.set_facecolor(C['bg']); ax_legend.set_facecolor(C['bg'])
    ax_legend.axis('off'); ax_donut.set_aspect('equal')
    sizes = [t['value'] for t in tokens_data]
    explode = [0.04 if i<3 else 0 for i in range(n)]
    wedges,_ = ax_donut.pie(sizes, startangle=90, colors=chart_colors, explode=explode,
        wedgeprops=dict(width=0.42,edgecolor=C['bg'],linewidth=2))
    for i,w in enumerate(wedges): w.set_alpha(0.92 if i>=3 else 1.0)
    centre = Circle((0,0),0.53,fc=C['card_bg'],linewidth=2.5,edgecolor=C['accent'],zorder=10)
    ax_donut.add_artist(centre)
    ax_donut.text(0,0.13,'TONG GIA TRI',ha='center',va='center',
        fontproperties=_chart_font(9),color=C['text_secondary'],zorder=11)
    ax_donut.text(0,-0.08,f'${total_value:,.0f}',ha='center',va='center',
        fontproperties=_chart_font(18,bold=True),color=C['accent'],zorder=11)
    for wedge,token in zip(wedges,tokens_data):
        if token['pct'] < 4: continue
        ang = (wedge.theta2+wedge.theta1)/2.0; r=0.70
        ax_donut.text(r*np.cos(np.deg2rad(ang)),r*np.sin(np.deg2rad(ang)),
            f"{token['pct']:.0f}%",ha='center',va='center',
            fontproperties=_chart_font(9,bold=True),color='white',zorder=12)
    ax_legend.text(0.5,0.97,'CHI TIET DANH MUC',transform=ax_legend.transAxes,
        ha='center',fontproperties=_chart_font(14,bold=True),color=C['accent'])
    row_h=0.072; y0=0.90
    for i,token in enumerate(tokens_data):
        if i>=12: break
        y=y0-i*row_h; clr=chart_colors[i] if i<len(chart_colors) else C['accent']
        ax_legend.add_patch(FancyBboxPatch((0.03,y-0.055),0.94,row_h*0.88,
            boxstyle="round,pad=0.008",facecolor=C['card_bg'],edgecolor=clr,
            linewidth=1.5,alpha=0.85,transform=ax_legend.transAxes))
        ax_legend.text(0.07,y-0.018,f"#{i+1}" if i<10 else "~",
            transform=ax_legend.transAxes,fontproperties=_chart_font(9,bold=True),color=clr,va='center')
        ax_legend.text(0.18,y-0.018,token['symbol'],transform=ax_legend.transAxes,
            fontproperties=_chart_font(11,bold=True),color=C['text_primary'],va='center')
        ax_legend.text(0.52,y-0.018,f"${token['value']:,.2f}",transform=ax_legend.transAxes,
            fontproperties=_chart_font(10),color=C['text_primary'],va='center',ha='center')
        ax_legend.text(0.72,y-0.018,f"{token['pct']:.1f}%",transform=ax_legend.transAxes,
            fontproperties=_chart_font(10,bold=True),color=clr,va='center',ha='center')
        fill_w = 0.18*min(token['pct']/100,1.0)
        ax_legend.add_patch(FancyBboxPatch((0.78,y-0.030),0.18,0.012,
            boxstyle="round,pad=0",facecolor=C['card_bg'],edgecolor=C['text_secondary'],
            linewidth=0.8,alpha=0.6,transform=ax_legend.transAxes))
        if fill_w>0:
            ax_legend.add_patch(FancyBboxPatch((0.78,y-0.030),fill_w,0.012,
                boxstyle="round,pad=0",facecolor=clr,edgecolor='none',
                alpha=0.85,transform=ax_legend.transAxes))
    num_tok = len([t for t in tokens_data if t['symbol']!='Others'])
    avg_val = total_value/max(num_tok,1); top_tok = tokens_data[0]
    stats_y = 0.97-12*row_h-0.04
    ax_legend.add_patch(FancyBboxPatch((0.03,stats_y-0.08),0.94,0.095,
        boxstyle="round,pad=0.01",facecolor=C['card_bg'],edgecolor=C['accent'],
        linewidth=1.5,alpha=0.9,transform=ax_legend.transAxes))
    ax_legend.text(0.5,stats_y-0.015,'THONG KE',transform=ax_legend.transAxes,
        ha='center',fontproperties=_chart_font(10,bold=True),color=C['accent'])
    ax_legend.text(0.5,stats_y-0.052,
        f"So token: {num_tok}  |  TB/token: ${avg_val:,.0f}  |  Lon nhat: {top_tok['symbol']} (${top_tok['value']:,.0f})",
        transform=ax_legend.transAxes,ha='center',fontproperties=_chart_font(8.5),color=C['text_secondary'])
    timestamp = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    fig.text(0.5,0.97,'PHAN BO DANH MUC DAU TU CRYPTO',
        ha='center',fontproperties=_chart_font(16,bold=True),color=C['accent'])
    fig.text(0.5,0.01,f'Cap nhat: {timestamp}',
        ha='center',fontproperties=_chart_font(8),color=C['text_secondary'],style='italic')
    fd,path = tempfile.mkstemp(prefix="donut_",suffix=".png"); os.close(fd)
    plt.savefig(path,dpi=180,bbox_inches='tight',facecolor=C['bg'],edgecolor='none')
    plt.close(fig)
    return path

def gen_portfolio_table_image(positions, prices, theme: str = DEFAULT_THEME) -> str:
    C = THEMES.get(theme, THEMES[DEFAULT_THEME])
    rows = []; total_inv = total_val = 0.0
    for p in positions:
        sym=p.get("symbol","").upper(); qty=float(p.get("qty",0) or 0)
        inv=float(max(p.get("invested_usd",0) or 0,0))
        cg=p.get("cg_id") or cg_guess_id_from_symbol(sym) or ""
        price_now=float(prices.get(cg,0) if cg else 0)
        value=qty*price_now; pnl=value-inv; pct=(pnl/inv*100) if inv>1e-12 else 0
        rows.append({'symbol':sym,'qty':f"{qty:,.4f}".rstrip('0').rstrip('.'),'price':f"${price_now:,.2f}",
                     'value':f"${value:,.2f}",'invested':f"${inv:,.2f}",'pnl':f"${pnl:+,.2f}",
                     'pct':pct,'pct_str':f"{pct:+.2f}%"})
        total_inv+=inv; total_val+=value
    total_pnl=total_val-total_inv; total_pct=(total_pnl/total_inv*100) if total_inv>1e-12 else 0
    headers=["Token","So luong","Gia (USD)","Gia tri","Von","Lai/Lo","%"]
    table_data=[headers]
    for r in rows: table_data.append([r['symbol'],r['qty'],r['price'],r['value'],r['invested'],r['pnl'],r['pct_str']])
    if not rows: table_data.append(["N/A","0","$0.00","$0.00","$0.00","$0.00","+0.00%"])
    table_data.append(["TONG CONG","","",f"${total_val:,.2f}",f"${total_inv:,.2f}",f"${total_pnl:+,.2f}",f"{total_pct:+.2f}%"])
    n_rows=len(table_data); fig_h=max(4.5,1.2+0.55*n_rows)
    fig=plt.figure(figsize=(15,fig_h),facecolor=C['bg_dark'])
    ax=fig.add_subplot(111,facecolor=C['bg_dark']); ax.axis('off')
    tbl=ax.table(cellText=table_data,loc='center',cellLoc='center')
    tbl.auto_set_font_size(False); tbl.set_fontsize(10); tbl.scale(1,1.9)
    fp_hdr=_chart_font(10,bold=True); fp_cell=_chart_font(10); fp_bold=_chart_font(10,bold=True)
    for (i,j),cell in tbl.get_celld().items():
        cell.set_edgecolor(C['border']); cell.set_linewidth(0.5)
        if i==0:
            cell.set_facecolor(C['bg_header']); cell.get_text().set_fontproperties(fp_hdr)
            cell.get_text().set_color(C['accent']); cell.set_height(0.08)
        elif i==n_rows-1:
            cell.set_facecolor(C['bg_total']); cell.get_text().set_fontproperties(fp_bold)
            if j in (5,6):
                clr=C['positive'] if total_pnl>0 else (C['negative'] if total_pnl<0 else C['neutral'])
                cell.get_text().set_color(clr)
            else: cell.get_text().set_color(C['accent'])
            cell.set_height(0.08)
        else:
            cell.set_facecolor(C['bg_row1'] if i%2==0 else C['bg_row2'])
            cell.get_text().set_fontproperties(fp_bold if j==0 else fp_cell)
            if j in (5,6) and (i-1)<len(rows):
                pv=rows[i-1]['pct']
                clr=C['positive'] if pv>0 else (C['negative'] if pv<0 else C['neutral'])
                cell.get_text().set_color(clr)
            else: cell.get_text().set_color(C['text_light'])
    pnl_color=C['positive'] if total_pnl>0 else (C['negative'] if total_pnl<0 else C['neutral'])
    ax.text(0.5,1.06,'DANH MUC DAU TU CRYPTO',transform=ax.transAxes,ha='center',
        fontproperties=_chart_font(16,bold=True),color=C['accent'])
    ax.text(0.5,1.02,
        f"Tong: ${total_val:,.2f}  |  Von: ${total_inv:,.2f}  |  Lai/Lo: ${total_pnl:+,.2f} ({total_pct:+.2f}%)",
        transform=ax.transAxes,ha='center',fontproperties=_chart_font(10),color=pnl_color,style='italic')
    if total_inv>0:
        bar_ax=fig.add_axes([0.1,0.01,0.8,0.018])
        bar_ax.set_xlim(0,1); bar_ax.set_ylim(0,1); bar_ax.axis('off')
        bar_ax.add_patch(mpatches.Rectangle((0,0),1,1,facecolor=C['bg_header'],edgecolor=C['border'],linewidth=1))
        ratio=min(abs(total_pct)/100,1.0); bclr=C['positive'] if total_pnl>=0 else C['negative']
        x0=0.5 if total_pnl>=0 else 0.5-ratio*0.5
        bar_ax.add_patch(mpatches.Rectangle((x0,0),ratio*0.5,1,facecolor=bclr,alpha=0.75))
        bar_ax.axvline(0.5,color=C['accent'],linewidth=1.5,alpha=0.8)
    timestamp=datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    ax.text(0.99,-0.04,f"Cap nhat: {timestamp}",transform=ax.transAxes,ha='right',
        fontproperties=_chart_font(8),color=C['text_dim'],style='italic')
    plt.subplots_adjust(top=0.88,bottom=0.06,left=0.02,right=0.98)
    fd,path=tempfile.mkstemp(prefix="table_",suffix=".png"); os.close(fd)
    plt.savefig(path,dpi=180,bbox_inches='tight',facecolor=C['bg_dark'],edgecolor='none')
    plt.close(fig)
    return path
# ===================== FINANCE CHARTS =====================
def get_period_dates(period_choice: str):
    today = datetime.datetime.now()
    if period_choice.startswith("m_"):          # m_YYYYMM
        y, m = int(period_choice[2:6]), int(period_choice[6:8])
        start = datetime.datetime(y, m, 1)
        end = ((start.replace(day=28)+datetime.timedelta(days=4))-datetime.timedelta(days=1)).replace(hour=23,minute=59,second=59)
        suffix = f"tháng {m:02d}/{y}"
    elif period_choice.startswith("y_"):        # y_YYYY
        y = int(period_choice[2:])
        start = datetime.datetime(y, 1, 1)
        end = datetime.datetime(y, 12, 31, 23, 59, 59)
        suffix = f"năm {y}"
    elif period_choice == "week":
        start = (today - datetime.timedelta(days=today.weekday())).replace(hour=0,minute=0,second=0,microsecond=0)
        end = start + datetime.timedelta(days=6,hours=23,minutes=59,seconds=59)
        suffix = f"tuần này ({start.strftime('%d/%m')} - {end.strftime('%d/%m')})"
    elif period_choice == "month":
        start = today.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
        end = ((start.replace(day=28)+datetime.timedelta(days=4)) - datetime.timedelta(days=1)).replace(hour=23,minute=59,second=59)
        suffix = f"tháng này ({today.strftime('%m/%Y')})"
    elif period_choice == "year":
        start = today.replace(month=1,day=1,hour=0,minute=0,second=0,microsecond=0)
        end = today.replace(month=12,day=31,hour=23,minute=59,second=59)
        suffix = f"năm nay ({today.year})"
    elif period_choice == "all":
        start = end = None; suffix = "tất cả thời gian"
    else:
        return None, None, None
    return start, end, suffix

def _initial_period_keyboard(prefix: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Tuần này",    callback_data=f"{prefix}_week")],
        [InlineKeyboardButton("📆 Chọn tháng", callback_data=f"{prefix}_pick_month")],
        [InlineKeyboardButton("🗓️ Chọn năm",  callback_data=f"{prefix}_pick_year")],
        [InlineKeyboardButton("⏰ Toàn bộ",    callback_data=f"{prefix}_all")],
        [InlineKeyboardButton("❌ Hủy bỏ",     callback_data="cancel_action")],
    ])

def _month_picker_keyboard(prefix: str) -> InlineKeyboardMarkup:
    """Last 13 months (current + 12 previous), 3 per row."""
    now = datetime.datetime.now()
    buttons, row = [], []
    for i in range(13):
        m = now.month - i; y = now.year
        while m <= 0: m += 12; y -= 1
        label = f"{m:02d}/{y}"
        row.append(InlineKeyboardButton(label, callback_data=f"{prefix}_m_{y}{m:02d}"))
        if len(row) == 3: buttons.append(row); row = []
    if row: buttons.append(row)
    buttons.append([InlineKeyboardButton("↩️ Quay lại", callback_data=f"{prefix}_back"),
                    InlineKeyboardButton("❌ Hủy",       callback_data="cancel_action")])
    return InlineKeyboardMarkup(buttons)

def _year_picker_keyboard(prefix: str) -> InlineKeyboardMarkup:
    """Last 5 years, all in one row."""
    now = datetime.datetime.now()
    row = [InlineKeyboardButton(str(now.year - i), callback_data=f"{prefix}_y_{now.year - i}") for i in range(5)]
    return InlineKeyboardMarkup([
        row,
        [InlineKeyboardButton("↩️ Quay lại", callback_data=f"{prefix}_back"),
         InlineKeyboardButton("❌ Hủy",       callback_data="cancel_action")],
    ])


def gen_finance_charts(user_id: int, period_choice: str) -> list:
    """Return list of (BytesIO, caption) tuples for all generated charts."""
    start_date, end_date, title_suffix = get_period_dates(period_choice)
    charts = []

    # 1. Donut chi tiêu
    expenses_data = db_list_expenses_grouped(user_id, start_date, end_date)
    if expenses_data:
        labels = [r[0] for r in expenses_data]; sizes = [r[1] for r in expenses_data]
        if sum(sizes) > 0:
            fig, ax = plt.subplots(figsize=(8,8))
            colors = ['#FF7A7A','#3498DB','#F1C40F','#27AE60','#E74C3C','#689F38','#F39C12','#2980B9','#9B59B6','#1ABC9C']
            wedges,_,autotexts = ax.pie(sizes, colors=colors[:len(sizes)], startangle=90,
                counterclock=False, wedgeprops={'width':0.4,'edgecolor':'white','linewidth':2},
                autopct="%1.1f%%", pctdistance=0.85)
            for at in autotexts: at.set_color('black'); at.set_fontsize(14); at.set_fontweight('bold')
            plt.gca().add_artist(plt.Circle((0,0),0.6,fc='white'))
            total = sum(sizes)
            legend_labels = [f'{lb} ({sz/total*100:.1f}%)' for lb,sz in zip(labels,sizes)]
            ax.legend(wedges, legend_labels, loc="center left", bbox_to_anchor=(1,0.5), fontsize=10)
            plt.title(f"Phân bổ Chi tiêu ({title_suffix})", fontsize=14, fontweight='bold', pad=15)
            buf = io.BytesIO(); plt.savefig(buf, format='png', bbox_inches='tight', dpi=150, facecolor='white')
            buf.seek(0); plt.close(fig)
            cap = f"🍩 Chi tiêu {title_suffix}\n💸 Tổng: {total:,.0f} VND"
            charts.append((buf, cap))

    # 2. Cột thu vs chi
    expenses_m, incomes_m = db_get_monthly_report(user_id, start_date, end_date)
    if expenses_m or incomes_m:
        all_months = sorted(set([e[0] for e in expenses_m]+[i[0] for i in incomes_m]))
        exp_dict = {m: a for m,a in expenses_m}; inc_dict = {m: a for m,a in incomes_m}
        exp_vals = [exp_dict.get(m,0) for m in all_months]
        inc_vals = [inc_dict.get(m,0) for m in all_months]
        x = np.arange(len(all_months))
        fig, ax = plt.subplots(figsize=(max(8,len(all_months)*1.2+2), 6))
        bars_inc = ax.bar(x-0.2, inc_vals, 0.4, label='Thu nhập', color='#2E8B57', alpha=0.8, edgecolor='white')
        bars_exp = ax.bar(x+0.2, exp_vals, 0.4, label='Chi tiêu', color='#DC143C', alpha=0.8, edgecolor='white')
        ax.set_xticks(x); ax.set_xticklabels(all_months, rotation=45, ha='right', fontsize=10)
        ax.set_ylabel("Số tiền (VND)"); ax.set_title(f"Thu Chi ({title_suffix})", fontsize=13, fontweight='bold')
        ax.grid(True, alpha=0.3); ax.legend(loc='upper left')
        for bar in bars_inc:
            h = bar.get_height()
            if h > 0: ax.text(bar.get_x()+bar.get_width()/2, h*1.01, f'{int(h):,}', ha='center', va='bottom', fontsize=8, color='#2E8B57', fontweight='bold')
        for bar in bars_exp:
            h = bar.get_height()
            if h > 0: ax.text(bar.get_x()+bar.get_width()/2, h*1.01, f'{int(h):,}', ha='center', va='bottom', fontsize=8, color='#DC143C', fontweight='bold')
        ti = sum(inc_vals); te = sum(exp_vals); nb = ti-te
        bbox_c = '#E8F5E8' if nb>=0 else '#FFE8E8'; tc = '#2E8B57' if nb>=0 else '#DC143C'
        ax.text(0.98,0.98,f"Tổng thu: {ti:,}\nTổng chi: {te:,}\nCân đối: {nb:,}",
            transform=ax.transAxes, va='top', ha='right', fontsize=9,
            bbox=dict(boxstyle='round,pad=0.4',facecolor=bbox_c,alpha=0.8),color=tc,fontweight='bold')
        plt.tight_layout()
        buf = io.BytesIO(); plt.savefig(buf,format='png',bbox_inches='tight',dpi=150,facecolor='white')
        buf.seek(0); plt.close(fig)
        cap = f"📊 Thu chi {title_suffix}\n💰 Thu: {ti:,.0f}  💸 Chi: {te:,.0f}  {'✅' if nb>=0 else '⚠️'} Cân đối: {nb:,.0f}"
        charts.append((buf, cap))

    # 3. Heatmap (week / specific month / specific year / month / year)
    is_day_view  = period_choice in ("week","month") or period_choice.startswith("m_")
    is_year_view = period_choice in ("year",)        or period_choice.startswith("y_")
    if is_day_view or is_year_view:
        conn = db_conn(); cur = conn.cursor()
        if start_date is None: start_date = datetime.datetime(2000,1,1)
        if end_date   is None: end_date   = datetime.datetime.now()
        if period_choice == "week":
            sd = start_date.date(); ed = end_date.date()
            grp = "strftime('%Y-%m-%d',created_at)"; x_label = "day"
            htitle = f"Heatmap Chi Tiêu Theo Ngày ({title_suffix})"
        elif is_day_view:                             # month or m_YYYYMM
            sd = start_date.date(); ed = end_date.date()
            grp = "strftime('%Y-%m-%d',created_at)"; x_label = "day"
            htitle = f"Heatmap Chi Tiêu Theo Ngày ({title_suffix})"
        else:                                         # year or y_YYYY
            sd = start_date.date(); ed = end_date.date()
            grp = "strftime('%Y-%m',created_at)"; x_label = "month"
            htitle = f"Heatmap Chi Tiêu Theo Tháng ({title_suffix})"
        end_incl = ed + datetime.timedelta(days=1)
        cur.execute(f"SELECT {grp},category,SUM(amount) FROM expenses WHERE user_id=? AND created_at>=? AND created_at<? GROUP BY 1,2 ORDER BY 1",
                    (user_id, sd.isoformat(), end_incl.isoformat()))
        rows = cur.fetchall(); conn.close()
        if rows:
            if x_label == "day":
                x_vals = []; d = sd
                while d <= ed: x_vals.append(d.isoformat()); d += datetime.timedelta(days=1)
            else:
                x_vals = []; y,m = sd.year,sd.month
                while (y<ed.year) or (y==ed.year and m<=ed.month):
                    x_vals.append(f"{y:04d}-{m:02d}"); m+=1
                    if m>12: m=1; y+=1
            cats = sorted({r[1] for r in rows})
            dmap = {(x,c):0 for x in x_vals for c in cats}
            for x,c,v in rows:
                if x in x_vals: dmap[(x,c)]=v
            heat = np.array([[dmap[(x,c)] for c in cats] for x in x_vals],dtype=float)
            cmap = mcolors.LinearSegmentedColormap.from_list('heat',['#f0f0f0','#ffcc66','#ff3300'])
            norm = mcolors.Normalize(vmin=0,vmax=heat.max() if heat.size else 1)
            fig,ax = plt.subplots(figsize=(max(8,len(cats)*1.2+2),max(4,len(x_vals)*0.4+2)))
            im = ax.imshow(heat,aspect='auto',cmap=cmap,norm=norm)
            plt.colorbar(im,ax=ax,label='Số tiền (VND)')
            ax.set_xticks(np.arange(len(cats))); ax.set_xticklabels(cats,rotation=45,ha='right',fontsize=9)
            ax.set_yticks(np.arange(len(x_vals))); ax.set_yticklabels(x_vals,fontsize=8)
            ax.set_title(htitle,fontsize=13,fontweight='bold',pad=10)
            maxv = heat.max() if heat.size else 0
            for i in range(len(x_vals)):
                for j in range(len(cats)):
                    val=heat[i,j]
                    if val>0:
                        tc='white' if maxv>0 and val>0.6*maxv else 'black'
                        ax.text(j,i,f"{int(val):,}",ha='center',va='center',fontsize=7,color=tc)
            plt.tight_layout()
            buf=io.BytesIO(); plt.savefig(buf,format='png',bbox_inches='tight',dpi=150,facecolor='white')
            buf.seek(0); plt.close(fig)
            charts.append((buf,f"🔥 {htitle}"))
    return charts

# ===================== HTML DASHBOARD =====================
def _make_html(user_id=None) -> str:  # noqa: C901
    now = datetime.datetime.now()
    ts  = now.strftime("%d/%m/%Y %H:%M:%S")
    uid_note = f"User ID: {user_id}" if user_id else "Thêm ?user_id=ID vào URL để xem dữ liệu"

    def _month_range(y, m):
        s = datetime.datetime(y, m, 1)
        e = ((s.replace(day=28) + datetime.timedelta(days=4)) - datetime.timedelta(days=1)).replace(hour=23, minute=59, second=59)
        return s, e

    cur_y, cur_m = now.year, now.month
    tm_s, tm_e = _month_range(cur_y, cur_m)
    lm_m = cur_m - 1 or 12
    lm_y = cur_y if cur_m > 1 else cur_y - 1
    lm_s, lm_e = _month_range(lm_y, lm_m)
    ty_s = datetime.datetime(cur_y, 1, 1)
    ty_e = datetime.datetime(cur_y, 12, 31, 23, 59, 59)

    crypto_summary_html = crypto_trades_html = ""
    finance_html = income_list_html = expense_list_html = ""
    bhtml = ""
    charts_json = "{}"

    if user_id:
        # ── Crypto summary ───────────────────────────
        try:
            positions = db_crypto_positions(user_id)
            cg_ids    = [p["cg_id"] for p in positions if p["cg_id"]]
            prices    = cg_simple_price_usd(cg_ids) if cg_ids else {}
            crypto_total = 0.0
            crow = ""
            for p in positions:
                sym  = p["symbol"]; qty = float(p["qty"]); inv = float(p["invested_usd"])
                price= float(prices.get(p["cg_id"] or "", 0))
                val  = qty * price; pnl = val - inv
                pct  = (pnl / inv * 100) if inv > 0 else 0
                crypto_total += val
                c = "#00ff88" if pnl >= 0 else "#ff4757"
                crow += (f"<tr><td>{sym}</td><td>{qty:g}</td><td>${price:,.2f}</td>"
                         f"<td>${val:,.2f}</td><td>${inv:,.2f}</td>"
                         f"<td style='color:{c}'>${pnl:+,.2f} ({pct:+.1f}%)</td></tr>")
        except Exception as ex:
            crow = f"<tr><td colspan='6' class='empty'>Lỗi: {ex}</td></tr>"; crypto_total = 0

        crypto_summary_html = f"""
        <div class="card">
          <div class="card-hdr">
            <h2>&#x1FA99; Danh M&#x1EE5;c Crypto</h2>
          </div>
          <table><thead><tr><th>Token</th><th>SL</th><th>Gi&#xE1;</th><th>Gi&#xE1; tr&#x1ECB;</th><th>V&#x1ED1;n</th><th>L&#xE3;i/L&#x1ED7;</th></tr></thead>
          <tbody>{crow or '<tr><td colspan="6" class="empty">Ch&#432;a c&#243; d&#7919;u li&#7879;u</td></tr>'}</tbody></table>
          <p class="total">T&#x1ED5;ng portfolio: <b>${crypto_total:,.2f} USD</b></p>
        </div>"""

        # ── Crypto trades list ───────────────────────
        try:
            crypto_trades = db_crypto_get_trades(user_id)
            ct_rows = ""
            for t in crypto_trades:
                tid, sym, cg, side, qty, price, fee, note, cat = t
                dt = str(cat or "")[:10]
                side_cls = "green" if side == "BUY" else "red"
                safe_note = str(note or "").replace("'", "&#39;").replace('"', "&quot;")
                ct_rows += (
                    f"<tr>"
                    f"<td>{dt}</td>"
                    f"<td><b>{sym}</b></td>"
                    f"<td class='{side_cls}'>{side}</td>"
                    f"<td>{qty:g}</td>"
                    f"<td>${price:,.4f}</td>"
                    f"<td>${fee:,.4f}</td>"
                    f"<td>{safe_note}</td>"
                    f"<td style='white-space:nowrap'>"
                    f"<button class='btn-sm' onclick=\"openEdit('crypto',{{id:{tid},qty:{qty},price:{price},fee:{fee},note:'{safe_note}'}})\">&#x270F;&#xFE0F;</button> "
                    f"<button class='btn-del' onclick='deleteCrypto({tid})'>&#x1F5D1;&#xFE0F;</button>"
                    f"</td></tr>"
                )
        except Exception as ex:
            ct_rows = f"<tr><td colspan='8' class='empty'>L&#x1ED7;i: {ex}</td></tr>"

        crypto_trades_html = f"""
        <div class="card">
          <div class="card-hdr">
            <h2>&#x1F4CB; L&#x1ECB;ch S&#x1EED; Giao D&#x1ECB;ch Crypto</h2>
            <div class="btns">
              <button class="btn-pri" onclick="openAddCrypto()">&#x2795; Th&#xEA;m</button>
              <button class="btn-sec" onclick="exportData('crypto','csv')">&#x2B07;&#xFE0F; CSV</button>
              <button class="btn-sec" onclick="exportData('crypto','excel')">&#x2B07;&#xFE0F; Excel</button>
              <label class="btn-sec" style="cursor:pointer">&#x2B06;&#xFE0F; Nh&#x1EAD;p file
                <input type="file" accept=".csv,.xlsx" style="display:none" onchange="importFile('crypto',this)">
              </label>
            </div>
          </div>
          <div style="overflow-x:auto">
          <table><thead><tr><th>Ng&#xE0;y</th><th>Symbol</th><th>Lo&#x1EA1;i</th><th>SL</th><th>Gi&#xE1;</th><th>Ph&#xED;</th><th>Ghi ch&#xFA;</th><th>Thao t&#xE1;c</th></tr></thead>
          <tbody>{ct_rows or '<tr><td colspan="8" class="empty">Ch&#432;a c&#243; giao d&#x1ECB;ch</td></tr>'}</tbody></table>
          </div>
        </div>"""

        # ── Finance overview ─────────────────────────
        try:
            all_i,  all_e,  all_b  = db_get_combined_summary(user_id)
            tm_i,   tm_e_,  tm_b   = db_get_combined_summary(user_id, tm_s, tm_e)
            lm_i,   lm_e_,  lm_b   = db_get_combined_summary(user_id, lm_s, lm_e)
            ty_i,   ty_e_,  ty_b   = db_get_combined_summary(user_id, ty_s, ty_e)

            tm_ig = db_list_incomes_grouped(user_id, tm_s, tm_e)
            tm_eg = db_list_expenses_grouped(user_id, tm_s, tm_e)
            lm_eg = db_list_expenses_grouped(user_id, lm_s, lm_e)
            budgets = db_get_budgets(user_id)

            # 13-month trend
            tl, ti, te = [], [], []
            for i in range(12, -1, -1):
                m = cur_m - i; y = cur_y
                while m <= 0: m += 12; y -= 1
                s, e = _month_range(y, m)
                inc_, exp_, _ = db_get_combined_summary(user_id, s, e)
                tl.append(f"{m:02d}/{y}"); ti.append(round(inc_)); te.append(round(exp_))

            charts_json = json.dumps({
                "trend":     {"labels": tl, "income": ti, "expense": te},
                "exp_donut": {"labels": [r[0] for r in tm_eg], "values": [round(r[1]) for r in tm_eg]},
                "inc_donut": {"labels": [r[0] for r in tm_ig], "values": [round(r[1]) for r in tm_ig]},
            })

            if budgets:
                bhtml = f"<div class='card'><h2>&#x1F3AF; Ng&#xE2;n S&#xE1;ch Th&#xE1;ng {cur_m:02d}/{cur_y}</h2>"
                for cat, lim in budgets.items():
                    spent = next((r[1] for r in tm_eg if r[0] == cat), 0)
                    pct2  = min(spent / lim * 100, 100) if lim > 0 else 0
                    bc    = "#ff4757" if pct2 >= 100 else "#ffa502" if pct2 >= 80 else "#00ff88"
                    bhtml += (
                        f"<div class='budget-row'>"
                        f"<div class='budget-label'><span>{cat}</span>"
                        f"<span class='{'red' if pct2>=100 else ''}'>{spent:,.0f} / {lim:,.0f} &#x111; ({pct2:.0f}%)</span></div>"
                        f"<div class='budget-bar'><div class='budget-fill' style='width:{pct2:.1f}%;background:{bc}'></div></div>"
                        f"</div>"
                    )
                bhtml += "</div>"

            def _pct(v, total): return f"{v/total*100:.1f}%" if total > 0 else "–"
            def _inc_rows(groups, total):
                if not groups: return "<tr><td colspan='3' class='empty'>Ch&#432;a c&#243; thu nh&#x1EAD;p</td></tr>"
                return "".join(f"<tr><td>{r[0]}</td><td class='amount green'>{r[1]:,.0f} &#x111;</td><td>{_pct(r[1],total)}</td></tr>" for r in groups)
            def _exp_rows(groups, total):
                if not groups: return "<tr><td colspan='3' class='empty'>Ch&#432;a c&#243; chi ti&#xEA;u</td></tr>"
                return "".join(f"<tr><td>{r[0]}</td><td class='amount red'>{r[1]:,.0f} &#x111;</td><td>{_pct(r[1],total)}</td></tr>" for r in groups)

            def _bc(v): return "#00ff88" if v >= 0 else "#ff4757"

            finance_html = f"""
            <div class="card">
              <h2>&#x1F4B5; T&#x1ED5;ng Quan T&#xE0;i Ch&#xED;nh</h2>
              <div class="period-tabs">
                <div class="period-card">
                  <div class="period-label">&#x1F4C5; Th&#xE1;ng {cur_m:02d}/{cur_y}</div>
                  <div class="prow green">Thu: {tm_i:,.0f} &#x111;</div>
                  <div class="prow red">Chi: {tm_e_:,.0f} &#x111;</div>
                  <div class="prow" style="color:{_bc(tm_b)}"><b>D&#432;: {tm_b:,.0f} &#x111;</b></div>
                </div>
                <div class="period-card">
                  <div class="period-label">&#x2B05;&#xFE0F; Th&#xE1;ng {lm_m:02d}/{lm_y}</div>
                  <div class="prow green">Thu: {lm_i:,.0f} &#x111;</div>
                  <div class="prow red">Chi: {lm_e_:,.0f} &#x111;</div>
                  <div class="prow" style="color:{_bc(lm_b)}"><b>D&#432;: {lm_b:,.0f} &#x111;</b></div>
                </div>
                <div class="period-card">
                  <div class="period-label">&#x1F5D3;&#xFE0F; N&#x103;m {cur_y}</div>
                  <div class="prow green">Thu: {ty_i:,.0f} &#x111;</div>
                  <div class="prow red">Chi: {ty_e_:,.0f} &#x111;</div>
                  <div class="prow" style="color:{_bc(ty_b)}"><b>D&#432;: {ty_b:,.0f} &#x111;</b></div>
                </div>
                <div class="period-card">
                  <div class="period-label">&#x23F0; T&#x1EA5;t c&#x1EA3;</div>
                  <div class="prow green">Thu: {all_i:,.0f} &#x111;</div>
                  <div class="prow red">Chi: {all_e:,.0f} &#x111;</div>
                  <div class="prow" style="color:{_bc(all_b)}"><b>D&#432;: {all_b:,.0f} &#x111;</b></div>
                </div>
              </div>
            </div>

            {bhtml}

            <div class="card">
              <h2>&#x1F4C5; Th&#xE1;ng {cur_m:02d}/{cur_y} &#x2014; Thu Chi Theo Danh M&#x1EE5;c</h2>
              <div class="two-col">
                <div>
                  <h3 class="sub-h">&#x1F4B5; Thu nh&#x1EAD;p theo ngu&#x1ED3;n</h3>
                  <table><thead><tr><th>Ngu&#x1ED3;n</th><th>S&#x1ED1; ti&#x1EC1;n</th><th>%</th></tr></thead>
                  <tbody>{_inc_rows(tm_ig, tm_i)}</tbody></table>
                </div>
                <div>
                  <h3 class="sub-h">&#x1F4B8; Chi ti&#xEA;u theo danh m&#x1EE5;c</h3>
                  <table><thead><tr><th>Danh m&#x1EE5;c</th><th>S&#x1ED1; ti&#x1EC1;n</th><th>%</th></tr></thead>
                  <tbody>{_exp_rows(tm_eg, tm_e_)}</tbody></table>
                </div>
              </div>
              <div class="two-col chart-row">
                <div><canvas id="incDonut"></canvas></div>
                <div><canvas id="expDonut"></canvas></div>
              </div>
            </div>

            <div class="card">
              <h2>&#x2B05;&#xFE0F; Th&#xE1;ng {lm_m:02d}/{lm_y} &#x2014; Chi Ti&#xEA;u Theo Danh M&#x1EE5;c</h2>
              <table><thead><tr><th>Danh m&#x1EE5;c</th><th>S&#x1ED1; ti&#x1EC1;n</th><th>%</th></tr></thead>
              <tbody>{_exp_rows(lm_eg, lm_e_)}</tbody></table>
            </div>

            <div class="card">
              <h2>&#x1F4C8; Xu H&#432;&#x1EDB;ng Thu Chi &#x2014; 13 Th&#xE1;ng G&#x1EA7;n Nh&#x1EA5;t</h2>
              <div class="chart-full"><canvas id="trendChart"></canvas></div>
            </div>"""

        except Exception as ex:
            finance_html = f"<div class='card'><p style='color:#ff4757'>L&#x1ED7;i: {ex}</p></div>"

        income_list_html = ""
        expense_list_html = ""

    # ── Assemble HTML ────────────────────────────────
    return f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Univer Bot Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#0a0e27;color:#e8e8e8;font-family:'Segoe UI',sans-serif;padding:20px;max-width:1400px;margin:0 auto}}
h1{{color:#00d4ff;text-align:center;margin-bottom:6px;font-size:1.8em}}
.subtitle{{text-align:center;color:#94a3b8;margin-bottom:24px;font-size:.85em}}
.card{{background:#151932;border-radius:12px;padding:20px;margin-bottom:20px;border:1px solid #2a2a3e}}
.card h2{{color:#00d4ff;font-size:1.15em;margin:0}}
h3.sub-h{{color:#94a3b8;font-size:.88em;margin-bottom:8px;font-weight:500;text-transform:uppercase;letter-spacing:.04em}}
table{{width:100%;border-collapse:collapse;font-size:.88em}}
th{{background:#16213e;color:#00d4ff;padding:9px 10px;text-align:left;font-weight:500}}
td{{padding:8px 10px;border-bottom:1px solid #1e2a3e}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:#1a2240}}
.amount{{text-align:right;font-family:'Courier New',monospace;font-weight:600}}
.green{{color:#00ff88}}.red{{color:#ff4757}}
.empty{{color:#535c68;text-align:center;padding:16px;font-style:italic}}
.total{{color:#00d4ff;font-weight:bold;margin-top:12px;font-size:.95em}}
.period-tabs{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px}}
.period-card{{background:#0f3460;border-radius:10px;padding:14px}}
.period-label{{color:#94a3b8;font-size:.78em;margin-bottom:8px;font-weight:600;text-transform:uppercase;letter-spacing:.05em}}
.prow{{font-size:.88em;margin:4px 0}}
.two-col{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}
.chart-row{{margin-top:24px}}
.chart-row canvas{{max-height:260px}}
.chart-full{{height:300px;position:relative}}
.budget-row{{margin:10px 0}}
.budget-label{{display:flex;justify-content:space-between;font-size:.84em;margin-bottom:5px;color:#b0b8cc}}
.budget-bar{{background:#1e2a3e;border-radius:6px;height:10px;overflow:hidden}}
.budget-fill{{height:100%;border-radius:6px}}
footer{{text-align:center;color:#535c68;margin-top:20px;font-size:.78em;padding-bottom:16px}}
.btn-pri{{background:#00d4ff;color:#0a0e27;border:none;padding:8px 18px;border-radius:6px;cursor:pointer;font-weight:600}}
.btn-pri:hover{{background:#00b8d9}}
.btn-sec{{background:#2a2a3e;color:#e8e8e8;border:1px solid #3a3a5e;padding:8px 18px;border-radius:6px;cursor:pointer}}
.btn-sec:hover{{background:#3a3a5e}}
.btn-sm{{background:#1e2a4a;color:#94a3b8;border:1px solid #2a3a5e;padding:4px 10px;border-radius:5px;cursor:pointer;font-size:.82em}}
.btn-sm:hover{{background:#2a3a5e}}
.btn-del{{background:#2d1515;color:#ff4757;border:1px solid #ff4757;padding:4px 10px;border-radius:5px;cursor:pointer;font-size:.82em}}
.btn-del:hover{{background:#ff4757;color:#fff}}
.card-hdr{{display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;border-bottom:1px solid #2a2a3e;padding-bottom:10px}}
.card-hdr h2{{color:#00d4ff;font-size:1.15em;margin:0}}
.card-hdr .btns{{display:flex;gap:8px;flex-wrap:wrap}}
label{{display:block;margin-bottom:10px;color:#94a3b8;font-size:.88em}}
.finput{{width:100%;margin-top:4px;background:#0f1730;color:#e8e8e8;border:1px solid #2a2a3e;padding:8px;border-radius:6px;font-size:.9em}}
.finput:focus{{outline:none;border-color:#00d4ff}}
.filter-bar{{display:flex;flex-wrap:wrap;gap:8px;align-items:flex-end;margin-bottom:14px;padding:14px;background:#0f1730;border-radius:8px;border:1px solid #2a2a3e}}
.filter-bar .fgroup{{display:flex;flex-direction:column;gap:4px;min-width:110px}}
.filter-bar .fgroup label{{margin:0;font-size:.78em;color:#94a3b8}}
.filter-bar input,.filter-bar select{{background:#151932;color:#e8e8e8;border:1px solid #2a3a5e;padding:6px 8px;border-radius:5px;font-size:.84em}}
.filter-bar input:focus,.filter-bar select:focus{{outline:none;border-color:#00d4ff}}
.quick-btns{{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:10px}}
.qbtn{{background:#16213e;color:#94a3b8;border:1px solid #2a3a5e;padding:5px 12px;border-radius:20px;cursor:pointer;font-size:.82em;transition:all .15s}}
.qbtn:hover,.qbtn.active{{background:#00d4ff;color:#0a0e27;border-color:#00d4ff;font-weight:600}}
.tbl-footer{{display:flex;justify-content:space-between;align-items:center;padding:10px 0 0;border-top:1px solid #2a2a3e;margin-top:6px;font-size:.88em;color:#94a3b8}}
.tbl-footer .total-val{{font-size:1em;font-weight:700}}
@media(max-width:768px){{
  .two-col{{grid-template-columns:1fr}}
  .period-tabs{{grid-template-columns:1fr 1fr}}
  .chart-full{{height:220px}}
  .card-hdr{{flex-direction:column;align-items:flex-start;gap:10px}}
  .filter-bar{{flex-direction:column}}
  .filter-bar .fgroup{{width:100%}}
  #financeTables{{grid-template-columns:1fr!important}}
}}
</style>
</head>
<body>
<h1>&#x1F916; Univer Bot Dashboard</h1>
<div class="subtitle">{uid_note} &middot; C&#x1EAD;p nh&#x1EAD;t: {ts}</div>
<div style="text-align:center;margin-bottom:18px;display:flex;gap:10px;justify-content:center;flex-wrap:wrap">
  <button class="btn-sec" onclick="downloadDB()" style="font-size:.82em">&#x1F4BE; T&#x1EA3;i xu&#x1ED1;ng DB</button>
  <label class="btn-sec" style="cursor:pointer;font-size:.82em">&#x2B06;&#xFE0F; Upload DB (ghi &#x111;&#xE8;)
    <input type="file" accept=".db,.sqlite,.sqlite3" style="display:none" onchange="uploadDB(this)">
  </label>
</div>

{crypto_summary_html}
{crypto_trades_html}
{finance_html}

<!-- Finance filter + tables (dynamic) -->
<div class="card" id="financeSection">
  <div class="card-hdr">
    <h2>&#x1F4CA; L&#x1ECD;c Thu Chi</h2>
    <div class="btns">
      <button class="btn-pri" onclick="openAddFinance('income')">&#x2795; Thu nh&#x1EAD;p</button>
      <button class="btn-pri" style="background:#ff4757" onclick="openAddFinance('expense')">&#x2795; Chi ti&#xEA;u</button>
      <button class="btn-sec" onclick="exportData('finance','csv')">&#x2B07;&#xFE0F; CSV</button>
      <button class="btn-sec" onclick="exportData('finance','excel')">&#x2B07;&#xFE0F; Excel</button>
      <label class="btn-sec" style="cursor:pointer">&#x2B06;&#xFE0F; Nh&#x1EAD;p file<input type="file" accept=".csv,.xlsx" style="display:none" onchange="importFile('finance',this)"></label>
    </div>
  </div>

  <div class="quick-btns">
    <button class="qbtn" onclick="setQuick('thismonth')">Th&#xE1;ng n&#xE0;y</button>
    <button class="qbtn" onclick="setQuick('lastmonth')">Th&#xE1;ng tr&#432;&#x1EDB;c</button>
    <button class="qbtn" onclick="setQuick('thisquarter')">Qu&#xFD; n&#xE0;y</button>
    <button class="qbtn" onclick="setQuick('thisyear')">N&#x103;m nay</button>
    <button class="qbtn" onclick="setQuick('lastyear')">N&#x103;m ngo&#xE1;i</button>
    <button class="qbtn" onclick="setQuick('all')">T&#x1EA5;t c&#x1EA3;</button>
  </div>

  <div class="filter-bar">
    <div class="fgroup"><label>T&#x1EEB; ng&#xE0;y</label><input type="date" id="fDateFrom"></div>
    <div class="fgroup"><label>&#x110;&#x1EBF;n ng&#xE0;y</label><input type="date" id="fDateTo"></div>
    <div class="fgroup" style="min-width:150px"><label>Ngu&#x1ED3;n thu nh&#x1EAD;p</label>
      <select id="fIncCat"><option value="">-- T&#x1EA5;t c&#x1EA3; --</option></select></div>
    <div class="fgroup" style="min-width:150px"><label>Danh m&#x1EE5;c chi ti&#xEA;u</label>
      <select id="fExpCat"><option value="">-- T&#x1EA5;t c&#x1EA3; --</option></select></div>
    <div class="fgroup" style="min-width:160px"><label>T&#xEC;m ki&#x1EBF;m</label>
      <input type="text" id="fSearch" placeholder="Ghi ch&#xFA;, ngu&#x1ED3;n..." onkeydown="if(event.key==='Enter')applyFilter()"></div>
    <div class="fgroup" style="justify-content:flex-end;padding-top:16px">
      <button class="btn-pri" onclick="applyFilter()" style="padding:6px 16px">&#x1F50D; L&#x1ECD;c</button>
    </div>
  </div>

  <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px" id="financeTables">
    <div>
      <h3 class="sub-h">&#x1F4B0; Thu Nh&#x1EAD;p</h3>
      <div style="overflow-x:auto">
        <table><thead><tr><th>Ng&#xE0;y</th><th>Ngu&#x1ED3;n</th><th>Ghi ch&#xFA;</th><th>S&#x1ED1; ti&#x1EC1;n</th><th></th></tr></thead>
        <tbody id="incTbody"><tr><td colspan="5" class="empty">&#x26A1; &#x110;ang t&#x1EA3;i...</td></tr></tbody></table>
      </div>
      <div class="tbl-footer"><span id="incCount">0 kho&#x1EA3;n</span><span class="total-val green" id="incTotal">0 &#x111;</span></div>
    </div>
    <div>
      <h3 class="sub-h">&#x1F4B8; Chi Ti&#xEA;u</h3>
      <div style="overflow-x:auto">
        <table><thead><tr><th>Ng&#xE0;y</th><th>Danh m&#x1EE5;c</th><th>Ghi ch&#xFA;</th><th>S&#x1ED1; ti&#x1EC1;n</th><th></th></tr></thead>
        <tbody id="expTbody"><tr><td colspan="5" class="empty">&#x26A1; &#x110;ang t&#x1EA3;i...</td></tr></tbody></table>
      </div>
      <div class="tbl-footer"><span id="expCount">0 kho&#x1EA3;n</span><span class="total-val red" id="expTotal">0 &#x111;</span></div>
    </div>
  </div>
  <div class="tbl-footer" style="margin-top:14px;border-top:2px solid #00d4ff;padding-top:12px">
    <span style="color:#e8e8e8;font-weight:600">C&#xE2;n &#x111;&#x1ED1;i</span>
    <span class="total-val" id="netTotal" style="color:#00d4ff">0 &#x111;</span>
  </div>
</div>

<footer>Univer All-in-One Bot &middot; Port {HTML_PORT}</footer>

<!-- Modal -->
<div id="modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.75);z-index:9999;align-items:center;justify-content:center">
  <div style="background:#151932;border:1px solid #2a2a3e;border-radius:12px;padding:24px;width:90%;max-width:480px;max-height:90vh;overflow-y:auto">
    <h3 id="modalTitle" style="color:#00d4ff;margin-bottom:16px"></h3>
    <div id="modalBody"></div>
    <div style="margin-top:20px;display:flex;gap:10px;justify-content:flex-end">
      <button onclick="closeModal()" class="btn-sec">H&#x1EE7;y</button>
      <button onclick="submitModal()" class="btn-pri">L&#432;u</button>
    </div>
  </div>
</div>

<script>
const D={charts_json};
const PAL=['#00d4ff','#00ff88','#ffa502','#ff4757','#a29bfe','#fd79a8','#55efc4','#fdcb6e','#e17055','#74b9ff','#b2bec3','#fab1a0'];
const fmtK=v=>v>=1e9?(v/1e9).toFixed(1)+'B':v>=1e6?(v/1e6).toFixed(1)+'M':v>=1e3?(v/1e3).toFixed(0)+'K':v;

if(D.trend){{
  new Chart(document.getElementById('trendChart'),{{
    type:'bar',
    data:{{labels:D.trend.labels,datasets:[
      {{label:'Thu nhập',data:D.trend.income,backgroundColor:'rgba(0,255,136,.65)',borderColor:'#00ff88',borderWidth:1,borderRadius:4}},
      {{label:'Chi ti\xeau',data:D.trend.expense,backgroundColor:'rgba(255,71,87,.65)',borderColor:'#ff4757',borderWidth:1,borderRadius:4}},
    ]}},
    options:{{responsive:true,maintainAspectRatio:false,
      plugins:{{legend:{{labels:{{color:'#e8e8e8',font:{{size:12}}}}}}}},
      scales:{{
        x:{{ticks:{{color:'#94a3b8',font:{{size:11}}}},grid:{{color:'#1a2240'}}}},
        y:{{ticks:{{color:'#94a3b8',font:{{size:11}},callback:fmtK}},grid:{{color:'#1a2240'}}}}
      }}
    }}
  }});
}}
function donut(id,title,data){{
  const el=document.getElementById(id); if(!el||!data||!data.labels.length) return;
  new Chart(el,{{
    type:'doughnut',
    data:{{labels:data.labels,datasets:[{{data:data.values,backgroundColor:PAL,borderWidth:0,hoverOffset:6}}]}},
    options:{{responsive:true,maintainAspectRatio:false,
      plugins:{{
        legend:{{position:'bottom',labels:{{color:'#e8e8e8',font:{{size:11}},padding:10}}}},
        title:{{display:true,text:title,color:'#94a3b8',font:{{size:12}}}}
      }}
    }}
  }});
}}
donut('incDonut','Thu nhập theo nguồn',D.inc_donut);
donut('expDonut','Chi ti\xeau theo danh mục',D.exp_donut);

// ── CRUD JS ──────────────────────────────────────────
const UID=new URLSearchParams(location.search).get('user_id')||'';
const TOK=new URLSearchParams(location.search).get('token')||'';

async function api(path,body){{
  const r=await fetch(`${{path}}?user_id=${{UID}}&token=${{TOK}}`,{{
    method:'POST',
    headers:{{'Content-Type':'application/json'}},
    body:JSON.stringify(body)
  }});
  return r.json();
}}

function exportData(scope,fmt){{
  window.location.href=`/api/export?user_id=${{UID}}&token=${{TOK}}&scope=${{scope}}&format=${{fmt}}`;
}}

function importFile(scope,input){{
  const file=input.files[0]; if(!file) return;
  const fmt=file.name.endsWith('.xlsx')?'excel':'csv';
  const reader=new FileReader();
  reader.onload=e=>{{
    const bytes=new Uint8Array(e.target.result);
    let bin=''; bytes.forEach(b=>bin+=String.fromCharCode(b));
    const b64=btoa(bin);
    api('/api/import',{{scope,format:fmt,data:b64}})
      .then(r=>{{alert(`Nhập xong: ${{r.imported||0}} th\xe0nh c\xf4ng, ${{r.failed||0}} lỗi`);location.reload();  }})
      .catch(e=>alert('Lỗi: '+e));
  }};
  reader.readAsArrayBuffer(file);
}}

let _ms={{}};

function closeModal(){{document.getElementById('modal').style.display='none';}}

function openAddCrypto(){{
  _ms={{action:'add',scope:'crypto'}};
  document.getElementById('modalTitle').textContent='➕ Th\xeam Giao Dịch Crypto';
  const today=new Date().toISOString().slice(0,10);
  document.getElementById('modalBody').innerHTML=`
    <label>Symbol<input id="f_symbol" class="finput" placeholder="BTC"></label>
    <label>Loại<select id="f_side" class="finput"><option value="BUY">Mua</option><option value="SELL">B\xe1n</option></select></label>
    <label>Số lượng<input id="f_qty" class="finput" type="number" step="any"></label>
    <label>Gi\xe1 (USD)<input id="f_price" class="finput" type="number" step="any"></label>
    <label>Ph\xed (USD)<input id="f_fee" class="finput" type="number" step="any" value="0"></label>
    <label>CoinGecko ID (tỹ chọn nếu trống)<input id="f_cgid" class="finput" placeholder="bitcoin"></label>
    <label>Ghi ch\xfa<input id="f_note" class="finput"></label>
    <label>Ng\xe0y<input id="f_date" class="finput" type="date" value="${{today}}"></label>
  `;
  document.getElementById('modal').style.display='flex';
}}

function openEdit(scope,data){{
  _ms={{action:'edit',scope,data}};
  if(scope==='crypto'){{
    document.getElementById('modalTitle').textContent='✏️ Sửa Giao Dịch Crypto';
    document.getElementById('modalBody').innerHTML=`
      <label>Số lượng<input id="f_qty" class="finput" type="number" step="any" value="${{data.qty}}"></label>
      <label>Gi\xe1 (USD)<input id="f_price" class="finput" type="number" step="any" value="${{data.price}}"></label>
      <label>Ph\xed (USD)<input id="f_fee" class="finput" type="number" step="any" value="${{data.fee}}"></label>
      <label>Ghi ch\xfa<input id="f_note" class="finput" value="${{data.note||''}}"></label>
    `;
  }} else if(scope==='income'){{
    document.getElementById('modalTitle').textContent='✏️ Sửa Thu Nhập';
    document.getElementById('modalBody').innerHTML=`
      <label>Số tiền<input id="f_amount" class="finput" type="number" step="any" value="${{data.amount}}"></label>
      <label>Nguồn<input id="f_cat" class="finput" value="${{data.cat||''}}"></label>
      <label>Ghi ch\xfa<input id="f_note" class="finput" value="${{data.note||''}}"></label>
      <label>Ng\xe0y<input id="f_date" class="finput" type="date" value="${{(data.date||'').slice(0,10)}}"></label>
    `;
  }} else {{
    document.getElementById('modalTitle').textContent='✏️ Sửa Chi Ti\xeau';
    document.getElementById('modalBody').innerHTML=`
      <label>Số tiền<input id="f_amount" class="finput" type="number" step="any" value="${{data.amount}}"></label>
      <label>Danh mục<input id="f_cat" class="finput" value="${{data.cat||''}}"></label>
      <label>Ghi ch\xfa<input id="f_note" class="finput" value="${{data.note||''}}"></label>
      <label>Ng\xe0y<input id="f_date" class="finput" type="date" value="${{(data.date||'').slice(0,10)}}"></label>
    `;
  }}
  document.getElementById('modal').style.display='flex';
}}

function openAddFinance(ttype){{
  _ms={{action:'add',scope:'finance',ttype}};
  const label=ttype==='income'?'Thu Nhập':'Chi Ti\xeau';
  const catLabel=ttype==='income'?'Nguồn':'Danh mục';
  const today=new Date().toISOString().slice(0,10);
  document.getElementById('modalTitle').textContent=`➕ Th\xeam ${{label}}`;
  document.getElementById('modalBody').innerHTML=`
    <label>Số tiền (VND)<input id="f_amount" class="finput" type="number" step="any"></label>
    <label>${{catLabel}}<input id="f_cat" class="finput"></label>
    <label>Ghi ch\xfa<input id="f_note" class="finput"></label>
    <label>Ng\xe0y<input id="f_date" class="finput" type="date" value="${{today}}"></label>
  `;
  document.getElementById('modal').style.display='flex';
}}

async function submitModal(){{
  const s=_ms;
  try{{
    let r;
    if(s.scope==='crypto'){{
      if(s.action==='add'){{
        r=await api('/api/crypto/add',{{
          symbol:v('f_symbol'),side:v('f_side'),
          qty:v('f_qty'),price:v('f_price'),fee:v('f_fee')||0,
          cg_id:v('f_cgid'),note:v('f_note'),date:v('f_date')
        }});
      }}else{{
        r=await api('/api/crypto/update',{{
          id:s.data.id,qty:v('f_qty'),price:v('f_price'),
          fee:v('f_fee')||0,note:v('f_note')
        }});
      }}
    }}else if(s.scope==='finance'){{
      r=await api('/api/finance/add',{{
        ttype:s.ttype,amount:v('f_amount'),
        category:v('f_cat'),note:v('f_note'),date:v('f_date')
      }});
    }}else{{
      r=await api('/api/finance/update',{{
        id:s.data.id,ttype:s.scope,
        amount:v('f_amount'),category:v('f_cat'),
        note:v('f_note'),date:v('f_date')
      }});
    }}
    if(r.ok){{
      closeModal();
      if(s.scope==='crypto'||s.scope==='crypto_edit') location.reload();
      else applyFilter();
    }}else alert('Lỗi: '+(r.error||'unknown'));
  }}catch(e){{alert('Lỗi: '+e);}}
}}

async function deleteCrypto(id){{
  if(!confirm('X\xf3a giao dịch n\xe0y?')) return;
  const r=await api('/api/crypto/delete',{{id}});
  if(r.ok) location.reload(); else alert('Lỗi: '+r.error);
}}

async function deleteFinance(id,ttype){{
  if(!confirm('X\xf3a giao dịch n\xe0y?')) return;
  const r=await api('/api/finance/delete',{{id,ttype}});
  if(r.ok) applyFilter(); else alert('Lỗi: '+r.error);
}}

function v(id){{const el=document.getElementById(id);return el?el.value:'';}}

document.getElementById('modal').addEventListener('click',function(e){{
  if(e.target===this) closeModal();
}});

// ── Finance filter ────────────────────────────────────
const fmtVND=n=>n.toLocaleString('vi-VN')+'₫';

function isoToday(){{return new Date().toISOString().slice(0,10);}}
function isoDate(d){{return d.toISOString().slice(0,10);}}

function setQuick(period){{
  document.querySelectorAll('.qbtn').forEach(b=>b.classList.remove('active'));
  event.target.classList.add('active');
  const now=new Date();
  let from,to;
  if(period==='thismonth'){{
    from=new Date(now.getFullYear(),now.getMonth(),1);
    to=new Date(now.getFullYear(),now.getMonth()+1,0);
  }}else if(period==='lastmonth'){{
    from=new Date(now.getFullYear(),now.getMonth()-1,1);
    to=new Date(now.getFullYear(),now.getMonth(),0);
  }}else if(period==='thisquarter'){{
    const q=Math.floor(now.getMonth()/3);
    from=new Date(now.getFullYear(),q*3,1);
    to=new Date(now.getFullYear(),q*3+3,0);
  }}else if(period==='thisyear'){{
    from=new Date(now.getFullYear(),0,1);
    to=new Date(now.getFullYear(),11,31);
  }}else if(period==='lastyear'){{
    from=new Date(now.getFullYear()-1,0,1);
    to=new Date(now.getFullYear()-1,11,31);
  }}else{{
    document.getElementById('fDateFrom').value='';
    document.getElementById('fDateTo').value='';
    applyFilter(); return;
  }}
  document.getElementById('fDateFrom').value=isoDate(from);
  document.getElementById('fDateTo').value=isoDate(to);
  applyFilter();
}}

async function loadCategories(){{
  try{{
    const r=await fetch(`/api/finance/categories?user_id=${{UID}}&token=${{TOK}}`);
    const d=await r.json();
    const incSel=document.getElementById('fIncCat');
    const expSel=document.getElementById('fExpCat');
    (d.income||[]).forEach(c=>{{const o=document.createElement('option');o.value=c;o.textContent=c;incSel.appendChild(o);}});
    (d.expense||[]).forEach(c=>{{const o=document.createElement('option');o.value=c;o.textContent=c;expSel.appendChild(o);}});
  }}catch(e){{}}
}}

async function applyFilter(){{
  const from=v('fDateFrom');
  const to=v('fDateTo');
  const incCat=v('fIncCat');
  const expCat=v('fExpCat');
  const search=v('fSearch');

  const qInc=new URLSearchParams({{user_id:UID,token:TOK,ttype:'income',date_from:from,date_to:to,category:incCat,search,limit:500}});
  const qExp=new URLSearchParams({{user_id:UID,token:TOK,ttype:'expense',date_from:from,date_to:to,category:expCat,search,limit:500}});

  document.getElementById('incTbody').innerHTML='<tr><td colspan="5" class="empty">Đang tải...</td></tr>';
  document.getElementById('expTbody').innerHTML='<tr><td colspan="5" class="empty">Đang tải...</td></tr>';

  const [rInc,rExp]=await Promise.all([
    fetch(`/api/finance/list?${{qInc}}`).then(r=>r.json()),
    fetch(`/api/finance/list?${{qExp}}`).then(r=>r.json())
  ]);

  renderTable('incTbody',rInc.income||[],'income');
  renderTable('expTbody',rExp.expense||[],'expense');

  const incT=rInc.income_total||0;
  const expT=rExp.expense_total||0;
  const net=incT-expT;
  document.getElementById('incCount').textContent=(rInc.income||[]).length+' khoản';
  document.getElementById('expCount').textContent=(rExp.expense||[]).length+' khoản';
  document.getElementById('incTotal').textContent=fmtVND(incT);
  document.getElementById('expTotal').textContent=fmtVND(expT);
  const netEl=document.getElementById('netTotal');
  netEl.textContent=(net>=0?'+':'')+fmtVND(net);
  netEl.style.color=net>=0?'#00ff88':'#ff4757';
}}

function renderTable(tbodyId,rows,ttype){{
  const tb=document.getElementById(tbodyId);
  if(!rows.length){{tb.innerHTML='<tr><td colspan="5" class="empty">Không có dữ liệu</td></tr>';return;}}
  const colorCls=ttype==='income'?'green':'red';
  tb.innerHTML=rows.map(r=>{{
    const safeC=(r.cat||'').replace(/'/g,"&#39;").replace(/"/g,"&quot;");
    const safeN=(r.note||'').replace(/'/g,"&#39;").replace(/"/g,"&quot;");
    return `<tr>
      <td>${{r.date}}</td>
      <td>${{safeC}}</td>
      <td>${{safeN}}</td>
      <td class="amount ${{colorCls}}">${{r.amount.toLocaleString('vi-VN')}} ₫</td>
      <td style="white-space:nowrap">
        <button class="btn-sm" onclick="openEdit('${{ttype}}',{{id:${{r.id}},amount:${{r.amount}},cat:'${{safeC}}',note:'${{safeN}}',date:'${{r.date}}'}})">✏️</button>
        <button class="btn-del" onclick="deleteFinanceR(${{r.id}},'${{ttype}}')">🗑️</button>
      </td></tr>`;
  }}).join('');
}}

async function deleteFinanceR(id,ttype){{
  if(!confirm('Xóa giao dịch này?')) return;
  const r=await api('/api/finance/delete',{{id,ttype}});
  if(r.ok) applyFilter(); else alert('Lỗi: '+r.error);
}}

// ── DB backup / restore ──────────────────────────────
function downloadDB(){{
  window.location.href=`/api/db/download?user_id=${{UID}}&token=${{TOK}}`;
}}

function uploadDB(input){{
  const file=input.files[0]; if(!file) return;
  if(!confirm(`Ghi đè toàn bộ dữ liệu bằng file "${{file.name}}"?\nHành động này không thể hoàn tác!`)) return;
  const reader=new FileReader();
  reader.onload=async e=>{{
    const bytes=new Uint8Array(e.target.result);
    let bin=''; bytes.forEach(b=>bin+=String.fromCharCode(b));
    const b64=btoa(bin);
    try{{
      const r=await api('/api/db/upload',{{data:b64}});
      if(r.ok){{ alert('Upload thành công! Trang sẽ tải lại.'); location.reload(); }}
      else alert('Lỗi: '+(r.error||'unknown'));
    }}catch(e){{ alert('Lỗi: '+e); }}
  }};
  reader.readAsArrayBuffer(file);
  input.value='';
}}

// init
loadCategories();
// default: this month
(function(){{
  const now=new Date();
  document.getElementById('fDateFrom').value=isoDate(new Date(now.getFullYear(),now.getMonth(),1));
  document.getElementById('fDateTo').value=isoDate(new Date(now.getFullYear(),now.getMonth()+1,0));
  document.querySelector('.qbtn').classList.add('active');
  applyFilter();
}})();
</script>
</body>
</html>"""

class _DashboardHandler(BaseHTTPRequestHandler):

    # ── helpers ────────────────────────────────────────

    def log_message(self, *a):
        pass

    def _send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def _send_403(self):
        body = b"<h2>403 Forbidden</h2><p>Thieu token hoac token sai.</p>"
        self.send_response(403)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _check_auth(self, qs):
        if not DASHBOARD_SECRET:
            return True
        return qs.get("token", [""])[0] == DASHBOARD_SECRET

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b"{}"
        return json.loads(raw.decode("utf-8"))

    def _uid(self, qs):
        try:
            return int(qs["user_id"][0])
        except Exception:
            return None

    # ── GET ────────────────────────────────────────────

    def do_GET(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        if not self._check_auth(qs):
            self._send_403(); return
        uid = self._uid(qs)
        if parsed.path == "/api/export":
            self._handle_export(uid, qs); return
        if parsed.path == "/api/finance/list":
            self._api_finance_list(uid, qs); return
        if parsed.path == "/api/finance/categories":
            self._send_json(db_get_finance_categories(uid) if uid else {"income":[],"expense":[]}); return
        if parsed.path == "/api/db/download":
            self._handle_db_download(); return
        html = _make_html(uid).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(html)))
        self.end_headers()
        self.wfile.write(html)

    # ── POST ───────────────────────────────────────────

    def do_POST(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        if not self._check_auth(qs):
            self._send_403(); return
        uid = self._uid(qs)
        try:
            body = self._read_body()
        except Exception as exc:
            self._send_json({"error": f"JSON parse error: {exc}"}, 400); return
        path = parsed.path
        try:
            if path == "/api/crypto/add":
                self._api_crypto_add(uid, body)
            elif path == "/api/crypto/delete":
                self._api_crypto_delete(body)
            elif path == "/api/crypto/update":
                self._api_crypto_update(body)
            elif path == "/api/finance/add":
                self._api_finance_add(uid, body)
            elif path == "/api/finance/delete":
                self._api_finance_delete(body)
            elif path == "/api/finance/update":
                self._api_finance_update(body)
            elif path == "/api/import":
                self._api_import(uid, body); return
            elif path == "/api/db/upload":
                self._handle_db_upload(body); return
            else:
                self._send_json({"error": "unknown path"}, 404)
        except Exception as exc:
            self._send_json({"error": str(exc), "trace": traceback.format_exc()}, 500)

    # ── finance list (filterable) ─────────────────────

    def _api_finance_list(self, uid, qs):
        if not uid:
            self._send_json({"error": "no user_id"}); return
        ttype     = qs.get("ttype",     ["all"])[0]
        date_from = qs.get("date_from", [""])[0].strip() or None
        date_to   = qs.get("date_to",   [""])[0].strip() or None
        category  = qs.get("category",  [""])[0].strip() or None
        search    = qs.get("search",    [""])[0].strip() or None
        limit     = int(qs.get("limit", ["500"])[0])
        result = {}
        if ttype in ("income", "all"):
            rows = db_get_incomes_filtered(uid, date_from, date_to, category if ttype=="income" else None, search, limit)
            result["income"] = [{"id":r[0],"amount":r[1],"cat":r[2],"note":r[3],"date":str(r[4])[:10]} for r in rows]
            result["income_total"] = sum(r[1] for r in rows)
        if ttype in ("expense", "all"):
            rows = db_get_expenses_filtered(uid, date_from, date_to, category if ttype=="expense" else None, search, limit)
            result["expense"] = [{"id":r[0],"amount":r[1],"cat":r[2],"note":r[3],"date":str(r[4])[:10]} for r in rows]
            result["expense_total"] = sum(r[1] for r in rows)
        self._send_json(result)

    # ── export ─────────────────────────────────────────

    def _handle_export(self, uid, qs):
        scope = qs.get("scope", ["finance"])[0]
        fmt   = qs.get("format", ["csv"])[0]
        try:
            if scope == "crypto":
                rows  = db_crypto_all_trades(uid) if uid else []
                heads = ["symbol","cg_id","side","qty","price_usd","fee_usd","note","created_at"]
            else:
                rows  = db_export_data(uid) if uid else []
                heads = ["Loại","Số tiền","Ghi chú","Danh mục/Nguồn","Ngày","ID"]

            fname = f"{scope}_export"
            if fmt == "excel":
                try:
                    from openpyxl import Workbook
                    wb = Workbook(); ws = wb.active
                    ws.title = scope.capitalize()
                    ws.append(heads)
                    for r in rows:
                        ws.append(list(r))
                    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                    data = buf.read()
                    ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    dname = fname + ".xlsx"
                except ImportError:
                    self._send_json({"error": "openpyxl not installed"}, 500); return
            else:
                buf = io.StringIO()
                w = csv.writer(buf)
                w.writerow(heads)
                for r in rows:
                    w.writerow(list(r))
                data = ("﻿" + buf.getvalue()).encode("utf-8")
                ctype = "text/csv; charset=utf-8-sig"
                dname = fname + ".csv"

            self.send_response(200)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Disposition", f'attachment; filename="{dname}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data if isinstance(data, bytes) else data.encode("utf-8"))
        except Exception as exc:
            self._send_json({"error": str(exc)}, 500)

    # ── crypto API ────────────────────────────────────

    def _api_crypto_add(self, uid, body):
        if not uid:
            self._send_json({"error": "no user_id"}); return
        symbol = str(body.get("symbol", "")).strip().upper()
        side   = str(body.get("side", "BUY")).strip().upper()
        qty    = float(body.get("qty", 0))
        price  = float(body.get("price", 0))
        fee    = float(body.get("fee", 0) or 0)
        note   = str(body.get("note", "") or "")
        cg_id  = str(body.get("cg_id", "") or "").strip() or None
        date_s = str(body.get("date", "") or "").strip()
        if not symbol or qty <= 0 or price < 0:
            self._send_json({"error": "symbol/qty/price invalid"}); return
        if not cg_id:
            cg_id = db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
        try:
            created_at = datetime.datetime.fromisoformat(date_s) if date_s else datetime.datetime.now()
        except Exception:
            created_at = datetime.datetime.now()
        db_crypto_add_trade(uid, symbol, side, qty, price, note, created_at, cg_id, fee)
        if cg_id:
            db_crypto_upsert_map(symbol, cg_id)
        self._send_json({"ok": True})

    def _api_crypto_delete(self, body):
        trade_id = int(body.get("id", 0))
        if not trade_id:
            self._send_json({"error": "no id"}); return
        db_crypto_delete_trade(trade_id)
        self._send_json({"ok": True})

    def _api_crypto_update(self, body):
        trade_id = int(body.get("id", 0))
        if not trade_id:
            self._send_json({"error": "no id"}); return
        qty   = float(body.get("qty", 0))
        price = float(body.get("price", 0))
        fee   = float(body.get("fee", 0) or 0)
        note  = str(body.get("note", "") or "")
        db_crypto_update_trade(trade_id, qty, price, fee, note)
        self._send_json({"ok": True})

    # ── finance API ───────────────────────────────────

    def _api_finance_add(self, uid, body):
        if not uid:
            self._send_json({"error": "no user_id"}); return
        ttype    = str(body.get("ttype", "expense"))
        amount   = float(body.get("amount", 0))
        category = str(body.get("category", "") or "")
        note     = str(body.get("note", "") or "")
        date_s   = str(body.get("date", "") or "").strip()
        if amount <= 0:
            self._send_json({"error": "amount must be > 0"}); return
        try:
            created_at = datetime.datetime.fromisoformat(date_s) if date_s else datetime.datetime.now()
        except Exception:
            created_at = datetime.datetime.now()
        if ttype == "income":
            db_add_income(uid, amount, category or "Khác", note, created_at)
        else:
            db_add_expense(uid, amount, note, category or "Khác", created_at)
        self._send_json({"ok": True})

    def _api_finance_delete(self, body):
        rec_id = int(body.get("id", 0))
        ttype  = str(body.get("ttype", "expense"))
        if not rec_id:
            self._send_json({"error": "no id"}); return
        db_delete_transaction(rec_id, ttype)
        self._send_json({"ok": True})

    def _api_finance_update(self, body):
        rec_id   = int(body.get("id", 0))
        ttype    = str(body.get("ttype", "expense"))
        amount   = float(body.get("amount", 0))
        category = str(body.get("category", "") or "")
        note     = str(body.get("note", "") or "")
        date_s   = str(body.get("date", "") or "").strip()
        if not rec_id:
            self._send_json({"error": "no id"}); return
        if not date_s:
            date_s = datetime.datetime.now().isoformat()
        if ttype == "income":
            db_update_income(rec_id, amount, category or "Khác", note, date_s)
        else:
            db_update_expense(rec_id, amount, category or "Khác", note, date_s)
        self._send_json({"ok": True})

    # ── import API ────────────────────────────────────

    def _api_import(self, uid, body):
        import base64
        if not uid:
            self._send_json({"error": "no user_id"}); return
        scope  = str(body.get("scope", "finance"))
        fmt    = str(body.get("format", "csv"))
        data64 = str(body.get("data", ""))
        try:
            raw = base64.b64decode(data64)
        except Exception as exc:
            self._send_json({"error": f"base64 decode failed: {exc}"}); return

        imported = failed = 0
        try:
            if fmt == "excel":
                from openpyxl import load_workbook as _lw
                wb = _lw(filename=io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows_iter = (
                    {headers[i]: v for i, v in enumerate(row) if i < len(headers)}
                    for row in ws.iter_rows(min_row=2, values_only=True)
                )
            else:
                text = raw.decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(text))
                # normalise header keys
                rows_iter = ({k.strip().lower(): v for k, v in r.items()} for r in reader)

            for row in rows_iter:
                if all((v is None or str(v).strip() == "") for v in row.values()):
                    continue
                try:
                    if scope == "crypto":
                        sym  = str(row.get("symbol") or row.get("asset") or "").strip().upper()
                        side = str(row.get("side") or "BUY").strip().upper()
                        qty  = float(row.get("qty") or row.get("quantity") or 0)
                        price= float(row.get("price_usd") or row.get("price") or 0)
                        fee  = float(row.get("fee_usd") or row.get("fee") or 0)
                        note = str(row.get("note") or "")
                        cg   = str(row.get("cg_id") or "").strip() or None
                        dat  = str(row.get("created_at") or row.get("date") or "").strip()
                        if not sym or qty <= 0:
                            failed += 1; continue
                        try:
                            cat = datetime.datetime.fromisoformat(dat) if dat else datetime.datetime.now()
                        except Exception:
                            cat = datetime.datetime.now()
                        cg = cg or db_crypto_get_map(sym) or cg_guess_id_from_symbol(sym)
                        db_crypto_add_trade(uid, sym, side, qty, price, note, cat, cg, fee)
                        if cg:
                            db_crypto_upsert_map(sym, cg)
                    else:  # finance
                        ttype  = str(row.get("type") or row.get("ttype") or row.get("loại") or "expense").strip()
                        amount = float(row.get("amount") or row.get("số tiền") or row.get("số_tiền") or 0)
                        cat_   = str(row.get("cat_or_source") or row.get("category") or row.get("source")
                                     or row.get("danh mục/nguồn") or row.get("danh_mục/nguồn") or "Khác")
                        note   = str(row.get("note") or row.get("ghi chú") or row.get("ghi_chú") or "")
                        dat    = str(row.get("created_at") or row.get("date") or row.get("ngày") or "").strip()
                        if amount <= 0:
                            failed += 1; continue
                        try:
                            cat_dt = datetime.datetime.fromisoformat(dat) if dat else datetime.datetime.now()
                        except Exception:
                            cat_dt = datetime.datetime.now()
                        is_income = (ttype.upper().startswith("INC") or
                                     ttype.lower() in ("income", "thu nhập", "thu nhap"))
                        if is_income:
                            db_add_income(uid, amount, cat_, note, cat_dt)
                        else:
                            db_add_expense(uid, amount, note, cat_, cat_dt)
                    imported += 1
                except Exception:
                    failed += 1
        except Exception as exc:
            self._send_json({"error": str(exc), "imported": imported, "failed": failed}); return

        self._send_json({"ok": True, "imported": imported, "failed": failed})

    def _handle_db_download(self):
        try:
            with open(DB_PATH, "rb") as f:
                data = f.read()
            fname = os.path.basename(DB_PATH)
            self.send_response(200)
            self.send_header("Content-Type", "application/octet-stream")
            self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except Exception as exc:
            self._send_json({"error": str(exc)}, 500)

    def _handle_db_upload(self, body):
        import base64, shutil
        data64 = str(body.get("data", ""))
        try:
            raw = base64.b64decode(data64)
        except Exception as e:
            self._send_json({"error": f"base64: {e}"}); return
        if not raw.startswith(b"SQLite format 3"):
            self._send_json({"error": "File không phải SQLite database hợp lệ"}); return
        bak = DB_PATH + ".bak"
        try:
            if os.path.exists(DB_PATH):
                shutil.copy2(DB_PATH, bak)
            with open(DB_PATH, "wb") as f:
                f.write(raw)
            run_migrations()
            self._send_json({"ok": True, "size": len(raw)})
        except Exception as exc:
            if os.path.exists(bak):
                shutil.copy2(bak, DB_PATH)
            self._send_json({"error": str(exc)}, 500)


# ── SECTION 3: _make_html (full replacement) ───────────

# ===================== AIOHTTP UNIFIED SERVER (webhook mode) =====================
# In webhook mode a single aiohttp server handles BOTH Telegram webhook
# updates and the HTML dashboard, so only one public port is needed.

_ptb_app_ref = None  # set by run_with_webhook()

def _aio_check_auth(qs) -> bool:
    if not DASHBOARD_SECRET: return True
    return qs.get("token", "") == DASHBOARD_SECRET

def _aio_uid(qs):
    try: return int(qs.get("user_id", ""))
    except: return None

def _dash_export_bytes(uid, scope: str, fmt: str):
    """Return (body_bytes, content_type, filename)."""
    if scope == "crypto":
        rows  = db_crypto_all_trades(uid) if uid else []
        heads = ["symbol","cg_id","side","qty","price_usd","fee_usd","note","created_at"]
    else:
        rows  = db_export_data(uid) if uid else []
        heads = ["Loại","Số tiền","Ghi chú","Danh mục/Nguồn","Ngày","ID"]
    fname = f"{scope}_export"
    if fmt == "excel":
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.title = scope.capitalize(); ws.append(heads)
        for r in rows: ws.append(list(r))
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return (buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fname + ".xlsx")
    buf = io.StringIO(); w = csv.writer(buf); w.writerow(heads)
    for r in rows: w.writerow(list(r))
    return (("﻿" + buf.getvalue()).encode("utf-8"), "text/csv; charset=utf-8-sig", fname + ".csv")

def _dash_finance_list_sync(uid, qs) -> dict:
    ttype     = qs.get("ttype",     "all")
    date_from = (qs.get("date_from","") or "").strip() or None
    date_to   = (qs.get("date_to",  "") or "").strip() or None
    category  = (qs.get("category", "") or "").strip() or None
    search    = (qs.get("search",   "") or "").strip() or None
    limit     = int(qs.get("limit", 500) or 500)
    result: dict = {}
    if ttype in ("income", "all"):
        rows = db_get_incomes_filtered(uid, date_from, date_to, category if ttype == "income" else None, search, limit)
        result["income"] = [{"id":r[0],"amount":r[1],"cat":r[2],"note":r[3],"date":str(r[4])[:10]} for r in rows]
        result["income_total"] = sum(r[1] for r in rows)
    if ttype in ("expense", "all"):
        rows = db_get_expenses_filtered(uid, date_from, date_to, category if ttype == "expense" else None, search, limit)
        result["expense"] = [{"id":r[0],"amount":r[1],"cat":r[2],"note":r[3],"date":str(r[4])[:10]} for r in rows]
        result["expense_total"] = sum(r[1] for r in rows)
    return result

def _dash_post_sync(uid, path: str, body: dict) -> dict:
    """Handles all dashboard POST API calls, returns JSON-serialisable dict."""
    import base64
    # ── crypto ──
    if path == "/api/crypto/add":
        if not uid: return {"error": "no user_id"}
        sym = str(body.get("symbol","")).strip().upper()
        side = str(body.get("side","BUY")).strip().upper()
        qty = float(body.get("qty",0)); price = float(body.get("price",0))
        fee = float(body.get("fee",0) or 0); note = str(body.get("note","") or "")
        cg = str(body.get("cg_id","") or "").strip() or None
        ds = str(body.get("date","") or "").strip()
        if not sym or qty <= 0 or price < 0: return {"error":"symbol/qty/price invalid"}
        if not cg: cg = db_crypto_get_map(sym) or cg_guess_id_from_symbol(sym)
        try: dt = datetime.datetime.fromisoformat(ds) if ds else datetime.datetime.now()
        except: dt = datetime.datetime.now()
        db_crypto_add_trade(uid, sym, side, qty, price, note, dt, cg, fee)
        if cg: db_crypto_upsert_map(sym, cg)
        return {"ok": True}
    if path == "/api/crypto/delete":
        tid = int(body.get("id",0))
        if not tid: return {"error":"no id"}
        db_crypto_delete_trade(tid); return {"ok": True}
    if path == "/api/crypto/update":
        tid = int(body.get("id",0))
        if not tid: return {"error":"no id"}
        db_crypto_update_trade(tid, float(body.get("qty",0)), float(body.get("price",0)),
                               float(body.get("fee",0) or 0), str(body.get("note","") or ""))
        return {"ok": True}
    # ── finance ──
    if path == "/api/finance/add":
        if not uid: return {"error":"no user_id"}
        ttype = str(body.get("ttype","expense")); amt = float(body.get("amount",0))
        cat = str(body.get("category","") or ""); note = str(body.get("note","") or "")
        ds = str(body.get("date","") or "").strip()
        if amt <= 0: return {"error":"amount must be > 0"}
        try: dt = datetime.datetime.fromisoformat(ds) if ds else datetime.datetime.now()
        except: dt = datetime.datetime.now()
        if ttype == "income": db_add_income(uid, amt, cat or "Khác", note, dt)
        else: db_add_expense(uid, amt, note, cat or "Khác", dt)
        return {"ok": True}
    if path == "/api/finance/delete":
        rid = int(body.get("id",0)); ttype = str(body.get("ttype","expense"))
        if not rid: return {"error":"no id"}
        db_delete_transaction(rid, ttype); return {"ok": True}
    if path == "/api/finance/update":
        rid = int(body.get("id",0)); ttype = str(body.get("ttype","expense"))
        amt = float(body.get("amount",0)); cat = str(body.get("category","") or "")
        note = str(body.get("note","") or ""); ds = str(body.get("date","") or "").strip()
        if not rid: return {"error":"no id"}
        if not ds: ds = datetime.datetime.now().isoformat()
        if ttype == "income": db_update_income(rid, amt, cat or "Khác", note, ds)
        else: db_update_expense(rid, amt, cat or "Khác", note, ds)
        return {"ok": True}
    # ── import ──
    if path == "/api/import":
        if not uid: return {"error":"no user_id"}
        scope = str(body.get("scope","finance")); fmt = str(body.get("format","csv"))
        try: raw = base64.b64decode(str(body.get("data","")))
        except Exception as e: return {"error": f"base64: {e}"}
        imported = failed = 0
        try:
            if fmt == "excel":
                from openpyxl import load_workbook as _lw
                wb = _lw(filename=io.BytesIO(raw), read_only=True, data_only=True); ws = wb.active
                hdrs = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows_it = ({hdrs[i]:v for i,v in enumerate(row) if i<len(hdrs)} for row in ws.iter_rows(min_row=2, values_only=True))
            else:
                rows_it = ({k.strip().lower():v for k,v in r.items()} for r in csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
            for row in rows_it:
                if all((v is None or str(v).strip()=="") for v in row.values()): continue
                try:
                    if scope == "crypto":
                        sym=str(row.get("symbol") or row.get("asset") or "").strip().upper()
                        side=str(row.get("side") or "BUY").strip().upper()
                        qty=float(row.get("qty") or row.get("quantity") or 0)
                        price=float(row.get("price_usd") or row.get("price") or 0)
                        fee=float(row.get("fee_usd") or row.get("fee") or 0)
                        note=str(row.get("note") or "")
                        cg=str(row.get("cg_id") or "").strip() or None
                        dat=str(row.get("created_at") or row.get("date") or "").strip()
                        if not sym or qty<=0: failed+=1; continue
                        try: cat=datetime.datetime.fromisoformat(dat) if dat else datetime.datetime.now()
                        except: cat=datetime.datetime.now()
                        cg=cg or db_crypto_get_map(sym) or cg_guess_id_from_symbol(sym)
                        db_crypto_add_trade(uid,sym,side,qty,price,note,cat,cg,fee)
                        if cg: db_crypto_upsert_map(sym,cg)
                    else:
                        ttype=str(row.get("type") or row.get("ttype") or row.get("loại") or "expense").strip()
                        amount=float(row.get("amount") or row.get("số tiền") or row.get("số_tiền") or 0)
                        cat_=str(row.get("cat_or_source") or row.get("category") or row.get("source") or row.get("danh mục/nguồn") or row.get("danh_mục/nguồn") or "Khác")
                        note=str(row.get("note") or row.get("ghi chú") or row.get("ghi_chú") or "")
                        dat=str(row.get("created_at") or row.get("date") or row.get("ngày") or "").strip()
                        if amount<=0: failed+=1; continue
                        try: cat_dt=datetime.datetime.fromisoformat(dat) if dat else datetime.datetime.now()
                        except: cat_dt=datetime.datetime.now()
                        is_inc=(ttype.upper().startswith("INC") or ttype.lower() in ("income","thu nhập","thu nhap"))
                        if is_inc: db_add_income(uid,amount,cat_,note,cat_dt)
                        else: db_add_expense(uid,amount,note,cat_,cat_dt)
                    imported+=1
                except: failed+=1
        except Exception as e: return {"error":str(e),"imported":imported,"failed":failed}
        return {"ok":True,"imported":imported,"failed":failed}
    if path == "/api/db/upload":
        import base64, shutil
        try: raw = base64.b64decode(str(body.get("data","")))
        except Exception as e: return {"error": f"base64: {e}"}
        if not raw.startswith(b"SQLite format 3"):
            return {"error": "File không phải SQLite database hợp lệ"}
        bak = DB_PATH + ".bak"
        try:
            if os.path.exists(DB_PATH): shutil.copy2(DB_PATH, bak)
            with open(DB_PATH, "wb") as f: f.write(raw)
            run_migrations()
            return {"ok": True, "size": len(raw)}
        except Exception as exc:
            if os.path.exists(bak): shutil.copy2(bak, DB_PATH)
            return {"error": str(exc)}
    return {"error":"unknown path"}

async def _aio_get(request):
    from aiohttp import web
    qs = request.rel_url.query
    if not _aio_check_auth(qs):
        return web.Response(status=403, text="403 Forbidden")
    uid = _aio_uid(qs)
    path = request.path
    try:
        if path == "/api/export":
            scope = qs.get("scope","finance"); fmt = qs.get("format","csv")
            data, ctype, fname = await asyncio.to_thread(_dash_export_bytes, uid, scope, fmt)
            return web.Response(body=data, content_type=ctype.split(";")[0].strip(),
                                headers={"Content-Disposition": f'attachment; filename="{fname}"',
                                         "Access-Control-Allow-Origin": "*"})
        if path == "/api/finance/list":
            if not uid: return web.Response(text='{"error":"no user_id"}', content_type="application/json")
            result = await asyncio.to_thread(_dash_finance_list_sync, uid, qs)
            return web.Response(text=json.dumps(result, ensure_ascii=False),
                                content_type="application/json", charset="utf-8")
        if path == "/api/finance/categories":
            cats = await asyncio.to_thread(db_get_finance_categories, uid) if uid else {"income":[],"expense":[]}
            return web.Response(text=json.dumps(cats, ensure_ascii=False),
                                content_type="application/json", charset="utf-8")
        if path == "/api/db/download":
            try:
                data = await asyncio.to_thread(lambda: open(DB_PATH,"rb").read())
                fname = os.path.basename(DB_PATH)
                return web.Response(body=data, content_type="application/octet-stream",
                                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})
            except Exception as exc:
                return web.Response(text=json.dumps({"error":str(exc)}), content_type="application/json", status=500)
        html = await asyncio.to_thread(_make_html, uid)
        return web.Response(text=html, content_type="text/html", charset="utf-8")
    except Exception as exc:
        logger.error(f"aio_get error: {exc}")
        return web.Response(text=json.dumps({"error": str(exc)}), content_type="application/json", status=500)

async def _aio_post(request):
    from aiohttp import web
    qs = request.rel_url.query
    if not _aio_check_auth(qs):
        return web.Response(status=403, text="403 Forbidden")
    uid = _aio_uid(qs)
    try:
        body = await request.json()
    except Exception as exc:
        return web.Response(text=json.dumps({"error": f"JSON: {exc}"}), content_type="application/json", status=400)
    try:
        result = await asyncio.to_thread(_dash_post_sync, uid, request.path, body)
    except Exception as exc:
        result = {"error": str(exc)}
    return web.Response(text=json.dumps(result, ensure_ascii=False),
                        content_type="application/json", charset="utf-8",
                        headers={"Access-Control-Allow-Origin": "*"})

async def _aio_telegram(request):
    from aiohttp import web
    if WEBHOOK_SECRET:
        tok = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
        if tok != WEBHOOK_SECRET:
            return web.Response(status=403, text="Forbidden")
    try:
        data = await request.json()
        update = Update.de_json(data, _ptb_app_ref.bot)
        await _ptb_app_ref.update_queue.put(update)
    except Exception as exc:
        logger.error(f"Telegram webhook error: {exc}")
    return web.Response(text="OK")

def _setup_job_queue(ptb_app) -> None:
    jq = ptb_app.job_queue
    if jq is None:
        logger.warning("JobQueue không khả dụng — pip install 'python-telegram-bot[job-queue]'")
        return
    try:
        import pytz; tz = pytz.timezone("Asia/Ho_Chi_Minh")
    except ImportError:
        tz = datetime.timezone(datetime.timedelta(hours=7))
    jq.run_daily(check_recurring_job, time=datetime.time(8, 0, tzinfo=tz), name="recurring_daily")
    jq.run_once(check_recurring_job, when=5, name="recurring_startup")

async def run_with_webhook(ptb_app) -> None:
    global _ptb_app_ref
    _ptb_app_ref = ptb_app
    from aiohttp import web
    _setup_job_queue(ptb_app)

    wh_path = f"/wh/{BOT_TOKEN}"
    aio_app = web.Application()
    aio_app.router.add_post(wh_path,                  _aio_telegram)
    aio_app.router.add_get("/api/export",             _aio_get)
    aio_app.router.add_get("/api/finance/list",       _aio_get)
    aio_app.router.add_get("/api/finance/categories", _aio_get)
    aio_app.router.add_get("/api/db/download",        _aio_get)
    aio_app.router.add_post("/api/{tail:.*}",         _aio_post)
    aio_app.router.add_get("/{tail:.*}",              _aio_get)

    runner = web.AppRunner(aio_app, access_log=None)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", WEBHOOK_PORT)
    await site.start()
    logger.info(f"Unified server on port {WEBHOOK_PORT} (webhook + dashboard)")

    async with ptb_app:
        await ptb_app.bot.set_webhook(
            url=f"{WEBHOOK_URL}{wh_path}",
            secret_token=WEBHOOK_SECRET or None,
            allowed_updates=list(Update.ALL_TYPES),
            drop_pending_updates=False,
        )
        logger.info(f"Webhook OK: {WEBHOOK_URL}{wh_path}")
        logger.info(f"Dashboard:  {WEBHOOK_URL}?user_id=YOUR_ID")
        await ptb_app.start()

        stop_event = asyncio.Event()
        import signal as _signal
        loop = asyncio.get_running_loop()
        for sig in (_signal.SIGINT, _signal.SIGTERM):
            try: loop.add_signal_handler(sig, stop_event.set)
            except (NotImplementedError, RuntimeError): pass

        await stop_event.wait()
        await ptb_app.stop()

    await runner.cleanup()

def start_html_server():
    global HTML_PORT
    for port in range(HTML_PORT, HTML_PORT + 10):
        try:
            server = HTTPServer(("0.0.0.0", port), _DashboardHandler)
            if port != HTML_PORT:
                logger.warning(f"Port {HTML_PORT} đã bị chiếm, dùng port {port}")
                HTML_PORT = port
            t = threading.Thread(target=server.serve_forever, daemon=True)
            t.start()
            logger.info(f"HTML dashboard running on port {HTML_PORT}")
            return server
        except OSError:
            continue
    logger.error("Không tìm được port trống (8080–8089). Dashboard không khởi động.")
    return None

# ===================== KEYBOARDS =====================
# ─── Home ───────────────────────────────────────────────────────────────
BTN_HOME_CRYPTO  = "📈 Crypto"
BTN_HOME_FINANCE = "💰 Thu Chi"
BTN_BACK         = "🔙 Trang Chủ"

# ─── Crypto submenu ─────────────────────────────────────────────────────
BTN_PORTFOLIO = "📊 Danh Mục"
BTN_BUY       = "➕ Mua"
BTN_SELL      = "➖ Bán"
BTN_MAP       = "🔗 Map Token"
BTN_CIMPORT   = "⬇️ Nhập"
BTN_CEXPORT   = "⬆️ Xuất"

# ─── Finance submenu ────────────────────────────────────────────────────
BTN_FADD    = "💵 Thêm"
BTN_FDEL      = "🗑️ Xóa"
BTN_BUDGET    = "🎯 Ngân Sách"
BTN_RECURRING = "🔁 Định Kỳ"
BTN_FIMPORT   = "⬇️ Nhập File"
BTN_FEXPORT   = "⬆️ Xuất File"

# ─── Shared (route by user_data['submenu']) ──────────────────────────────
BTN_CHART  = "📈 Biểu Đồ"
BTN_REPORT = "📋 Báo Cáo"

ALL_MENU_BTNS = {
    BTN_HOME_CRYPTO, BTN_HOME_FINANCE, BTN_BACK,
    BTN_PORTFOLIO, BTN_BUY, BTN_SELL, BTN_MAP, BTN_CIMPORT, BTN_CEXPORT,
    BTN_FADD, BTN_FDEL, BTN_BUDGET, BTN_FIMPORT, BTN_FEXPORT,
    BTN_CHART, BTN_REPORT,
}

def home_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [KeyboardButton(BTN_HOME_CRYPTO)],
        [KeyboardButton(BTN_HOME_FINANCE)],
    ], resize_keyboard=True)

def main_menu_keyboard() -> ReplyKeyboardMarkup:
    return home_keyboard()

def crypto_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [KeyboardButton(BTN_PORTFOLIO), KeyboardButton(BTN_BUY),     KeyboardButton(BTN_SELL)],
        [KeyboardButton(BTN_CHART),     KeyboardButton(BTN_REPORT),  KeyboardButton(BTN_MAP)],
        [KeyboardButton(BTN_CIMPORT),   KeyboardButton(BTN_CEXPORT), KeyboardButton(BTN_FDEL)],
        [KeyboardButton(BTN_BACK)],
    ], resize_keyboard=True)

def finance_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [KeyboardButton(BTN_FADD),    KeyboardButton(BTN_FDEL),    KeyboardButton(BTN_BUDGET)],
        [KeyboardButton(BTN_RECURRING), KeyboardButton(BTN_REPORT), KeyboardButton(BTN_CHART)],
        [KeyboardButton(BTN_FIMPORT), KeyboardButton(BTN_FEXPORT), KeyboardButton(BTN_BACK)],
    ], resize_keyboard=True)

def _submenu_keyboard(user_data: dict) -> ReplyKeyboardMarkup:
    submenu = user_data.get('submenu')
    if submenu == 'finance': return finance_menu_keyboard()
    if submenu == 'crypto':  return crypto_menu_keyboard()
    return home_keyboard()

def _clear_keep_submenu(user_data: dict) -> None:
    submenu = user_data.get('submenu')
    user_data.clear()
    if submenu: user_data['submenu'] = submenu


# ===================== CRYPTO HELPERS =====================
def _build_positions_and_prices(user_id: int):
    positions = db_crypto_positions(user_id)
    ids = [p["cg_id"] or cg_guess_id_from_symbol(p["symbol"]) for p in positions if (p["cg_id"] or cg_guess_id_from_symbol(p["symbol"]))]
    return positions, cg_simple_price_usd(ids)

def _parse_price(s: str) -> Optional[float]:
    s = s.lower().replace(',','').strip()
    try:
        if s.endswith('k'): return float(s[:-1])*1_000
        if s.endswith('m'): return float(s[:-1])*1_000_000
        if s.endswith('b'): return float(s[:-1])*1_000_000_000
        return float(s)
    except Exception: return None

def _parse_natural_trade(text: str, user_id: int = 0):
    text = text.strip()
    side = None
    for kw in ['bán','ban','sell']:
        if text.lower().startswith(kw): side='SELL'; text=text[len(kw):].strip(); break
    if side is None:
        for kw in ['mua','buy']:
            if text.lower().startswith(kw): side='BUY'; text=text[len(kw):].strip(); break
    if side is None: return None
    parts = text.split()
    if not parts: return None
    symbol = parts[0].upper(); text = ' '.join(parts[1:]).strip()
    text = re.sub(r'\b(gia|giá|price|@)\b',' ',text,flags=re.IGNORECASE).strip()
    text = re.sub(r'\s+',' ',text)
    parts = text.split()
    if len(parts) < 2: return None
    qty_raw = parts[0].lower(); price_raw = parts[1]
    note = ' '.join(parts[2:]).strip() if len(parts)>2 else ''
    qty = None
    if qty_raw in ('tất cả','tat ca','hết','het','all','toàn bộ','toan bo'):
        if side=='SELL' and user_id: qty = _get_held_qty(user_id,symbol)
        else: return None
    elif re.fullmatch(r'[\d.,]+\s*%',qty_raw) or re.fullmatch(r'[\d.,]+pct',qty_raw):
        pct_val = float(re.sub(r'[%pct]','',qty_raw).replace(',',''))
        if side=='SELL' and user_id: qty = _get_held_qty(user_id,symbol)*pct_val/100
        else: return None
    else: qty = _parse_price(qty_raw)
    if qty is None or qty<=0: return None
    price = _parse_price(price_raw)
    if price is None or price<=0: return None
    return side, symbol, qty, price, note

async def _execute_trade(update: Update, context: ContextTypes.DEFAULT_TYPE,
                         side: str, symbol: str, qty: float, price: float, note: str):
    user_id = update.effective_user.id
    if side=='SELL':
        held = _get_held_qty(user_id,symbol)
        if qty > held+1e-9:
            await update.message.reply_text(f"⚠️ Bạn chỉ đang giữ {held:g} {symbol}, không thể bán {qty:g}.")
            return
    cg_id = db_crypto_get_map(symbol) or cg_guess_id_from_symbol(symbol)
    if cg_id: db_crypto_upsert_map(symbol,cg_id)
    db_crypto_add_trade(user_id,symbol,side,qty,price,note=note,cg_id=cg_id)
    action="MUA" if side=='BUY' else "BAN"
    msg=f"✅ {action} {symbol}: {qty:g} @ ${price:,.4g} = ${qty*price:,.2f} USD"
    if note: msg+=f"\nGhi chú: {note}"
    await update.message.reply_text(msg)

async def send_portfolio_images(chat_id: int, user_id: int, context: ContextTypes.DEFAULT_TYPE):
    positions, prices = _build_positions_and_prices(user_id)
    if not positions:
        await context.bot.send_message(chat_id,"Danh mục trống. Dùng ➕ Mua để thêm giao dịch.")
        return
    theme = db_get_theme(user_id)
    donut = table = None
    try:
        donut = await asyncio.to_thread(gen_portfolio_donut_image,positions,prices,theme)
        table = await asyncio.to_thread(gen_portfolio_table_image,positions,prices,theme)
        with open(donut,"rb") as f1: await context.bot.send_photo(chat_id,f1)
        with open(table,"rb") as f2: await context.bot.send_photo(chat_id,f2,caption="📋 Danh mục")
        if user_id==OWNER_USER_ID:
            try:
                with open(donut,"rb") as fc:
                    await context.bot.send_photo(CHANNEL_CHAT_ID,fc,
                        caption=f"📊 Phân bổ danh mục — {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
            except Exception as ce:
                await context.bot.send_message(chat_id,f"⚠️ Không gửi được lên channel: {ce!r}")
    except Exception as e:
        await context.bot.send_message(chat_id,f"⚠️ Lỗi khi tạo ảnh: {e!r}")
    finally:
        for fp in (donut,table):
            try:
                if fp and os.path.exists(fp): os.remove(fp)
            except Exception: pass

# ===================== CRYPTO IMPORT/EXPORT =====================
def _normalize_datetime_str(val) -> Optional[str]:
    if val is None: return None
    try:
        if hasattr(val,"year") and hasattr(val,"month"): return val.strftime("%Y-%m-%d %H:%M:%S")
    except Exception: pass
    s = str(val).strip()
    if not s: return None
    for f in ["%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M","%Y-%m-%d","%d/%m/%Y %H:%M","%d/%m/%Y"]:
        try: return datetime.datetime.strptime(s,f).strftime("%Y-%m-%d %H:%M:%S")
        except Exception: continue
    try: return datetime.datetime.fromisoformat(s).strftime("%Y-%m-%d %H:%M:%S")
    except Exception: pass
    raise ValueError(f"Invalid datetime: {s}")

def _parse_trades_from_excel(bio) -> list:
    if load_workbook is None:
        raise RuntimeError("Excel import requires openpyxl: pip install openpyxl")
    bio.seek(0)
    wb = load_workbook(filename=bio,read_only=True,data_only=True)
    ws = wb["Trades"] if "Trades" in wb.sheetnames else wb.active
    header_cells = next(ws.iter_rows(min_row=1,max_row=1,values_only=False))
    headers = [c.value.strip() if isinstance(c.value,str) else (c.value or "") for c in header_cells]
    norm = {str(h).strip().lower():i for i,h in enumerate(headers) if str(h).strip()}
    def gv(rv,*names):
        for n in names:
            k=n.lower()
            if k in norm and norm[k]<len(rv): return rv[norm[k]]
        return None
    rows = []
    for r in ws.iter_rows(min_row=2,values_only=True):
        if r is None or all((v is None or (isinstance(v,str) and not v.strip())) for v in r): continue
        rows.append({"date":gv(r,"date","created_at"),"exchange":gv(r,"exchange"),"symbol":gv(r,"symbol","asset","token"),
                     "pair":gv(r,"pair"),"side":gv(r,"side","action"),"quantity":gv(r,"quantity","qty","amount"),
                     "price":gv(r,"price","price_usd"),"fee":gv(r,"fee","fee_usd"),
                     "fee_currency":gv(r,"feecurrency","fee_currency"),"note":gv(r,"note","memo","comment"),
                     "created_at":gv(r,"created_at")})
    if not rows: raise ValueError(f"No data rows. Headers: {list(norm.keys())}")
    return rows

async def cp_import_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_rate(update): return
    doc = None
    if update.message and update.message.document: doc = update.message.document
    if not doc and update.message and update.message.reply_to_message and update.message.reply_to_message.document:
        doc = update.message.reply_to_message.document
    if not doc:
        await update.message.reply_text("Gửi file CSV hoặc Excel rồi reply `/cp_import`.",parse_mode="Markdown"); return
    await update.message.reply_text("📥 Đang xử lý file...")
    fobj = await context.bot.get_file(doc.file_id)
    bio = io.BytesIO(); await fobj.download_to_memory(out=bio); bio.seek(0)
    filename=(doc.file_name or "").lower(); mime=(doc.mime_type or "").lower()
    ok=fail=0; user=update.effective_user
    try:
        if filename.endswith((".xlsx",".xlsm")) or "spreadsheetml" in mime:
            for row in _parse_trades_from_excel(bio):
                try:
                    sym=(row.get("symbol") or "").strip().upper()
                    side=(row.get("side") or "").strip().upper()
                    qty=float(row.get("quantity") or 0); price=float(row.get("price") or 0)
                    fee=float(row.get("fee") or 0) if str(row.get("fee") or "").strip() else 0.0
                    note=(row.get("note") or "").strip()
                    at_raw=row.get("created_at"); created_at=None
                    if at_raw not in (None,""):
                        try: created_at=datetime.datetime.fromisoformat(_normalize_datetime_str(at_raw))
                        except Exception: pass
                    cg_id=db_crypto_get_map(sym) or cg_guess_id_from_symbol(sym)
                    if cg_id: db_crypto_upsert_map(sym,cg_id)
                    if not sym or side not in ("BUY","SELL") or qty<=0 or price<=0: fail+=1; continue
                    db_crypto_add_trade(user.id,sym,side,qty,price,note=note,cg_id=cg_id,fee_usd=fee,created_at=created_at)
                    ok+=1
                except Exception: fail+=1
        else:
            for row in csv.DictReader(io.TextIOWrapper(bio,encoding="utf-8")):
                try:
                    sym=(row.get("symbol") or row.get("Symbol") or "").strip().upper()
                    side=(row.get("side") or row.get("Side") or "").strip().upper()
                    qty=float(row.get("qty") or row.get("quantity") or row.get("Quantity") or 0)
                    price=float(row.get("price") or row.get("price_usd") or row.get("Price") or 0)
                    fee=float(row.get("fee") or row.get("Fee") or 0) if str(row.get("fee") or row.get("Fee") or "").strip() else 0.0
                    note=(row.get("note") or row.get("Note") or "").strip()
                    cg_id=(row.get("cg_id") or row.get("CG_ID") or "").strip() or None
                    at_raw=(row.get("created_at") or row.get("CreatedAt") or "").strip()
                    if not sym or side not in ("BUY","SELL") or qty<=0 or price<=0: fail+=1; continue
                    created_at=None
                    if at_raw:
                        try: created_at=datetime.datetime.fromisoformat(at_raw)
                        except Exception: pass
                    if not cg_id: cg_id=db_crypto_get_map(sym) or cg_guess_id_from_symbol(sym)
                    if cg_id: db_crypto_upsert_map(sym,cg_id)
                    db_crypto_add_trade(user.id,sym,side,qty,price,note=note,cg_id=cg_id,fee_usd=fee,created_at=created_at)
                    ok+=1
                except Exception: fail+=1
    except RuntimeError as e:
        await update.message.reply_text(f"❌ {e}"); return
    await update.message.reply_text(f"✅ Đã nhập {ok} giao dịch, lỗi {fail}.")

# ===================== CRYPTO COMMAND HANDLERS =====================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_migrated()
    uid = update.effective_user.id
    await update.message.reply_text(
        "👋 Chào mừng! Bot quản lý *Crypto* + *Thu Chi* tài chính.\n\n"
        f"🪪 Telegram ID của bạn: `{uid}`\n"
        f"🌐 HTML Dashboard: `http://localhost:{HTML_PORT}?user_id={uid}`\n\n"
        "Dùng bàn phím bên dưới hoặc /help để xem hướng dẫn.",
        parse_mode="Markdown", reply_markup=main_menu_keyboard())

async def dashboard_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    token_part = f"&token={DASHBOARD_SECRET}" if DASHBOARD_SECRET else ""
    if WEBHOOK_URL:
        link = f"{WEBHOOK_URL}?user_id={uid}{token_part}"
        source = "☁️ Server (Railway/Cloud)"
    else:
        link = f"http://localhost:{HTML_PORT}?user_id={uid}{token_part}"
        source = "🖥️ Local (máy tính)"
    secret_note = (f"\n🔑 Token: `{DASHBOARD_SECRET}`" if DASHBOARD_SECRET else
                   "\n⚠️ Chưa đặt DASHBOARD\\_SECRET — nên thêm vào .env")
    await update.message.reply_text(
        f"🌐 *HTML Dashboard*\n\n"
        f"🪪 Telegram ID: `{uid}`\n"
        f"📡 Chế độ: {source}\n"
        f"🔗 Link: `{link}`{secret_note}\n\n"
        "Tính năng:\n"
        "• 📊 Xem portfolio crypto & thu chi\n"
        "• 🔍 Lọc theo ngày, danh mục, từ khoá\n"
        "• ✏️ Thêm / sửa / xóa giao dịch\n"
        "• 📥 Import / 📤 Export CSV & Excel\n"
        "• 💾 Backup & restore database",
        parse_mode="Markdown")

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = (
        "📖 *HƯỚNG DẪN NHANH*\n\n"
        "*📈 CRYPTO*\n"
        "• 📊 Danh Mục — portfolio + lợi nhuận\n"
        "• ➕ Mua / ➖ Bán — nhập lệnh thủ công\n"
        "• 📈 Biểu Đồ — giá lịch sử & cơ cấu\n"
        "• 📋 Báo Cáo — tóm tắt theo kỳ\n"
        "• 🔗 Map Token — gắn CoinGecko ID\n"
        "• ⬇️ Nhập / ⬆️ Xuất — CSV & Excel\n"
        "• Lệnh nhanh: `mua BTC 0.01 giá 70k`\n"
        "• `/cp_add SYMBOL SL GIA [note]`\n"
        "• `/cp_sell SYMBOL SL GIA [note]`\n\n"
        "*💰 THU CHI*\n"
        "• 💵 Thêm — thu nhập hoặc chi tiêu\n"
        "• 🗑️ Xóa — xóa giao dịch\n"
        "• 🎯 Ngân Sách — đặt hạn mức & cảnh báo\n"
        "• 🔁 Định Kỳ — thu/chi tự động hàng tháng\n"
        "• 📋 Báo Cáo / 📈 Biểu Đồ\n"
        "• ⬇️ Nhập File / ⬆️ Xuất File\n"
        "• Lệnh nhanh: `20k ăn sáng`, `+500k lương`\n\n"
        "*🌐 HTML Dashboard*\n"
        "• Lọc thu chi theo ngày, danh mục, từ khoá\n"
        "• Thêm / sửa / xóa không cần reload\n"
        "• Import / Export CSV & Excel\n"
        "• 💾 Backup & restore database\n"
        f"• Dùng /dashboard để lấy link\n\n"
        f"🪪 ID của bạn: `{uid}`"
    )
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=_submenu_keyboard(context.user_data))

async def cp_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_rate(update): return
    await send_portfolio_images(update.effective_chat.id,update.effective_user.id,context)

async def cp_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_rate(update): return
    ensure_migrated()
    args = context.args
    if len(args)<3:
        await update.message.reply_text("Cú pháp: /cp_add SYMBOL SL GIA [note]\nVD: /cp_add BTC 0.01 70000"); return
    symbol=args[0].upper()
    try: qty=float(args[1]); price=float(args[2])
    except Exception: await update.message.reply_text("Số lượng & giá không hợp lệ."); return
    note=" ".join(args[3:]).strip() if len(args)>3 else ""
    if qty<=0 or price<=0: await update.message.reply_text("Số lượng & giá phải > 0"); return
    await _execute_trade(update,context,'BUY',symbol,qty,price,note)

async def cp_sell(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_rate(update): return
    ensure_migrated()
    args = context.args
    if len(args)<3:
        await update.message.reply_text("Cú pháp: /cp_sell SYMBOL SL GIA [note]\nVD: /cp_sell SOL 10 87"); return
    symbol=args[0].upper()
    try: qty=float(args[1]); price=float(args[2])
    except Exception: await update.message.reply_text("Số lượng & giá không hợp lệ."); return
    note=" ".join(args[3:]).strip() if len(args)>3 else ""
    if qty<=0 or price<=0: await update.message.reply_text("Số lượng & giá phải > 0"); return
    await _execute_trade(update,context,'SELL',symbol,qty,price,note)

async def cp_map_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_rate(update): return
    args = context.args
    if len(args)<2: await update.message.reply_text("Cú pháp: /cp_map SYMBOL CG_ID"); return
    symbol,cg_id = args[0].upper(),args[1]
    db_crypto_upsert_map(symbol,cg_id)
    await update.message.reply_text(f"✅ Đã map {symbol} → {cg_id}")

async def theme_select_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    theme_key = q.data.split(":")[1]
    if theme_key not in THEMES: await q.edit_message_text("❌ Theme không hợp lệ."); return
    db_set_theme(q.from_user.id,theme_key)
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(
        f"{'✅ ' if k==theme_key else ''}{v['name']}",callback_data=f"SET_THEME:{k}")
        for k,v in THEMES.items()]])
    await q.edit_message_text(f"✅ Theme: {THEMES[theme_key]['name']}. Bấm 📋 Danh mục để thấy kết quả.",reply_markup=kb)

async def handle_crypto_menu_btns(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_rate(update): return
    text=(update.message.text or "").strip(); user_id=update.effective_user.id
    if text==BTN_PORTFOLIO:
        try: await send_portfolio_images(update.effective_chat.id,user_id,context)
        except Exception as e: await update.message.reply_text(f"Chưa có dữ liệu.\nLỗi: {e}")
    elif text==BTN_BUY:
        await update.message.reply_text("➕ *MUA*: `/cp_add SYMBOL SL GIA [note]`\nVD: `/cp_add BTC 0.01 70000`\nHoặc gõ: `mua BTC 0.01 giá 70k`",parse_mode="Markdown")
    elif text==BTN_SELL:
        await update.message.reply_text("➖ *BÁN*: `/cp_sell SYMBOL SL GIA [note]`\nVD: `/cp_sell SOL 10 87`\nHoặc gõ: `bán SOL tất cả giá 100`",parse_mode="Markdown")
    elif text==BTN_MAP:
        await update.message.reply_text("🔗 *MAP TOKEN*: `/cp_map SYMBOL CG_ID`\nVD: `/cp_map ATOM cosmos`",parse_mode="Markdown")
    elif text==BTN_CIMPORT:
        tmp_fd,tmp=tempfile.mkstemp(prefix="crypto_tpl_",suffix=".csv"); os.close(tmp_fd)
        with open(tmp,"w",newline="",encoding="utf-8") as tf:
            tw=csv.writer(tf)
            tw.writerow(["symbol","cg_id","side","qty","price_usd","fee_usd","note","created_at"])
            tw.writerow(["ATOM","cosmos","BUY","1","10.0","0","","2025-01-01 12:00:00"])
            tw.writerow(["BTC","bitcoin","BUY","0.01","50000","2.5","spot","2025-02-01 20:30:00"])
        try:
            with open(tmp,"rb") as f:
                await update.message.reply_document(f,filename="crypto_import_template.csv",caption="📄 File mẫu. Điền rồi gửi lại + reply `/cp_import`.")
        finally:
            try: os.remove(tmp)
            except Exception: pass
    elif text==BTN_CEXPORT:
        rows=db_crypto_all_trades(user_id)
        if not rows: await update.message.reply_text("Chưa có giao dịch để xuất."); return
        fd,path=tempfile.mkstemp(prefix="portfolio_",suffix=".csv"); os.close(fd)
        with open(path,"w",newline="",encoding="utf-8") as f:
            w=csv.writer(f); w.writerow(["symbol","cg_id","side","qty","price_usd","fee_usd","note","created_at"])
            for r in rows: w.writerow(r)
        try:
            with open(path,"rb") as f: await update.message.reply_document(f,filename="crypto_portfolio.csv",caption="✅ Đã xuất.")
        finally:
            try: os.remove(path)
            except Exception: pass

# ===================== 2-LEVEL MENU NAVIGATION =====================
async def handle_home_crypto_btn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['submenu'] = 'crypto'
    await update.message.reply_text("📈 *Crypto*:", parse_mode="Markdown", reply_markup=crypto_menu_keyboard())

async def handle_home_finance_btn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['submenu'] = 'finance'
    await update.message.reply_text("💰 *Thu Chi*:", parse_mode="Markdown", reply_markup=finance_menu_keyboard())

async def handle_back_btn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop('submenu', None)
    await update.message.reply_text("Menu chính:", reply_markup=home_keyboard())

async def handle_budget_btn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🎯 Đặt ngân sách", callback_data="budget_menu_set")],
        [InlineKeyboardButton("📋 Xem ngân sách",  callback_data="budget_menu_view")],
        [InlineKeyboardButton("❌ Hủy",            callback_data="cancel_action")],
    ])
    await update.message.reply_text("🎯 *Ngân Sách*:", parse_mode="Markdown", reply_markup=kb)

async def handle_budget_menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    if q.data == "budget_menu_set":
        await q.edit_message_text(
            "📝 Dùng lệnh:\n`/ngansach <Danh mục> <Số tiền>`\nVD: `/ngansach Ăn uống 3000000`",
            parse_mode="Markdown")
    elif q.data == "budget_menu_view":
        budgets = db_get_budgets(q.from_user.id)
        if not budgets:
            await q.edit_message_text("Chưa có ngân sách nào. Dùng /ngansach để đặt.")
            return
        lines = ["🎯 *Ngân Sách Hiện Tại:*"]
        for cat, amt in budgets.items():
            lines.append(f"• {cat}: `{amt:,.0f} đ`")
        await q.edit_message_text("\n".join(lines), parse_mode="Markdown")

async def handle_shared_btn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Routes 📈 Biểu Đồ and 📋 Báo Cáo based on user_data['submenu']."""
    submenu = context.user_data.get('submenu', 'finance')
    text = (update.message.text or "").strip()
    if submenu == 'crypto':
        if text == BTN_CHART:
            try: await send_portfolio_images(update.effective_chat.id, update.effective_user.id, context)
            except Exception as e: await update.message.reply_text(f"Chưa có dữ liệu.\nLỗi: {e}")
        elif text == BTN_REPORT:
            await cp_cmd(update, context)
    else:
        if text == BTN_CHART:
            await chart_start(update, context)
        elif text == BTN_REPORT:
            await report_start(update, context)

# ===================== FINANCE CONVERSATION STATES =====================
CHOOSING_INCOME_AMOUNT, CHOOSING_INCOME_NOTE, CHOOSING_INCOME_DATE, CHOOSING_INCOME_SOURCE, CONFIRM_INCOME_AMOUNT = range(5)
CHOOSING_EXPENSE_AMOUNT, CHOOSING_EXPENSE_NOTE, CHOOSING_EXPENSE_DATE, CHOOSING_EXPENSE_CATEGORY, CONFIRM_EXPENSE_AMOUNT = range(5,10)
CHOOSING_ADD_TYPE = 10
CHOOSING_REPORT_PERIOD = 11
CHOOSING_SUPPLEMENT_TYPE = 12
DELETE_CHOOSING_ACTION = 13
DELETE_INDIVIDUAL_LISTING = 14
DELETE_INDIVIDUAL_CONFIRM = 15
RESET_DATA_CONFIRM_STEP1 = 16
RESET_DATA_CONFIRM_STEP2 = 17
CHOOSING_CHART_PERIOD = 18
EXPORT_CHOOSING_FORMAT = 19
CHOOSING_IMPORT_ACTION = 20
IMPORT_FILE_UPLOADED = 21
RECURRING_LIST = 22
RECURRING_ADD_TYPE = 23
RECURRING_ADD_AMOUNT = 24
RECURRING_ADD_CATEGORY = 25
RECURRING_ADD_NOTE = 26
RECURRING_ADD_DAY = 27
RECURRING_ADD_FREQ = 28
RECURRING_ADD_MONTH = 29
RECURRING_EDIT_FIELD = 30
RECURRING_EDIT_VALUE = 31
RECURRING_DELETE_CONFIRM = 32

# ===================== FINANCE HELPERS =====================
def anonymize_name(name: str) -> str:
    if not name: return "****"
    if len(name)<2: return "*"
    return '*'*(len(name)-1)+name[-1]

async def post_to_channel(context: ContextTypes.DEFAULT_TYPE, message: str):
    try:
        await context.bot.send_message(chat_id=FINANCE_CHANNEL_ID,text=message,parse_mode='Markdown')
    except Exception as e:
        logger.error(f"Không đăng được lên kênh: {e}")

async def send_or_update_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, message_text: str = "Menu chính:"):
    kbd = _submenu_keyboard(context.user_data)
    if update.callback_query:
        try: await update.callback_query.edit_message_text(message_text,reply_markup=kbd)
        except Exception: await update.callback_query.message.reply_text(message_text,reply_markup=kbd)
    elif update.message:
        await update.message.reply_text(message_text,reply_markup=kbd)

async def end_conv(update: Update, context: ContextTypes.DEFAULT_TYPE, msg: str) -> int:
    _clear_keep_submenu(context.user_data)
    await send_or_update_main_menu(update,context,msg)
    return ConversationHandler.END

async def cancel_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await end_conv(update,context,'Đã hủy thao tác.')

async def cancel_conversation_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await end_conv(update,context,'Đã hủy thao tác.')

# ===================== ADD FINANCE FLOW =====================
async def handle_add_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _clear_keep_submenu(context.user_data)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Thu nhập (Hôm nay)",callback_data="add_type_income_today")],
        [InlineKeyboardButton("Chi tiêu (Hôm nay)",callback_data="add_type_expense_today")],
        [InlineKeyboardButton("Bổ sung giao dịch",callback_data="add_type_supplement")],
        [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]
    ])
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text("Bạn muốn thêm loại giao dịch nào?",reply_markup=kb)
    else:
        await update.message.reply_text("Bạn muốn thêm loại giao dịch nào?",reply_markup=kb)
    return CHOOSING_ADD_TYPE

async def choose_add_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer(); choice=q.data
    if choice=="add_type_income_today":
        context.user_data['transaction_type']="income"; context.user_data['created_at']=datetime.datetime.now()
        await q.edit_message_text("Nhập số tiền thu nhập (VD: 5tr, 200k):"); return CHOOSING_INCOME_AMOUNT
    elif choice=="add_type_expense_today":
        context.user_data['transaction_type']="expense"; context.user_data['created_at']=datetime.datetime.now()
        await q.edit_message_text("Nhập số tiền chi tiêu (VD: 5tr2, 100k):"); return CHOOSING_EXPENSE_AMOUNT
    elif choice=="add_type_supplement":
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("Thu nhập",callback_data="supplement_type_income")],
            [InlineKeyboardButton("Chi tiêu",callback_data="supplement_type_expense")],
            [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
        await q.edit_message_text("Bổ sung Thu nhập hay Chi tiêu?",reply_markup=kb); return CHOOSING_SUPPLEMENT_TYPE
    return await cancel_conversation_callback(update,context)

async def choose_supplement_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer(); choice=q.data
    context.user_data.pop('created_at',None)
    if choice=="supplement_type_income":
        context.user_data['transaction_type']="income"
        await q.edit_message_text("Nhập số tiền thu nhập:"); return CHOOSING_INCOME_AMOUNT
    elif choice=="supplement_type_expense":
        context.user_data['transaction_type']="expense"
        await q.edit_message_text("Nhập số tiền chi tiêu:"); return CHOOSING_EXPENSE_AMOUNT
    return await cancel_conversation_callback(update,context)

def _income_source_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Lương",callback_data="income_source_Lương"),InlineKeyboardButton("Thưởng",callback_data="income_source_Thưởng")],
        [InlineKeyboardButton("Kinh doanh",callback_data="income_source_Kinh doanh"),InlineKeyboardButton("Được cho",callback_data="income_source_Được cho")],
        [InlineKeyboardButton("Thu hồi nợ",callback_data="income_source_Thu hồi nợ"),InlineKeyboardButton("Bán tài sản",callback_data="income_source_Bán tài sản")],
        [InlineKeyboardButton("Lãi tiết kiệm",callback_data="income_source_Lãi tiết kiệm"),InlineKeyboardButton("Lãi đầu tư",callback_data="income_source_Lãi đầu tư")],
        [InlineKeyboardButton("Khác",callback_data="income_source_Khác")],
        [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]
    ])

def _expense_cat_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Ăn uống",callback_data="expense_category_Ăn uống"),InlineKeyboardButton("Di chuyển",callback_data="expense_category_Di chuyển")],
        [InlineKeyboardButton("Mua sắm",callback_data="expense_category_Mua sắm"),InlineKeyboardButton("Hóa đơn",callback_data="expense_category_Hóa đơn")],
        [InlineKeyboardButton("Gia đình",callback_data="expense_category_Gia đình"),InlineKeyboardButton("Sức khỏe",callback_data="expense_category_Sức khỏe")],
        [InlineKeyboardButton("Giải trí",callback_data="expense_category_Giải trí"),InlineKeyboardButton("Đầu tư",callback_data="expense_category_Đầu tư")],
        [InlineKeyboardButton("Tiết kiệm",callback_data="expense_category_Tiết kiệm"),InlineKeyboardButton("Khác",callback_data="expense_category_Khác")],
        [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]
    ])

async def get_income_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        amount=parse_amount(update.message.text)
        if amount<=0: await update.message.reply_text("Số tiền phải > 0:"); return CHOOSING_INCOME_AMOUNT
        if amount>LARGE_AMOUNT_THRESHOLD:
            context.user_data['amount_to_confirm']=amount
            kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Đúng vậy",callback_data="confirm_amount_yes")],
                [InlineKeyboardButton("❌ Nhập lại",callback_data="confirm_amount_no")],
                [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
            await update.message.reply_text(f"Chắc chắn thu {amount:,.0f} đ? (số lớn)",reply_markup=kb)
            return CONFIRM_INCOME_AMOUNT
        context.user_data['amount']=amount
        await update.message.reply_text("Nhập ghi chú cho khoản thu:"); return CHOOSING_INCOME_NOTE
    except Exception: await update.message.reply_text("Không hợp lệ. Thử lại (VD: 5tr, 200k):"); return CHOOSING_INCOME_AMOUNT

async def confirm_income_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="confirm_amount_yes":
        context.user_data['amount']=context.user_data.pop('amount_to_confirm')
        await q.edit_message_text("Đã xác nhận. Nhập ghi chú:"); return CHOOSING_INCOME_NOTE
    elif q.data=="confirm_amount_no":
        context.user_data.pop('amount_to_confirm',None)
        await q.edit_message_text("Nhập lại số tiền:"); return CHOOSING_INCOME_AMOUNT
    return await cancel_conversation_callback(update,context)

async def get_income_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['note']=update.message.text
    if 'created_at' in context.user_data:
        await update.message.reply_text("Chọn nguồn thu nhập:",reply_markup=_income_source_kb()); return CHOOSING_INCOME_SOURCE
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("Hôm nay",callback_data="date_today_income")],
        [InlineKeyboardButton("Ngày khác",callback_data="date_other_income")],
        [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
    await update.message.reply_text("Giao dịch vào ngày nào?",reply_markup=kb); return CHOOSING_INCOME_DATE

async def choose_income_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="date_today_income":
        context.user_data['created_at']=datetime.datetime.now()
        await q.edit_message_text("Chọn nguồn thu nhập:",reply_markup=_income_source_kb()); return CHOOSING_INCOME_SOURCE
    elif q.data=="date_other_income":
        await q.edit_message_text("Nhập ngày (VD: 25/06/2024 hoặc 15/1):"); return CHOOSING_INCOME_DATE
    return await cancel_conversation_callback(update,context)

async def get_income_date_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['created_at']=parse_date(update.message.text)
        await update.message.reply_text("Chọn nguồn thu nhập:",reply_markup=_income_source_kb()); return CHOOSING_INCOME_SOURCE
    except ValueError as e:
        await update.message.reply_text(f"{e} Nhập lại ngày:"); return CHOOSING_INCOME_DATE

async def get_income_source(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    source=q.data.replace("income_source_","")
    user_id=q.from_user.id; user_name=q.from_user.full_name
    amount=context.user_data.get('amount'); note=context.user_data.get('note')
    created_at=context.user_data.get('created_at')
    db_add_income(user_id,amount,source,note,created_at)
    await q.edit_message_text(f"✅ Thu nhập: {amount:,.0f} đ - {source} ({note}) ngày {created_at.strftime('%d/%m/%Y')}")
    await post_to_channel(context,
        f"🔔 *Thu nhập mới*\n💰 `{amount:,.0f} đ`\n📌 {source}\n📝 {note}\n📅 {created_at.strftime('%d/%m/%Y')}\n👤 `{anonymize_name(user_name)}`")
    return await end_conv(update,context,"✅ Thu nhập đã lưu.")

async def get_expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        amount=parse_amount(update.message.text)
        if amount<=0: await update.message.reply_text("Số tiền phải > 0:"); return CHOOSING_EXPENSE_AMOUNT
        if amount>LARGE_AMOUNT_THRESHOLD:
            context.user_data['amount_to_confirm']=amount
            kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Đúng vậy",callback_data="confirm_amount_yes")],
                [InlineKeyboardButton("❌ Nhập lại",callback_data="confirm_amount_no")],
                [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
            await update.message.reply_text(f"Chắc chắn chi {amount:,.0f} đ? (số lớn)",reply_markup=kb)
            return CONFIRM_EXPENSE_AMOUNT
        context.user_data['amount']=amount
        await update.message.reply_text("Nhập ghi chú cho khoản chi:"); return CHOOSING_EXPENSE_NOTE
    except Exception: await update.message.reply_text("Không hợp lệ. Thử lại (VD: 5tr2, 100k):"); return CHOOSING_EXPENSE_AMOUNT

async def confirm_expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="confirm_amount_yes":
        context.user_data['amount']=context.user_data.pop('amount_to_confirm')
        await q.edit_message_text("Đã xác nhận. Nhập ghi chú:"); return CHOOSING_EXPENSE_NOTE
    elif q.data=="confirm_amount_no":
        context.user_data.pop('amount_to_confirm',None)
        await q.edit_message_text("Nhập lại số tiền:"); return CHOOSING_EXPENSE_AMOUNT
    return await cancel_conversation_callback(update,context)

async def get_expense_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['note']=update.message.text
    if 'created_at' in context.user_data:
        await update.message.reply_text("Chọn danh mục chi tiêu:",reply_markup=_expense_cat_kb()); return CHOOSING_EXPENSE_CATEGORY
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("Hôm nay",callback_data="date_today_expense")],
        [InlineKeyboardButton("Ngày khác",callback_data="date_other_expense")],
        [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
    await update.message.reply_text("Giao dịch vào ngày nào?",reply_markup=kb); return CHOOSING_EXPENSE_DATE

async def choose_expense_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="date_today_expense":
        context.user_data['created_at']=datetime.datetime.now()
        await q.edit_message_text("Chọn danh mục chi tiêu:",reply_markup=_expense_cat_kb()); return CHOOSING_EXPENSE_CATEGORY
    elif q.data=="date_other_expense":
        await q.edit_message_text("Nhập ngày (VD: 25/06/2024 hoặc 15/1):"); return CHOOSING_EXPENSE_DATE
    return await cancel_conversation_callback(update,context)

async def get_expense_date_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['created_at']=parse_date(update.message.text)
        await update.message.reply_text("Chọn danh mục chi tiêu:",reply_markup=_expense_cat_kb()); return CHOOSING_EXPENSE_CATEGORY
    except ValueError as e:
        await update.message.reply_text(f"{e} Nhập lại ngày:"); return CHOOSING_EXPENSE_DATE

async def get_expense_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    category=q.data.replace("expense_category_","")
    user_id=q.from_user.id; user_name=q.from_user.full_name
    amount=context.user_data.get('amount'); note=context.user_data.get('note')
    created_at=context.user_data.get('created_at')
    db_add_expense(user_id,amount,note,category,created_at)
    await q.edit_message_text(f"✅ Chi tiêu: {amount:,.0f} đ - {note} ({category}) ngày {created_at.strftime('%d/%m/%Y')}")
    budgets=db_get_budgets(user_id); budget_amt=budgets.get(category)
    if budget_amt:
        m_start,m_end=get_month_range(created_at)
        spent=db_sum_expense_by_category_in_period(user_id,category,m_start,m_end)
        ratio=spent/budget_amt if budget_amt>0 else 0
        if ratio>=1.0:
            await q.message.reply_text(f"🚨 *VƯỢT NGÂN SÁCH* {category}: {spent:,.0f}/{budget_amt:,.0f} đ",parse_mode='Markdown')
        elif ratio>=0.9:
            await q.message.reply_text(f"⚠️ *GẦN HẾT NGÂN SÁCH* {category}: {spent:,.0f}/{budget_amt:,.0f} đ (≥90%)",parse_mode='Markdown')
    await post_to_channel(context,
        f"🔔 *Chi tiêu mới*\n💸 `{amount:,.0f} đ`\n📌 {category}\n📝 {note}\n📅 {created_at.strftime('%d/%m/%Y')}\n👤 `{anonymize_name(user_name)}`")
    return await end_conv(update,context,"✅ Chi tiêu đã lưu.")

# ===================== DELETE FLOW =====================
async def handle_delete_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _clear_keep_submenu(context.user_data)
    submenu = context.user_data.get('submenu', 'finance')
    if submenu == 'crypto':
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("Xóa giao dịch Crypto", callback_data="delete_action_individual")],
            [InlineKeyboardButton("Đặt lại toàn bộ Crypto", callback_data="delete_action_reset_crypto")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]])
        prompt = "🗑️ Xóa dữ liệu Crypto?"
    else:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("Xóa giao dịch", callback_data="delete_action_individual")],
            [InlineKeyboardButton("Đặt lại toàn bộ", callback_data="delete_action_reset_all")],
            [InlineKeyboardButton("❌ Hủy bỏ", callback_data="cancel_action")]])
        prompt = "Bạn muốn làm gì?"
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(prompt, reply_markup=kb)
    else:
        await update.message.reply_text(prompt, reply_markup=kb)
    return DELETE_CHOOSING_ACTION

async def choose_delete_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="delete_action_individual":
        user_id=update.effective_user.id
        submenu=context.user_data.get('submenu','finance')
        if submenu=='crypto':
            trades=db_crypto_get_trades(user_id)[:5]
            if not trades: return await end_conv(update,context,"Chưa có giao dịch Crypto để xóa.")
            msg="5 giao dịch Crypto gần nhất:\n\n"; kb=[]
            for i,t in enumerate(trades):
                tid,sym,_cg,side,qty,price,_fee,_note,cat_str=t
                dt=datetime.datetime.fromisoformat(str(cat_str)).strftime('%d/%m/%Y') if cat_str else ''
                msg+=f"*{i+1}.* [{side}] {sym} {qty} @ ${price:,.2f} - {dt}\n"
                kb.append([InlineKeyboardButton(f"Xóa #{i+1}",callback_data=f"delete_id_{tid}_crypto")])
            kb.append([InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")])
            await q.edit_message_text(msg,reply_markup=InlineKeyboardMarkup(kb),parse_mode='Markdown')
        else:
            txs=db_get_last_n_transactions(user_id,5)
            if not txs: return await end_conv(update,context,"Chưa có giao dịch để xóa.")
            msg="5 giao dịch gần nhất:\n\n"; kb=[]
            for i,t in enumerate(txs):
                tid,ttype,amount,note,cat,cat_str=t
                tname="Chi tiêu" if ttype=="expense" else "Thu nhập"
                dt=datetime.datetime.fromisoformat(cat_str).strftime('%d/%m/%Y')
                msg+=f"*{i+1}.* [{tname}] {amount:,.0f} đ - `{note}` ({cat}) - {dt}\n"
                kb.append([InlineKeyboardButton(f"Xóa mục {i+1}",callback_data=f"delete_id_{tid}_{ttype}")])
            kb.append([InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")])
            await q.edit_message_text(msg,reply_markup=InlineKeyboardMarkup(kb),parse_mode='Markdown')
        return DELETE_INDIVIDUAL_LISTING
    elif q.data=="delete_action_reset_all":
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Có, chắc chắn",callback_data="reset_confirm_1_yes")],
            [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
        await q.edit_message_text("Xóa TẤT CẢ dữ liệu tài chính? Không thể hoàn tác.",reply_markup=kb)
        return RESET_DATA_CONFIRM_STEP1
    elif q.data=="delete_action_reset_crypto":
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Có, chắc chắn",callback_data="reset_confirm_1_crypto")],
            [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
        await q.edit_message_text("Xóa TẤT CẢ giao dịch Crypto? Không thể hoàn tác.",reply_markup=kb)
        return RESET_DATA_CONFIRM_STEP1
    return await cancel_conversation_callback(update,context)

async def delete_choose_item(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    parts=q.data.split('_'); trans_id=int(parts[2]); trans_type=parts[3]
    context.user_data['id_to_delete']=trans_id; context.user_data['type_to_delete']=trans_type
    if trans_type=='crypto':
        label=f"Crypto ID {trans_id}"
        confirm_cb="confirm_delete_crypto_yes"
    else:
        label=f"ID {trans_id} ({'Thu nhập' if trans_type=='income' else 'Chi tiêu'})"
        confirm_cb="confirm_delete_yes"
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Xác nhận XÓA",callback_data=confirm_cb)],
        [InlineKeyboardButton("❌ Hủy",callback_data="cancel_action")]])
    await q.edit_message_text(f"Xóa giao dịch {label}?",reply_markup=kb)
    return DELETE_INDIVIDUAL_CONFIRM

async def delete_confirm_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="confirm_delete_yes":
        db_delete_transaction(context.user_data.get('id_to_delete'),context.user_data.get('type_to_delete'))
        return await end_conv(update,context,"✅ Đã xóa giao dịch.")
    elif q.data=="confirm_delete_crypto_yes":
        db_crypto_delete_trade(context.user_data.get('id_to_delete'))
        return await end_conv(update,context,"✅ Đã xóa giao dịch Crypto.")
    return await end_conv(update,context,"Đã hủy.")

async def reset_data_confirmation_1_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="reset_confirm_1_yes":
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("🔥 XÓA TẤT CẢ",callback_data="reset_confirm_2_yes")],
            [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
        await q.edit_message_text("Xác nhận cuối: XÓA toàn bộ dữ liệu? Không thể khôi phục.",reply_markup=kb)
        return RESET_DATA_CONFIRM_STEP2
    elif q.data=="reset_confirm_1_crypto":
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("🔥 XÓA TẤT CẢ CRYPTO",callback_data="reset_confirm_2_crypto")],
            [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
        await q.edit_message_text("Xác nhận cuối: XÓA toàn bộ Crypto? Không thể khôi phục.",reply_markup=kb)
        return RESET_DATA_CONFIRM_STEP2
    return await cancel_conversation_callback(update,context)

async def reset_data_confirmation_2_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="reset_confirm_2_yes":
        db_reset_all_data(q.from_user.id)
        return await end_conv(update,context,"✅ Đã xóa toàn bộ dữ liệu tài chính.")
    elif q.data=="reset_confirm_2_crypto":
        db_crypto_reset_all(q.from_user.id)
        return await end_conv(update,context,"✅ Đã xóa toàn bộ dữ liệu Crypto.")
    return await end_conv(update,context,"Đã hủy.")

# ===================== REPORT & CHART FLOWS =====================
async def report_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _clear_keep_submenu(context.user_data)
    kb = _initial_period_keyboard("report_period")
    msg = "📊 *CHỌN KỲ BÁO CÁO*\nChọn khoảng thời gian:"
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(msg, reply_markup=kb, parse_mode='Markdown')
    else:
        await update.message.reply_text(msg, reply_markup=kb, parse_mode='Markdown')
    return CHOOSING_REPORT_PERIOD

async def generate_report_for_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    if q.data == "cancel_action":
        return await cancel_conversation_callback(update, context)
    period = q.data.replace("report_period_", "")
    # Sub-navigation: back to initial menu
    if period == "back":
        await q.edit_message_text("📊 *CHỌN KỲ BÁO CÁO*\nChọn khoảng thời gian:",
                                  reply_markup=_initial_period_keyboard("report_period"), parse_mode='Markdown')
        return CHOOSING_REPORT_PERIOD
    # Sub-navigation: show month picker
    if period == "pick_month":
        await q.edit_message_text("📆 Chọn tháng:", reply_markup=_month_picker_keyboard("report_period"))
        return CHOOSING_REPORT_PERIOD
    # Sub-navigation: show year picker
    if period == "pick_year":
        await q.edit_message_text("🗓️ Chọn năm:", reply_markup=_year_picker_keyboard("report_period"))
        return CHOOSING_REPORT_PERIOD
    # Generate report
    user_id = q.from_user.id
    start_date, end_date, title_suffix = get_period_dates(period)
    if title_suffix is None:
        return CHOOSING_REPORT_PERIOD
    loading = await q.edit_message_text("⏳ Đang tạo báo cáo...")
    income,expense,balance = db_get_combined_summary(user_id, start_date, end_date)
    incomes_g  = db_list_incomes_grouped(user_id, start_date, end_date)
    expenses_g = db_list_expenses_grouped(user_id, start_date, end_date)
    report = (f"📋 *BÁO CÁO {title_suffix.upper()}*\n{'='*30}\n\n"
              f"💼 *TỔNG QUAN*\n"
              f"┃ 📈 Thu nhập: `{income:,.0f} VND`\n"
              f"┃ 📉 Chi tiêu: `{expense:,.0f} VND`\n"
              f"┃ 💰 Số dư: `{balance:,.0f} VND`\n\n")
    if incomes_g:
        report += "💵 *CHI TIẾT THU NHẬP*\n"
        for i,(src,tot) in enumerate(incomes_g, 1):
            pct = (tot/income*100) if income>0 else 0
            report += f"{i}. **{src}**: `{tot:,.0f}` ({pct:.1f}%)\n"
        report += "\n"
    else: report += "⚠️ *Chưa có thu nhập trong kỳ*\n\n"
    if expenses_g:
        report += "💸 *CHI TIẾT CHI TIÊU*\n"
        for i,(cat,tot) in enumerate(expenses_g, 1):
            pct = (tot/expense*100) if expense>0 else 0
            report += f"{i}. **{cat}**: `{tot:,.0f}` ({pct:.1f}%)\n"
        report += "\n"
    else: report += "✅ *Không có chi tiêu trong kỳ*\n\n"
    budgets = db_get_budgets(user_id)
    if (period in ("month",) or period.startswith("m_")) and budgets and expenses_g:
        report += "⚠️ *NGÂN SÁCH*\n"
        for cat,tot in expenses_g:
            lim = budgets.get(cat)
            if not lim: continue
            ratio = tot/lim if lim>0 else 0
            if ratio>=1.0: report += f"- {cat}: `{tot:,.0f}/{lim:,.0f}` 🚨 VƯỢT\n"
            elif ratio>=0.9: report += f"- {cat}: `{tot:,.0f}/{lim:,.0f}` ⚠️ GẦN\n"
        report += "\n"
    report += ("📊 *ĐÁNH GIÁ*: Tích cực! 🎉" if balance>0 else
               "📊 *ĐÁNH GIÁ*: Cân bằng ⚖️" if balance==0 else
               "📊 *ĐÁNH GIÁ*: Cần cân nhắc ⚠️")
    await loading.delete()
    await q.message.reply_text(report, parse_mode='Markdown')
    return await end_conv(update, context, "✅ Báo cáo hoàn tất.")

async def chart_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _clear_keep_submenu(context.user_data)
    kb = _initial_period_keyboard("chart_period")
    msg = "📊 Xem biểu đồ theo kỳ nào?"
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(msg, reply_markup=kb)
    else:
        await update.message.reply_text(msg, reply_markup=kb)
    return CHOOSING_CHART_PERIOD

async def generate_charts_for_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    if q.data == "cancel_action":
        return await cancel_conversation_callback(update, context)
    period = q.data.replace("chart_period_", "")
    # Sub-navigation: back to initial menu
    if period == "back":
        await q.edit_message_text("📊 Xem biểu đồ theo kỳ nào?",
                                  reply_markup=_initial_period_keyboard("chart_period"))
        return CHOOSING_CHART_PERIOD
    # Sub-navigation: show month picker
    if period == "pick_month":
        await q.edit_message_text("📆 Chọn tháng:", reply_markup=_month_picker_keyboard("chart_period"))
        return CHOOSING_CHART_PERIOD
    # Sub-navigation: show year picker
    if period == "pick_year":
        await q.edit_message_text("🗓️ Chọn năm:", reply_markup=_year_picker_keyboard("chart_period"))
        return CHOOSING_CHART_PERIOD
    # Generate charts
    user_id = q.from_user.id
    _, _, title_suffix = get_period_dates(period)
    if title_suffix is None:
        return CHOOSING_CHART_PERIOD
    msg = await q.edit_message_text(f"⏳ Đang tạo biểu đồ {title_suffix}...")
    charts = await asyncio.to_thread(gen_finance_charts, user_id, period)
    if not charts:
        await msg.delete()
        return await end_conv(update, context, f"Không có dữ liệu trong {title_suffix}.")
    await msg.delete()
    for buf,cap in charts:
        buf.seek(0)
        await q.message.reply_photo(photo=buf, caption=cap)
    return await end_conv(update, context, "✅ Biểu đồ hoàn tất.")

# ===================== EXPORT / IMPORT FINANCE FLOWS =====================
async def export_file_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _clear_keep_submenu(context.user_data)
    user_id=update.effective_user.id
    rows=db_export_data(user_id)
    if not rows:
        return await end_conv(update,context,"⚠️ Chưa có dữ liệu để xuất.")
    kb=InlineKeyboardMarkup([
        [InlineKeyboardButton("Xuất ra CSV",callback_data="export_csv")],
        [InlineKeyboardButton("Xuất ra Excel",callback_data="export_excel")],
        [InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]
    ])
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text("Xuất dữ liệu ra định dạng nào?",reply_markup=kb)
    else:
        await update.message.reply_text("Xuất dữ liệu ra định dạng nào?",reply_markup=kb)
    return EXPORT_CHOOSING_FORMAT

async def handle_export_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q=update.callback_query; await q.answer()
    if q.data=="cancel_action": return await cancel_conversation_callback(update,context)
    user_id=q.from_user.id; rows=db_export_data(user_id)
    if not rows: return await end_conv(update,context,"⚠️ Không có dữ liệu.")
    loading=await q.edit_message_text("⏳ Đang chuẩn bị file...")
    prefix=f"finance_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    headers=['Loại','Số tiền','Ghi chú','Danh mục/Nguồn','Ngày','ID']
    if q.data=="export_csv":
        buf=io.StringIO()
        w=csv.writer(buf); w.writerow(headers)
        for r in rows: w.writerow(r)
        data=buf.getvalue().encode('utf-8-sig')
        await q.message.reply_document(document=data,filename=f'{prefix}.csv',caption="📁 Dữ liệu tài chính (CSV)")
    elif q.data=="export_excel":
        if openpyxl is None:
            await loading.delete()
            return await end_conv(update,context,"❌ Cần cài openpyxl: pip install openpyxl")
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Finance"
        ws.append(headers)
        for r in rows: ws.append(list(r))
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        await q.message.reply_document(document=buf.getvalue(),filename=f'{prefix}.xlsx',caption="📁 Dữ liệu tài chính (Excel)")
    await loading.delete()
    return await end_conv(update,context,"✅ Xuất file hoàn tất.")

async def import_file_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _clear_keep_submenu(context.user_data)
    msg=("📥 *NHẬP DỮ LIỆU TÀI CHÍNH*\n\n"
         "File cần có 5 cột:\n`Loại` | `Số tiền` | `Ghi chú` | `Danh mục/Nguồn` | `Ngày`\n\n"
         "Ví dụ:\n`Chi tiêu | 50000 | Ăn trưa | Ăn uống | 2024-07-28`\n\n"
         "Gửi file .csv hoặc .xlsx ngay bây giờ.")
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Hủy bỏ",callback_data="cancel_action")]])
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(msg,reply_markup=kb,parse_mode='Markdown')
    else:
        await update.message.reply_text(msg,reply_markup=kb,parse_mode='Markdown')
    return IMPORT_FILE_UPLOADED

async def handle_import_file_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id=update.effective_user.id; doc=update.message.document
    fname=(doc.file_name or "").lower()
    if not (fname.endswith('.csv') or fname.endswith('.xlsx')):
        await update.message.reply_text("⚠️ Chỉ hỗ trợ .csv hoặc .xlsx. Thử lại.")
        return IMPORT_FILE_UPLOADED
    await update.message.reply_text("⏳ Đang xử lý file...")
    fobj=await doc.get_file(); bio=io.BytesIO()
    await fobj.download_to_memory(bio); bio.seek(0)
    ok_income=ok_expense=fail=0
    try:
        if fname.endswith('.xlsx'):
            if load_workbook is None:
                return await end_conv(update,context,"❌ Cần cài openpyxl: pip install openpyxl")
            wb=load_workbook(filename=bio,read_only=True,data_only=True)
            ws=wb.active
            hrow=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
            hdrs=[str(h).strip().lower().replace(' ','_') if h else '' for h in hrow]
            col=lambda n: hdrs.index(n) if n in hdrs else -1
            for r in ws.iter_rows(min_row=2,values_only=True):
                if not r or all(v is None or str(v).strip()=='' for v in r): continue
                try:
                    ttype=str(r[col('loại')]).strip().lower() if col('loại')>=0 else ''
                    amount=float(r[col('số_tiền')]) if col('số_tiền')>=0 else 0
                    note=str(r[col('ghi_chú')]).strip() if col('ghi_chú')>=0 else ''
                    cat=str(r[col('danh_mục/nguồn')]).strip() if col('danh_mục/nguồn')>=0 else ''
                    raw_date=r[col('ngày')] if col('ngày')>=0 else None
                    if hasattr(raw_date,'year'): dt=raw_date if hasattr(raw_date,'hour') else datetime.datetime.combine(raw_date,datetime.time())
                    else: dt=datetime.datetime.fromisoformat(str(raw_date).split('T')[0]) if raw_date else datetime.datetime.now()
                    if ttype in ('thu nhập','thu nhap'):
                        db_add_income(user_id,amount,cat,note,dt); ok_income+=1
                    elif ttype in ('chi tiêu','chi tieu'):
                        db_add_expense(user_id,amount,note,cat,dt); ok_expense+=1
                    else: fail+=1
                except Exception: fail+=1
        else:
            reader=csv.DictReader(io.TextIOWrapper(bio,encoding='utf-8'))
            for r in reader:
                try:
                    cols={k.strip().lower().replace(' ','_'):v for k,v in r.items()}
                    ttype=(cols.get('loại') or '').strip().lower()
                    amount=float(cols.get('số_tiền') or cols.get('so_tien') or 0)
                    note=(cols.get('ghi_chú') or cols.get('ghi_chu') or '').strip()
                    cat=(cols.get('danh_mục/nguồn') or cols.get('danh_muc/nguon') or '').strip()
                    raw_date=(cols.get('ngày') or cols.get('ngay') or '').strip()
                    try: dt=datetime.datetime.fromisoformat(raw_date)
                    except Exception: dt=datetime.datetime.now()
                    if ttype in ('thu nhập','thu nhap'):
                        db_add_income(user_id,amount,cat,note,dt); ok_income+=1
                    elif ttype in ('chi tiêu','chi tieu'):
                        db_add_expense(user_id,amount,note,cat,dt); ok_expense+=1
                    else: fail+=1
                except Exception: fail+=1
    except Exception as e:
        logger.error(f"Import error user {user_id}: {e}")
        return await end_conv(update,context,f"❌ Lỗi xử lý file: {e}")
    total=ok_income+ok_expense
    return await end_conv(update,context,
        f"✅ Nhập thành công {total} giao dịch:\n- {ok_income} thu nhập\n- {ok_expense} chi tiêu\n- {fail} lỗi bỏ qua")

# ===================== BUDGET COMMANDS =====================
async def budget_set_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tokens=(update.message.text or "").strip().split()
    if len(tokens)<3:
        await update.message.reply_text("Cú pháp: /ngansach <Danh mục> <Số tiền>\nVD: /ngansach Ăn uống 3tr"); return
    category=" ".join(tokens[1:-1]).strip()
    try:
        amount=parse_amount(tokens[-1])
        if amount<=0: raise ValueError
        db_set_budget(update.effective_user.id,category,amount)
        await update.message.reply_text(f"✅ Đặt ngân sách *{category}*: `{amount:,.0f} đ`/tháng",parse_mode='Markdown')
    except Exception:
        await update.message.reply_text("Số tiền không hợp lệ. Thử: /ngansach Ăn uống 3tr")

async def budget_list_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id=update.effective_user.id; budgets=db_get_budgets(user_id)
    if not budgets: await update.message.reply_text("Chưa đặt ngân sách nào."); return
    now=datetime.datetime.now(); m_start,m_end=get_month_range(now)
    lines=["📋 *NGÂN SÁCH THÁNG NÀY*"]
    for cat,lim in budgets.items():
        spent=db_sum_expense_by_category_in_period(user_id,cat,m_start,m_end)
        ratio=spent/lim if lim>0 else 0
        status="🚨 VƯỢT" if ratio>=1.0 else ("⚠️ GẦN" if ratio>=0.9 else "✅ OK")
        lines.append(f"- {cat}: {spent:,.0f}/{lim:,.0f} đ ({ratio*100:.0f}%)  {status}")
    await update.message.reply_text("\n".join(lines),parse_mode='Markdown')

async def budget_delete_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args=(update.message.text or "").split(maxsplit=1)
    if len(args)<2: await update.message.reply_text("Cú pháp: /xoa_ngansach <Danh mục>"); return
    category=args[1].strip(); db_delete_budget(update.effective_user.id,category)
    await update.message.reply_text(f"🗑️ Đã xóa ngân sách *{category}*",parse_mode='Markdown')

# ===================== UNIFIED TEXT HANDLER (catch-all) =====================
async def unified_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_rate(update): return
    text=(update.message.text or "").strip(); user_id=update.effective_user.id
    # Try crypto natural trade first
    parsed=_parse_natural_trade(text,user_id)
    if parsed:
        side,symbol,qty,price,note=parsed
        await _execute_trade(update,context,side,symbol,qty,price,note); return
    # Try finance NLP
    parsed_f=nlp_parse_transaction(text)
    if parsed_f and parsed_f["amount"]>0:
        if parsed_f["type"]=="income":
            db_add_income(user_id,parsed_f["amount"],parsed_f["category"],parsed_f["note"],parsed_f["created_at"])
            msg=(f"✅ *Thu nhập*: `{parsed_f['amount']:,.0f} đ`\n"
                 f"• Nguồn: *{parsed_f['category']}*\n• Ghi chú: _{parsed_f['note']}_\n• Ngày: {parsed_f['created_at'].strftime('%d/%m/%Y')}")
        else:
            db_add_expense(user_id,parsed_f["amount"],parsed_f["note"],parsed_f["category"],parsed_f["created_at"])
            msg=(f"✅ *Chi tiêu*: `{parsed_f['amount']:,.0f} đ`\n"
                 f"• Danh mục: *{parsed_f['category']}*\n• Ghi chú: _{parsed_f['note']}_\n• Ngày: {parsed_f['created_at'].strftime('%d/%m/%Y')}")
            budgets=db_get_budgets(user_id); budget_amt=budgets.get(parsed_f["category"])
            if budget_amt:
                m_start,m_end=get_month_range(parsed_f["created_at"])
                spent=db_sum_expense_by_category_in_period(user_id,parsed_f["category"],m_start,m_end)
                ratio=spent/budget_amt if budget_amt>0 else 0
                if ratio>=1.0: msg+=f"\n\n🚨 *VƯỢT NGÂN SÁCH* {parsed_f['category']}: {spent:,.0f}/{budget_amt:,.0f} đ"
                elif ratio>=0.9: msg+=f"\n\n⚠️ *GẦN HẾT NGÂN SÁCH* {parsed_f['category']}: {spent:,.0f}/{budget_amt:,.0f} đ"
        await update.message.reply_text(msg,parse_mode='Markdown')
        await update.message.reply_text("Giao dịch đã ghi nhận.",reply_markup=main_menu_keyboard())
        return
    # Fallback
    await update.message.reply_text("Không nhận diện được lệnh. Dùng các nút menu bên dưới:",reply_markup=main_menu_keyboard())

# ===================== ERROR HANDLER =====================
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    try: err="".join(traceback.format_exception(None,context.error,context.error.__traceback__))
    except Exception: err=str(context.error)
    logger.error(f"Update {update} caused error:\n{err}")
    try:
        chat_id=update.effective_chat.id if update and hasattr(update,"effective_chat") else None
        if chat_id: await context.bot.send_message(chat_id,"⚠️ Có lỗi xảy ra. Hãy thử lại.")
    except Exception: pass

# ===================== BUILD APP =====================
def build_app() -> "Application":
    run_migrations()
    app = Application.builder().token(BOT_TOKEN).build()

    # Finance ConversationHandlers (register FIRST for priority)
    add_conv=ConversationHandler(
        entry_points=[CommandHandler("them",handle_add_button),
                      MessageHandler(filters.Regex(f"^{re.escape(BTN_FADD)}$"),handle_add_button)],
        states={
            CHOOSING_ADD_TYPE:[CallbackQueryHandler(choose_add_type_callback,pattern="^add_type_")],
            CHOOSING_SUPPLEMENT_TYPE:[CallbackQueryHandler(choose_supplement_type_callback,pattern="^supplement_type_")],
            CHOOSING_INCOME_AMOUNT:[MessageHandler(filters.TEXT & ~filters.COMMAND,get_income_amount)],
            CONFIRM_INCOME_AMOUNT:[CallbackQueryHandler(confirm_income_amount,pattern="^confirm_amount_")],
            CHOOSING_INCOME_NOTE:[MessageHandler(filters.TEXT & ~filters.COMMAND,get_income_note)],
            CHOOSING_INCOME_DATE:[CallbackQueryHandler(choose_income_date,pattern="^date_"),
                                  MessageHandler(filters.TEXT & ~filters.COMMAND,get_income_date_input)],
            CHOOSING_INCOME_SOURCE:[CallbackQueryHandler(get_income_source,pattern="^income_source_")],
            CHOOSING_EXPENSE_AMOUNT:[MessageHandler(filters.TEXT & ~filters.COMMAND,get_expense_amount)],
            CONFIRM_EXPENSE_AMOUNT:[CallbackQueryHandler(confirm_expense_amount,pattern="^confirm_amount_")],
            CHOOSING_EXPENSE_NOTE:[MessageHandler(filters.TEXT & ~filters.COMMAND,get_expense_note)],
            CHOOSING_EXPENSE_DATE:[CallbackQueryHandler(choose_expense_date,pattern="^date_"),
                                   MessageHandler(filters.TEXT & ~filters.COMMAND,get_expense_date_input)],
            CHOOSING_EXPENSE_CATEGORY:[CallbackQueryHandler(get_expense_category,pattern="^expense_category_")],
        },
        fallbacks=[CommandHandler("cancel",cancel_conversation),
                   CallbackQueryHandler(cancel_conversation_callback,pattern="^cancel_action$")],
    )
    delete_conv=ConversationHandler(
        entry_points=[CommandHandler("xoa",handle_delete_button),
                      MessageHandler(filters.Regex(f"^{re.escape(BTN_FDEL)}$"),handle_delete_button)],
        states={
            DELETE_CHOOSING_ACTION:[CallbackQueryHandler(choose_delete_action,pattern="^delete_action_")],
            DELETE_INDIVIDUAL_LISTING:[CallbackQueryHandler(delete_choose_item,pattern="^delete_id_")],
            DELETE_INDIVIDUAL_CONFIRM:[CallbackQueryHandler(delete_confirm_action,pattern="^confirm_delete_")],
            RESET_DATA_CONFIRM_STEP1:[CallbackQueryHandler(reset_data_confirmation_1_handler,pattern="^reset_confirm_1_")],
            RESET_DATA_CONFIRM_STEP2:[CallbackQueryHandler(reset_data_confirmation_2_handler,pattern="^reset_confirm_2_")],
        },
        fallbacks=[CommandHandler("cancel",cancel_conversation),
                   CallbackQueryHandler(cancel_conversation_callback,pattern="^cancel_action$")],
    )
    report_conv=ConversationHandler(
        entry_points=[CommandHandler("baocao",report_start)],
        states={CHOOSING_REPORT_PERIOD:[CallbackQueryHandler(generate_report_for_period,pattern="^report_period_|^cancel_action$")]},
        fallbacks=[CommandHandler("cancel",cancel_conversation),
                   CallbackQueryHandler(cancel_conversation_callback,pattern="^cancel_action$")],
    )
    chart_conv=ConversationHandler(
        entry_points=[CommandHandler("bieudo",chart_start)],
        states={CHOOSING_CHART_PERIOD:[CallbackQueryHandler(generate_charts_for_period,pattern="^chart_period_|^cancel_action$")]},
        fallbacks=[CommandHandler("cancel",cancel_conversation),
                   CallbackQueryHandler(cancel_conversation_callback,pattern="^cancel_action$")],
    )
    export_conv=ConversationHandler(
        entry_points=[CommandHandler("xuatfile",export_file_start),
                      MessageHandler(filters.Regex(f"^{re.escape(BTN_FEXPORT)}$"),export_file_start)],
        states={EXPORT_CHOOSING_FORMAT:[CallbackQueryHandler(handle_export_choice,pattern="^export_|^cancel_action$")]},
        fallbacks=[CommandHandler("cancel",cancel_conversation),
                   CallbackQueryHandler(cancel_conversation_callback,pattern="^cancel_action$")],
    )
    import_conv=ConversationHandler(
        entry_points=[CommandHandler("nhapfile",import_file_start),
                      MessageHandler(filters.Regex(f"^{re.escape(BTN_FIMPORT)}$"),import_file_start)],
        states={IMPORT_FILE_UPLOADED:[MessageHandler(filters.Document.ALL,handle_import_file_upload)]},
        fallbacks=[CommandHandler("cancel",cancel_conversation),
                   CallbackQueryHandler(cancel_conversation_callback,pattern="^cancel_action$")],
    )

    app.add_handler(add_conv)
    app.add_handler(delete_conv)
    app.add_handler(report_conv)
    app.add_handler(chart_conv)
    app.add_handler(export_conv)
    app.add_handler(import_conv)

    recurring_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_RECURRING)}$"), handle_recurring_btn)],
        states={
            RECURRING_LIST:[
                CallbackQueryHandler(recurring_list_callback,    pattern="^recur_add$|^recur_edit_|^recur_del_|^recur_back_list$|^recur_noop_")],
            RECURRING_ADD_TYPE:[
                CallbackQueryHandler(recurring_pick_type,        pattern="^recur_type_")],
            RECURRING_ADD_AMOUNT:[
                MessageHandler(filters.TEXT & ~filters.COMMAND,  recurring_get_amount)],
            RECURRING_ADD_CATEGORY:[
                CallbackQueryHandler(recurring_get_category,     pattern="^income_source_|^expense_category_")],
            RECURRING_ADD_NOTE:[
                MessageHandler(filters.TEXT & ~filters.COMMAND,  recurring_get_note)],
            RECURRING_ADD_DAY:[
                MessageHandler(filters.TEXT & ~filters.COMMAND,  recurring_get_day)],
            RECURRING_ADD_FREQ:[
                CallbackQueryHandler(recurring_pick_freq,        pattern="^recur_freq_")],
            RECURRING_ADD_MONTH:[
                CallbackQueryHandler(recurring_pick_month,       pattern="^recur_month_")],
            RECURRING_EDIT_FIELD:[
                CallbackQueryHandler(recurring_edit_field,       pattern="^recur_field_|^recur_back_list$")],
            RECURRING_EDIT_VALUE:[
                CallbackQueryHandler(recurring_edit_value,       pattern="^income_source_|^expense_category_|^recur_freq_|^recur_edit_month_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND,  recurring_edit_value)],
            RECURRING_DELETE_CONFIRM:[
                CallbackQueryHandler(recurring_delete_confirm,   pattern="^recur_del_confirm$|^recur_back_list$")],
        },
        fallbacks=[CommandHandler("cancel", cancel_conversation),
                   CallbackQueryHandler(cancel_conversation_callback, pattern="^cancel_action$")],
    )
    app.add_handler(recurring_conv)

    # Commands
    app.add_handler(CommandHandler("start",start))
    app.add_handler(CommandHandler("help",help_cmd))
    app.add_handler(CommandHandler("dashboard",dashboard_cmd))
    app.add_handler(CommandHandler("cp",cp_cmd))
    app.add_handler(CommandHandler("cp_add",cp_add))
    app.add_handler(CommandHandler("cp_sell",cp_sell))
    app.add_handler(CommandHandler("cp_map",cp_map_cmd))
    app.add_handler(CommandHandler("cp_import",cp_import_cmd))
    app.add_handler(CommandHandler("ngansach",budget_set_cmd))
    app.add_handler(CommandHandler("xem_ngansach",budget_list_cmd))
    app.add_handler(CommandHandler("xoa_ngansach",budget_delete_cmd))
    app.add_handler(CommandHandler("cancel",cancel_conversation))

    # Home navigation buttons
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(f"^{re.escape(BTN_HOME_CRYPTO)}$"), handle_home_crypto_btn))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(f"^{re.escape(BTN_HOME_FINANCE)}$"), handle_home_finance_btn))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(f"^{re.escape(BTN_BACK)}$"), handle_back_btn))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(f"^{re.escape(BTN_BUDGET)}$"), handle_budget_btn))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(f"^{re.escape(BTN_RECURRING)}$"), handle_recurring_btn))

    # Shared chart/report buttons (routed by submenu state)
    shared_btn_pattern = f"^({'|'.join(re.escape(b) for b in [BTN_CHART, BTN_REPORT])})$"
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(shared_btn_pattern), handle_shared_btn))

    # Crypto menu buttons
    crypto_btn_pattern = f"^({'|'.join(re.escape(b) for b in [BTN_PORTFOLIO,BTN_BUY,BTN_SELL,BTN_MAP,BTN_CIMPORT,BTN_CEXPORT])})$"
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex(crypto_btn_pattern), handle_crypto_menu_btns))

    # Callbacks
    app.add_handler(CallbackQueryHandler(theme_select_cb,         pattern="^SET_THEME:"))
    app.add_handler(CallbackQueryHandler(handle_budget_menu_callback, pattern="^budget_menu_"))
    # Global fallbacks for chart/report when triggered via handle_shared_btn (outside ConvHandler)
    app.add_handler(CallbackQueryHandler(generate_charts_for_period,  pattern="^chart_period_"))
    app.add_handler(CallbackQueryHandler(generate_report_for_period,  pattern="^report_period_"))

    # Unified catch-all (natural trade + finance NLP)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, unified_text_handler))

    app.add_error_handler(error_handler)
    return app

# ===================== RECURRING TRANSACTIONS =====================

def _recurring_freq_str(month) -> str:
    return "hàng tháng" if month is None else f"hàng năm (tháng {month})"

def _recurring_list_kb(items) -> InlineKeyboardMarkup:
    kb = []
    for r in items:
        rid, ttype, amount, category, note, day, month, _ = r
        freq = "tháng" if month is None else f"năm/T{month}"
        icon = "💰" if ttype == "income" else "💸"
        label = f"{icon} {note or category} {amount:,.0f}đ ngày {day}/{freq}"
        kb.append([InlineKeyboardButton(label[:48], callback_data=f"recur_noop_{rid}")])
        kb.append([InlineKeyboardButton("✏️ Sửa", callback_data=f"recur_edit_{rid}"),
                   InlineKeyboardButton("🗑️ Xóa", callback_data=f"recur_del_{rid}")])
    kb.append([InlineKeyboardButton("➕ Thêm khoản mới", callback_data="recur_add")])
    kb.append([InlineKeyboardButton("❌ Đóng", callback_data="cancel_action")])
    return InlineKeyboardMarkup(kb)

async def handle_recurring_btn(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    _clear_keep_submenu(context.user_data)
    user_id = update.effective_user.id
    items = db_recurring_list(user_id)
    msg = (f"🔁 *Giao dịch định kỳ* ({len(items)} khoản):"
           if items else "🔁 *Giao dịch định kỳ*\nChưa có khoản nào. Nhấn ➕ để thêm.")
    kb = _recurring_list_kb(items)
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.edit_message_text(msg, reply_markup=kb, parse_mode='Markdown')
        except Exception: await update.callback_query.message.reply_text(msg, reply_markup=kb, parse_mode='Markdown')
    else:
        await update.message.reply_text(msg, reply_markup=kb, parse_mode='Markdown')
    return RECURRING_LIST

async def recurring_list_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    uid = q.from_user.id
    if q.data == "recur_add":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💰 Thu nhập", callback_data="recur_type_income"),
             InlineKeyboardButton("💸 Chi tiêu",  callback_data="recur_type_expense")],
            [InlineKeyboardButton("❌ Hủy", callback_data="cancel_action")]])
        await q.edit_message_text("🔁 *THÊM KHOẢN ĐỊNH KỲ*\nLoại giao dịch?",
                                  reply_markup=kb, parse_mode='Markdown')
        return RECURRING_ADD_TYPE
    elif q.data.startswith("recur_edit_"):
        rec_id = int(q.data.split("_")[-1])
        row = db_recurring_get(rec_id)
        if not row: await q.edit_message_text("Không tìm thấy khoản này."); return ConversationHandler.END
        context.user_data.update({'recur_edit_id': rec_id, 'recur_edit_row': row})
        _, _, ttype, amount, category, note, day, month, _ = row
        info = (f"📋 *Khoản định kỳ hiện tại:*\n"
                f"Loại: {'💰 Thu nhập' if ttype=='income' else '💸 Chi tiêu'}\n"
                f"Số tiền: `{amount:,.0f} đ`\n"
                f"Danh mục: {category}\n"
                f"Ghi chú: {note or '—'}\n"
                f"Lịch: ngày {day} {_recurring_freq_str(month)}\n\nSửa trường nào?")
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💰 Số tiền",       callback_data="recur_field_amount"),
             InlineKeyboardButton("📌 Danh mục",      callback_data="recur_field_category")],
            [InlineKeyboardButton("📝 Ghi chú",       callback_data="recur_field_note"),
             InlineKeyboardButton("📅 Ngày/Tần suất", callback_data="recur_field_schedule")],
            [InlineKeyboardButton("❌ Hủy", callback_data="recur_back_list")]])
        await q.edit_message_text(info, reply_markup=kb, parse_mode='Markdown')
        return RECURRING_EDIT_FIELD
    elif q.data.startswith("recur_del_"):
        rec_id = int(q.data.split("_")[-1])
        row = db_recurring_get(rec_id)
        if not row: await q.edit_message_text("Không tìm thấy."); return ConversationHandler.END
        context.user_data['recur_del_id'] = rec_id
        _, _, ttype, amount, category, note, day, month, _ = row
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Xóa", callback_data="recur_del_confirm"),
             InlineKeyboardButton("◀️ Quay lại", callback_data="recur_back_list")]])
        await q.edit_message_text(
            f"Xóa khoản định kỳ:\n*{note or category}* — `{amount:,.0f} đ`\nNgày {day} {_recurring_freq_str(month)}?",
            reply_markup=kb, parse_mode='Markdown')
        return RECURRING_DELETE_CONFIRM
    elif q.data == "recur_back_list":
        items = db_recurring_list(uid)
        msg = (f"🔁 *Giao dịch định kỳ* ({len(items)} khoản):"
               if items else "🔁 *Giao dịch định kỳ*\nChưa có khoản nào.")
        await q.edit_message_text(msg, reply_markup=_recurring_list_kb(items), parse_mode='Markdown')
        return RECURRING_LIST
    return RECURRING_LIST

# ── Add flow ─────────────────────────────────────────────

async def recurring_pick_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    context.user_data['recur_ttype'] = q.data.replace("recur_type_", "")
    await q.edit_message_text("💰 Nhập số tiền (VD: 250k, 1.5tr, 200000):")
    return RECURRING_ADD_AMOUNT

async def recurring_get_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    amount = parse_amount_loose(update.message.text)
    if not amount or amount <= 0:
        await update.message.reply_text("⚠️ Số tiền không hợp lệ. Nhập lại (VD: 250k):"); return RECURRING_ADD_AMOUNT
    context.user_data['recur_amount'] = amount
    if context.user_data.get('recur_ttype') == 'income':
        await update.message.reply_text("📌 Chọn nguồn thu nhập:", reply_markup=_income_source_kb())
    else:
        await update.message.reply_text("📌 Chọn danh mục chi tiêu:", reply_markup=_expense_cat_kb())
    return RECURRING_ADD_CATEGORY

async def recurring_get_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    cat = q.data.replace("income_source_", "").replace("expense_category_", "")
    context.user_data['recur_category'] = cat
    await q.edit_message_text("📝 Nhập ghi chú (VD: Hóa đơn VNPT), hoặc gõ `-` để bỏ qua:")
    return RECURRING_ADD_NOTE

async def recurring_get_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    txt = update.message.text.strip()
    context.user_data['recur_note'] = '' if txt == '-' else txt
    await update.message.reply_text("📅 Ngày mấy trong tháng sẽ tự ghi? Nhập số (1–28):")
    return RECURRING_ADD_DAY

async def recurring_get_day(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        day = int(update.message.text.strip())
        if not 1 <= day <= 28: raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ Nhập số từ 1 đến 28:"); return RECURRING_ADD_DAY
    context.user_data['recur_day'] = day
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Hàng tháng", callback_data="recur_freq_monthly"),
         InlineKeyboardButton("📆 Hàng năm",   callback_data="recur_freq_yearly")]])
    await update.message.reply_text(f"Lặp lại như thế nào?", reply_markup=kb)
    return RECURRING_ADD_FREQ

async def recurring_pick_freq(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    if q.data == "recur_freq_monthly":
        context.user_data['recur_month'] = None
        return await _recurring_save_new(q, context)
    months = ["T1","T2","T3","T4","T5","T6","T7","T8","T9","T10","T11","T12"]
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(m, callback_data=f"recur_month_{i+1}")
                                for i, m in enumerate(row, start=ri*4)]
                               for ri, row in enumerate([months[i:i+4] for i in range(0,12,4)])])
    await q.edit_message_text("Hàng năm vào tháng nào?", reply_markup=kb)
    return RECURRING_ADD_MONTH

async def recurring_pick_month(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    context.user_data['recur_month'] = int(q.data.replace("recur_month_", ""))
    return await _recurring_save_new(q, context)

async def _recurring_save_new(q, context) -> int:
    uid = q.from_user.id; d = context.user_data
    db_recurring_add(uid, d['recur_ttype'], d['recur_amount'], d['recur_category'],
                     d.get('recur_note', ''), d['recur_day'], d.get('recur_month'))
    freq = _recurring_freq_str(d.get('recur_month'))
    icon = "💰" if d['recur_ttype'] == 'income' else "💸"
    saved = (f"✅ *Đã thêm khoản định kỳ*\n"
             f"{icon} `{d['recur_amount']:,.0f} đ` — {d['recur_category']}\n"
             f"📝 {d.get('recur_note') or '—'}\n"
             f"📅 Ngày {d['recur_day']} {freq}")
    items = db_recurring_list(uid)
    try:
        await q.edit_message_text(saved + f"\n\n🔁 *Danh sách ({len(items)} khoản):*",
                                  reply_markup=_recurring_list_kb(items), parse_mode='Markdown')
    except Exception:
        await q.message.reply_text(saved, parse_mode='Markdown')
    return RECURRING_LIST

# ── Delete flow ───────────────────────────────────────────

async def recurring_delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer(); uid = q.from_user.id
    if q.data == "recur_del_confirm":
        db_recurring_delete(context.user_data.get('recur_del_id', 0))
    items = db_recurring_list(uid)
    msg = (f"✅ Đã xóa.\n\n🔁 *Danh sách ({len(items)} khoản):*"
           if q.data == "recur_del_confirm" else
           f"🔁 *Giao dịch định kỳ* ({len(items)} khoản):" if items else
           "🔁 *Giao dịch định kỳ*\nChưa có khoản nào.")
    await q.edit_message_text(msg, reply_markup=_recurring_list_kb(items), parse_mode='Markdown')
    return RECURRING_LIST

# ── Edit flow ─────────────────────────────────────────────

async def recurring_edit_field(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    if q.data == "recur_back_list":
        uid = q.from_user.id; items = db_recurring_list(uid)
        msg = f"🔁 *Giao dịch định kỳ* ({len(items)} khoản):" if items else "🔁 Chưa có khoản nào."
        await q.edit_message_text(msg, reply_markup=_recurring_list_kb(items), parse_mode='Markdown')
        return RECURRING_LIST
    field = q.data.replace("recur_field_", "")
    context.user_data['recur_edit_field'] = field
    _, _, ttype, amount, category, note, day, month, _ = context.user_data['recur_edit_row']
    if field == 'amount':
        await q.edit_message_text(f"Số tiền hiện tại: `{amount:,.0f} đ`\nNhập số tiền mới:", parse_mode='Markdown')
    elif field == 'category':
        kb = _income_source_kb() if ttype == 'income' else _expense_cat_kb()
        await q.edit_message_text(f"Danh mục hiện tại: *{category}*\nChọn danh mục mới:", reply_markup=kb, parse_mode='Markdown')
    elif field == 'note':
        await q.edit_message_text(f"Ghi chú hiện tại: *{note or '—'}*\nNhập ghi chú mới (hoặc `-` để xóa):", parse_mode='Markdown')
    elif field == 'schedule':
        await q.edit_message_text(f"Lịch hiện tại: ngày *{day}* {_recurring_freq_str(month)}\nNhập ngày mới (1–28):", parse_mode='Markdown')
    return RECURRING_EDIT_VALUE

async def recurring_edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    field = context.user_data.get('recur_edit_field')
    rec_id = context.user_data.get('recur_edit_id')
    _, _, ttype, amount, category, note, day, month, _ = context.user_data['recur_edit_row']
    uid = update.effective_user.id

    if update.callback_query:
        q = update.callback_query; await q.answer()
        if field == 'category':
            new_cat = q.data.replace("income_source_", "").replace("expense_category_", "")
            db_recurring_update(rec_id, amount, new_cat, note, day, month)
            info = f"✅ Đã cập nhật danh mục → *{new_cat}*"
        elif field == 'schedule' and q.data.startswith("recur_freq_"):
            if q.data == "recur_freq_yearly":
                new_day = context.user_data.get('recur_edit_new_day', day)
                context.user_data['recur_edit_day_tmp'] = new_day
                months = ["T1","T2","T3","T4","T5","T6","T7","T8","T9","T10","T11","T12"]
                kb = InlineKeyboardMarkup([[InlineKeyboardButton(m, callback_data=f"recur_edit_month_{i+1}")
                                           for i, m in enumerate(row, start=ri*4)]
                                          for ri, row in enumerate([months[i:i+4] for i in range(0,12,4)])])
                await q.edit_message_text("Vào tháng nào?", reply_markup=kb); return RECURRING_EDIT_VALUE
            else:
                new_day = context.user_data.get('recur_edit_new_day', day)
                db_recurring_update(rec_id, amount, category, note, new_day, None)
                info = f"✅ Đã cập nhật lịch → ngày {new_day} hàng tháng"
        elif field == 'schedule' and q.data.startswith("recur_edit_month_"):
            new_month = int(q.data.replace("recur_edit_month_", ""))
            new_day = context.user_data.get('recur_edit_day_tmp', day)
            db_recurring_update(rec_id, amount, category, note, new_day, new_month)
            info = f"✅ Đã cập nhật lịch → ngày {new_day} tháng {new_month} hàng năm"
        else:
            info = ""
        items = db_recurring_list(uid)
        await q.edit_message_text(
            (info + f"\n\n🔁 *Danh sách ({len(items)} khoản):*") if info else f"🔁 *Danh sách ({len(items)} khoản):*",
            reply_markup=_recurring_list_kb(items), parse_mode='Markdown')
        return RECURRING_LIST

    # Text input
    text = update.message.text.strip()
    if field == 'amount':
        new_amount = parse_amount_loose(text)
        if not new_amount or new_amount <= 0:
            await update.message.reply_text("⚠️ Số tiền không hợp lệ. Nhập lại:"); return RECURRING_EDIT_VALUE
        db_recurring_update(rec_id, new_amount, category, note, day, month)
        await update.message.reply_text(f"✅ Đã cập nhật số tiền → `{new_amount:,.0f} đ`", parse_mode='Markdown')
    elif field == 'note':
        new_note = '' if text == '-' else text
        db_recurring_update(rec_id, amount, category, new_note, day, month)
        await update.message.reply_text(f"✅ Đã cập nhật ghi chú → *{new_note or '—'}*", parse_mode='Markdown')
    elif field == 'schedule':
        try:
            new_day = int(text)
            if not 1 <= new_day <= 28: raise ValueError
        except ValueError:
            await update.message.reply_text("⚠️ Nhập số từ 1 đến 28:"); return RECURRING_EDIT_VALUE
        context.user_data['recur_edit_new_day'] = new_day
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 Hàng tháng", callback_data="recur_freq_monthly"),
             InlineKeyboardButton("📆 Hàng năm",   callback_data="recur_freq_yearly")]])
        await update.message.reply_text(f"Ngày {new_day} — tần suất?", reply_markup=kb)
        return RECURRING_EDIT_VALUE
    items = db_recurring_list(uid)
    await update.message.reply_text(
        f"🔁 *Danh sách ({len(items)} khoản):*",
        reply_markup=_recurring_list_kb(items), parse_mode='Markdown')
    return RECURRING_LIST

# ── Daily job ─────────────────────────────────────────────

async def check_recurring_job(context: ContextTypes.DEFAULT_TYPE):
    today = datetime.datetime.now()
    rows = db_recurring_all()
    for row in rows:
        rec_id, user_id, ttype, amount, category, note, day, month, last_triggered = row
        if today.day != day: continue
        period_key = today.strftime('%Y-%m') if month is None else today.strftime('%Y')
        if month is not None and today.month != month: continue
        if last_triggered == period_key: continue
        created_at = today
        if ttype == 'income':
            db_add_income(user_id, amount, category, note or '', created_at)
        else:
            db_add_expense(user_id, amount, note or '', category, created_at)
        db_recurring_update_triggered(rec_id, period_key)
        freq = _recurring_freq_str(month)
        icon = "💰" if ttype == 'income' else "💸"
        msg = (f"🔁 *Giao dịch định kỳ tự động*\n"
               f"{icon} `{amount:,.0f} đ` — {category}\n"
               f"📝 {note or '—'}\n"
               f"📅 {freq.capitalize()} ngày {day}")
        try:
            await context.bot.send_message(chat_id=user_id, text=msg, parse_mode='Markdown')
        except Exception as e:
            logger.warning(f"Recurring notify failed user {user_id}: {e}")


# ===================== MAIN =====================
if __name__ == "__main__":
    # Windows: aiohttp + asyncio cần SelectorEventLoop (không dùng ProactorEventLoop)
    import sys
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    if WEBHOOK_URL:
        # Webhook mode: single aiohttp server handles Telegram + dashboard
        app = build_app()
        logger.info(f"Webhook mode — port {WEBHOOK_PORT}")
        asyncio.run(run_with_webhook(app))
    else:
        # Polling mode: separate thread for dashboard, PTB polls Telegram
        start_html_server()
        app = build_app()
        _setup_job_queue(app)
        logger.info(f"Polling mode — dashboard port {HTML_PORT}")
        app.run_polling(allowed_updates=Update.ALL_TYPES)
